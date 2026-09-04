#!/usr/bin/env python3
"""Export creative performance table to xlsx with formatting."""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Load data
with open(r"D:\claude-projects\projects\幻宠\creative_table_data.json", "r", encoding="utf-8") as f:
    data = json.load(f)

rows = data["rows"]

wb = Workbook()
ws = wb.active
ws.title = "素材表现汇总"

# ============================================================
# Styles
# ============================================================
header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2972F4", end_color="2972F4", fill_type="solid")
sub_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
april_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
june_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # light green
summary_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
aeo_font = Font(name="Microsoft YaHei", size=10, color="C00000", bold=True)
install_font = Font(name="Microsoft YaHei", size=10, color="1F4E79")
both_font = Font(name="Microsoft YaHei", size=10, color="7030A0")
normal_font = Font(name="Microsoft YaHei", size=10)
summary_font = Font(name="Microsoft YaHei", size=10, bold=True, color="333333")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

# ============================================================
# Headers
# ============================================================
headers = ["月份", "素材名称", "素材类型", "出价方式", "安装数", "CPI", "R1", "R3", "R7", "备注"]
col_widths = [8, 30, 10, 14, 12, 12, 10, 10, 10, 55]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 28

# ============================================================
# Data rows
# ============================================================
current_month = None
row_num = 2

for r in rows:
    month = r["月份"]

    # Month separator
    if month != current_month:
        current_month = month
        # Add a month separator row
        sep_cell = ws.cell(row=row_num, column=1, value=f"◆ {month}")
        sep_cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="1F4E79")
        sep_fill = april_fill if "4月" in month else june_fill
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=c).fill = sep_fill
            ws.cell(row=row_num, column=c).border = thin_border
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        ws.row_dimensions[row_num].height = 24
        row_num += 1

    # Determine if summary row
    is_summary = r["素材名称"].startswith("【")
    is_april = month == "4月"
    bidding = r["出价方式"]

    # Choose font based on bidding
    if is_summary:
        cell_font = summary_font
        cell_fill = summary_fill
    elif bidding == "AEO":
        cell_font = aeo_font
        cell_fill = april_fill if is_april else june_fill
    elif bidding == "Install":
        cell_font = install_font
        cell_fill = april_fill if is_april else june_fill
    elif bidding == "Install+AEO":
        cell_font = both_font
        cell_fill = june_fill
    else:
        cell_font = normal_font
        cell_fill = april_fill if is_april else june_fill

    # Write cells
    values = [
        r["月份"], r["素材名称"], r["素材类型"], r["出价方式"],
        r["安装数"], r["CPI"], r["R1"], r["R3"], r["R7"], r["备注"]
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = cell_font
        cell.fill = cell_fill
        cell.border = thin_border
        cell.alignment = center_align if col_idx in (1, 3, 4, 5, 6, 7, 8, 9) else left_align

    # Bold for summary rows
    if is_summary:
        for c in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=c).font = summary_font

    ws.row_dimensions[row_num].height = 22
    row_num += 1

# ============================================================
# Legend sheet
# ============================================================
ws2 = wb.create_sheet("说明")
legend_data = [
    ["幻宠帝国 — 素材表现汇总表"],
    [""],
    ["数据范围", "4月留存测试 (4/2-4/4) + 6月测试 (6/11-6/15)，仅美国"],
    ["行定义", "每素材 × 出价方式 = 一行，同时跑AEO+Install的分两行"],
    [""],
    ["颜色说明", ""],
    ["黄色背景", "4月数据"],
    ["绿色背景", "6月数据"],
    ["红色字体", "AEO出价"],
    ["蓝色字体", "Install出价"],
    ["紫色字体", "Install+AEO合并行（数据源未拆分）"],
    [""],
    ["数据源", ""],
    ["4月AEO明细", "飞书Wiki: 幻想宠物留存测试报告 → 内嵌表格 Z0l2s2bs1he1TotM2ERcgBNHnEf → Sheet 8HCs2U"],
    ["4月Install", "仅汇总数据（541安装，CPI $3.91，R1=28.9%），无每素材明细"],
    ["6月明细", "飞书表格: 6月11日幻宠测试(Bnkpsh602hZL49t39cTcT1f4nBq) → 详细的(FfBRvO)"],
    ["6月出价方式", "来自整理表(z89zqR)的广告组名称：US-AAA-install / US-核心用户-AEO"],
    [""],
    ["已知限制", ""],
    ["1. 4月Install", "源数据只有汇总层级，每素材具体安装数/CPI/留存不可得"],
    ["2. 6月Install+AEO", "V-抓宠战斗、P-海啸、V-场景展示、V-核心玩法、V-天灾重建-长版同时跑两种出价，详细的表数据合并，无法按出价方式拆分R1"],
    ["3. R3/R7", "两个测试的源数据均无每素材R3/R7数据，仅6月有广告组级别的R3"],
    ["4. R7", "所有数据源均无R7"],
    [""],
    ["6月广告组级留存（参考）", ""],
    ["US-AAA-install (旧商店页)", "DNU=250, R1=23.20%, R2=8.40%, R3=8.40%"],
    ["素材方向测试-0611 (新商店页)", "DNU=964, R1=26.35%, R2=15.04%, R3=9.34%"],
    ["AEO", "DNU=273, R1=29.67%, R2=16.48%, R3=12.09%"],
]

for row_idx, row_data in enumerate(legend_data, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = Font(name="Microsoft YaHei", size=14, bold=True)
        elif val and (val.startswith("颜色") or val.startswith("数据源") or val.startswith("已知限制")):
            cell.font = Font(name="Microsoft YaHei", size=11, bold=True)
        else:
            cell.font = Font(name="Microsoft YaHei", size=10)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 80

# ============================================================
# Save
# ============================================================
output_path = r"D:\claude-projects\projects\幻宠\幻宠素材表现汇总_4月6月_US.xlsx"
wb.save(output_path)
print(f"Saved to: {output_path}")
