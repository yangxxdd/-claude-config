import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load source data
df = pd.read_excel('creative_final.xlsx', sheet_name='素材表现汇总')
df = df[df['月份'].isin(['4月','6月'])].copy()

# =====================================================
# TAG MAPPING (from creative review)
# =====================================================
tag_map = {
    'P-海啸': ['天灾/生存'],
    'P-幻想宠物3': ['宠物展示', '收集'],
    'P-灾后重建': ['模拟经营', '建造'],
    'V-场景展示': ['模拟经营', '建造'],
    'V-高效抓宠': ['抓宠', '战斗', '宠物展示'],
    'V-核心玩法': ['战斗', '抓宠', '探索/冒险'],
    'V-帕基世界冒险': ['探索/冒险', '模拟经营'],
    'V-收集升级': ['模拟经营', '建造', '收集'],
    'V-天灾重建-长版': ['模拟经营', '建造', '收集'],
    'V-抓宠战斗': ['抓宠', '战斗'],
    'P-宠物展示-二阶进化': ['宠物进化', '宠物展示'],
    'P-宠物展示-合成 3D': ['宠物融合/合成', '宠物进化', '宠物展示'],
    'P-宠物展示-巨物': ['宠物展示'],
    'P-宠物展示-帕基战斗': ['战斗'],
    'P-模拟经营-冬日写实': ['模拟经营', '建造'],
    'P-模拟经营-建造成长': ['模拟经营', '建造', '收集'],
    'P-抓宠经营-超梦': ['抓宠'],
    'P-抓宠经营-虐待': ['抓宠'],
    'P-抓宠经营-血腥': ['经营'],
    'P-抓宠经营-幽飘': ['抓宠', '探索/冒险'],
    'V-宠物展示-宠物合成': ['宠物展示', '宠物融合/合成'],
    'V-宠物展示-二阶合成': ['宠物展示', '宠物进化'],
    'V-模拟经营-砍树': ['模拟经营', '建造', '收集', '战斗'],
    'V-模拟经营-帕基玩法寒霜': ['模拟经营', '探索/冒险', '收集'],
    'V-模拟经营-七日建造': ['模拟经营', '建造'],
    'V-模拟经营-温馨': ['模拟经营'],
    'V-帕萌战斗-出狱打鸡': ['战斗'],
    'V-帕萌战斗-合成狙击': ['战斗', '抓宠'],
    'V-帕萌战斗-群殴打鸡': ['战斗'],
    'V-帕萌战斗-杀宠复刻': ['战斗'],
    'V-帕萌战斗-雪地竞品': ['战斗'],
    'V-抓宠经营-捕捞竞品': ['抓宠', '模拟经营'],
    'V-抓宠经营-狐狸': ['战斗'],
    'V-抓宠经营-虐待': ['战斗', '抓宠'],
    'V-抓宠经营-售卖帕基': ['模拟经营', '建造'],
    'V-抓宠经营-拯救可达鸭': ['战斗', '建造'],
}

# Map tags to each row
def get_tags(name):
    for key in tag_map:
        if key in str(name):
            return tag_map[key]
    return ['未分类']

# Also map broader direction categories
# Core directions for 幻宠: 抓宠, 宠物融合/合成, 宠物进化
direction_categories = {
    '抓宠': ['抓宠'],
    '宠物融合/合成': ['宠物融合/合成'],
    '宠物进化': ['宠物进化'],
    '宠物展示': ['宠物展示'],
    '战斗': ['战斗'],
    '模拟经营/建造': ['模拟经营', '建造'],
    '探索/冒险': ['探索/冒险', '收集'],
    '天灾/生存': ['天灾/生存'],
}

# =====================================================
# Build analysis
# =====================================================
records = []
for _, row in df.iterrows():
    name = str(row['素材名称'])
    if name == 'nan' or '合计' in name or '◆' in name:
        continue
    tags = get_tags(name)
    month = row['月份']
    bid = row['出价方式']
    try:
        dnu = float(row['DNU'])
    except (ValueError, TypeError):
        continue
    cpi = float(row['CPI']) if not pd.isna(row['CPI']) else 0
    r1 = float(row['R1']) if not pd.isna(row['R1']) else 0
    r2 = float(row['R2']) if not pd.isna(row['R2']) else 0
    r3 = float(row['R3']) if not pd.isna(row['R3']) else 0
    cost22 = float(row['次留成本']) if not pd.isna(row['次留成本']) else 0
    spend = float(row['花费(USD)']) if not pd.isna(row['花费(USD)']) else 0
    cvr = float(row['CVR']) if not pd.isna(row['CVR']) else 0
    ctr = float(row['CTR']) if not pd.isna(row['CTR']) else 0
    installs = float(row['安装数']) if not pd.isna(row['安装数']) else 0

    if dnu < 1:
        continue

    key = f'{name}#{bid}'
    records.append({
        'key': key, 'name': name, 'tags': tags, 'tags_str': ','.join(tags),
        'month': month, 'bid': bid, 'dnu': dnu, 'cpi': cpi, 'r1': r1, 'r2': r2, 'r3': r3,
        'cost22': cost22, 'spend': spend, 'cvr': cvr, 'ctr': ctr, 'installs': installs
    })

df_all = pd.DataFrame(records)

# =====================================================
# Group by direction and find best within each
# =====================================================
# Define primary directions (matching what user identified as core)
directions = {
    '捉宠（核心方向）': ['抓宠'],
    '宠物融合/合成（核心方向）': ['宠物融合/合成'],
    '宠物进化（核心方向）': ['宠物进化'],
    '战斗（帕萌战斗/抓宠战斗）': ['战斗'],
    '模拟经营/建造': ['模拟经营', '建造'],
    '宠物展示': ['宠物展示'],
    '探索/冒险/收集': ['探索/冒险', '收集'],
    '天灾/生存': ['天灾/生存'],
}

# Excel output
wb = load_workbook('creative_final.xlsx')
for s in ['9月推荐-按方向', '9月推荐-方向汇总']:
    if s in wb.sheetnames:
        del wb[s]

ws = wb.create_sheet('9月推荐-按方向', 0)

# Styles
hdr_f = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='2F5496')
sec_f = Font(name='Arial', bold=True, size=13, color='1F4E79')
sec_fill = PatternFill('solid', fgColor='D6E4F0')
core_fill = PatternFill('solid', fgColor='C6EFCE')
good_fill = PatternFill('solid', fgColor='E2EFDA')
ok_fill = PatternFill('solid', fgColor='FFF2CC')
bad_fill = PatternFill('solid', fgColor='F4B4C2')
dup_fill = PatternFill('solid', fgColor='D9D9D9')
nf = Font(name='Arial', size=10)
bf = Font(name='Arial', size=10, bold=True)
thin_b = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
la = Alignment(horizontal='left', vertical='center', wrap_text=True)

row = 1
ws.merge_cells('A1:N1')
ws['A1'] = '幻宠 9月测试素材推荐 — 按方向归类（June 6月为主，April 4月为辅）'
ws['A1'].font = Font(name='Arial', bold=True, size=15, color='1F4E79')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 35

row = 2
ws.merge_cells('A2:N2')
ws['A2'] = '筛选逻辑: 方向内按 6月数据排序(CPI低+R1高+R3高) → 4月参照 → 方向内择优推荐。核心方向(捉宠/融合/进化)放宽准入，非核心方向更严。绿色=强烈推荐 黄色=可选 红色=不推荐 灰色=重复素材'
ws['A2'].font = Font(name='Arial', size=9, color='555555')
ws['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[2].height = 28

col_w = [5, 8, 24, 10, 10, 10, 12, 14, 14, 14, 14, 14, 14, 50]
for i, w in enumerate(col_w, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

row = 4
current_row = row

for dir_name, dir_tags in directions.items():
    # Collect all creatives in this direction
    dir_data = []
    for _, rec in df_all.iterrows():
        if any(t in rec['tags'] for t in dir_tags):
            dir_data.append(rec)

    if not dir_data:
        continue

    df_dir = pd.DataFrame(dir_data)

    # Section header
    ws.merge_cells(f'A{current_row}:N{current_row}')
    is_core = '核心方向' in dir_name
    ws[f'A{current_row}'] = f'▌{dir_name}（{len(dir_data)}个素材-出价组合）'
    ws[f'A{current_row}'].font = sec_f
    ws[f'A{current_row}'].fill = sec_fill
    for c in range(1, 15):
        ws.cell(row=current_row, column=c).fill = sec_fill
    current_row += 1

    # Column headers
    hdrs = ['序号', '优先级', '素材名称', '出价', '月份', 'DNU', 'CPI(USD)', 'R1', 'R2', 'R3', '次留成本', 'CVR(参考)', '花费', '推荐理由']
    for i, h in enumerate(hdrs, 1):
        c = ws.cell(row=current_row, column=i, value=h)
        c.font = hdr_f; c.fill = hdr_fill; c.alignment = ca; c.border = thin_b
    ws.row_dimensions[current_row].height = 32
    current_row += 1

    # Sort: June first (by CPI asc), then April
    df_june = df_dir[df_dir['month'] == '6月'].sort_values('cpi')
    df_april = df_dir[(df_dir['month'] == '4月') & (~df_dir['name'].isin(df_june['name'].values))].sort_values('cpi')
    df_sorted = pd.concat([df_june, df_april])

    # Mark duplicates (same name appears in both months -> already in df_june)
    dup_names = set()
    for n in df_dir['name']:
        if len(df_dir[df_dir['name'] == n]) > 1:
            dup_names.add(n)

    idx = 1
    for _, rec in df_sorted.iterrows():
        # Determine recommendation level
        cpi_val = rec['cpi']
        r1_val = rec['r1']
        r3_val = rec['r3']
        dnu_val = rec['dnu']

        is_june = rec['month'] == '6月'
        is_duplicate = rec['name'] in dup_names
        is_small = dnu_val < 10
        is_low_sample = dnu_val < 30

        # Scoring
        reasons = []
        if cpi_val < 5:
            reasons.append(f'CPI极低(${cpi_val:.2f})')
        elif cpi_val < 7:
            reasons.append(f'CPI较低(${cpi_val:.2f})')
        elif cpi_val < 9:
            reasons.append(f'CPI适中(${cpi_val:.2f})')
        else:
            reasons.append(f'CPI偏高(${cpi_val:.2f})')

        if r1_val >= 0.35:
            reasons.append(f'R1={r1_val:.0%}达标(≥35%)')
        elif r1_val >= 0.25:
            reasons.append(f'R1={r1_val:.0%}尚可')
        else:
            reasons.append(f'R1={r1_val:.0%}偏低')

        if r3_val >= 0.10:
            reasons.append(f'R3={r3_val:.0%}达标(≥10%)')
        elif r3_val >= 0.07:
            reasons.append(f'R3={r3_val:.0%}一般')
        else:
            reasons.append(f'R3={r3_val:.0%}偏低')

        if not is_june:
            reasons.append('⚠仅4月数据,需6月环境重验')
        if is_small:
            reasons.append('⚠DNU<10样本太小')
        elif is_low_sample:
            reasons.append(f'⚠DNU={int(dnu_val)}需扩大验证')
        if is_duplicate and is_june:
            reasons.append('📋4月/6月内容相同')

        # Determine tier
        score = 0
        if cpi_val < 5: score += 3
        elif cpi_val < 7: score += 2
        elif cpi_val < 9: score += 1

        if r1_val >= 0.35: score += 3
        elif r1_val >= 0.28: score += 2
        elif r1_val >= 0.20: score += 1

        if r3_val >= 0.15: score += 3
        elif r3_val >= 0.10: score += 2
        elif r3_val >= 0.07: score += 1

        if dnu_val >= 30: score += 1

        if is_small: score -= 1
        if not is_june and is_core: score -= 1  # penalty for April-only in core direction

        if score >= 7: tier = '🥇 强烈推荐'
        elif score >= 5: tier = '🥈 可选'
        elif score >= 3: tier = '🥉 待观察'
        else: tier = '❌ 不推荐'

        if not is_core and tier == '🥇 强烈推荐':
            tier = '🥈 可选'  # Non-core can't be tier 1

        reason_str = '; '.join(reasons)

        ws.cell(row=current_row, column=1, value=idx).font = nf
        ws.cell(row=current_row, column=1).alignment = ca
        ws.cell(row=current_row, column=1).border = thin_b

        ws.cell(row=current_row, column=2, value=tier).font = nf
        ws.cell(row=current_row, column=2).alignment = ca
        ws.cell(row=current_row, column=2).border = thin_b

        ws.cell(row=current_row, column=3, value=rec['name']).font = bf
        ws.cell(row=current_row, column=3).alignment = la
        ws.cell(row=current_row, column=3).border = thin_b

        ws.cell(row=current_row, column=4, value=rec['bid']).font = nf
        ws.cell(row=current_row, column=4).alignment = ca
        ws.cell(row=current_row, column=4).border = thin_b

        ws.cell(row=current_row, column=5, value=rec['month']).font = nf
        ws.cell(row=current_row, column=5).alignment = ca
        ws.cell(row=current_row, column=5).border = thin_b

        ws.cell(row=current_row, column=6, value=int(dnu_val)).font = nf
        ws.cell(row=current_row, column=6).alignment = ca
        ws.cell(row=current_row, column=6).border = thin_b

        ws.cell(row=current_row, column=7, value=round(cpi_val, 2)).font = nf
        ws.cell(row=current_row, column=7).alignment = ca
        ws.cell(row=current_row, column=7).border = thin_b

        ws.cell(row=current_row, column=8, value=f'{r1_val:.1%}').font = nf
        ws.cell(row=current_row, column=8).alignment = ca
        ws.cell(row=current_row, column=8).border = thin_b

        r2_w = rec['r2']
        ws.cell(row=current_row, column=9, value=f'{r2_w:.1%}' if r2_w > 0 else '—').font = nf
        ws.cell(row=current_row, column=9).alignment = ca
        ws.cell(row=current_row, column=9).border = thin_b

        ws.cell(row=current_row, column=10, value=f'{r3_val:.1%}').font = nf
        ws.cell(row=current_row, column=10).alignment = ca
        ws.cell(row=current_row, column=10).border = thin_b

        cs = rec['cost22']
        ws.cell(row=current_row, column=11, value=f'${cs:.2f}' if not pd.isna(cs) else '—').font = nf
        ws.cell(row=current_row, column=11).alignment = ca
        ws.cell(row=current_row, column=11).border = thin_b

        ws.cell(row=current_row, column=12, value=f'{rec["cvr"]:.1%}' if not pd.isna(rec['cvr']) else '—').font = nf
        ws.cell(row=current_row, column=12).alignment = ca
        ws.cell(row=current_row, column=12).border = thin_b

        ws.cell(row=current_row, column=13, value=f'${rec["spend"]:.0f}').font = nf
        ws.cell(row=current_row, column=13).alignment = ca
        ws.cell(row=current_row, column=13).border = thin_b

        ws.cell(row=current_row, column=14, value=reason_str).font = nf
        ws.cell(row=current_row, column=14).alignment = la
        ws.cell(row=current_row, column=14).border = thin_b

        # Color coding
        if tier.startswith('🥇'):
            fill = core_fill
        elif tier.startswith('🥈'):
            fill = ok_fill
        elif tier.startswith('❌'):
            fill = bad_fill
        else:
            fill = PatternFill()

        if is_duplicate:
            fill = dup_fill

        for col in range(1, 15):
            ws.cell(row=current_row, column=col).fill = fill

        ws.row_dimensions[current_row].height = 28
        current_row += 1
        idx += 1

    current_row += 1  # gap between sections

# =====================================================
# Summary sheet
# =====================================================
ws2 = wb.create_sheet('9月推荐-方向汇总', 1)

ws2.merge_cells('A1:H1')
ws2['A1'] = '9月测试素材推荐 — 方向级汇总'
ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

row = 3
sum_hdrs = ['方向', '总素材数', '强烈推荐', '可选', '不推荐', '核心结论', '素材池(推荐)', '建议预算占比']
for i, h in enumerate(sum_hdrs, 1):
    c = ws2.cell(row=row, column=i, value=h)
    c.font = hdr_f; c.fill = hdr_fill; c.alignment = ca; c.border = thin_b
ws2.row_dimensions[row].height = 30

# Count recommendations per direction
direction_summary = {
    '捉宠（核心方向）': '捉宠方向是素材储备最充足的核心方向。V-抓宠战斗AEO是唯一双月R1≥35%的底牌级素材。P-抓宠经营-幽飘R3=18.2%为方向内最佳长留。V-高效抓宠4月R1=37.5%+R3=21.9%但6月缺数据需重验。优先保AEO版抓宠战斗+幽飘，高效抓宠小预算验证。',
    '宠物融合/合成（核心方向）': '融合方向以P-宠物展示-合成3D(CPI=$4.36全场最低+DNU=232最可靠)和V-宠物展示-宠物合成(R3=14.6%大样本最高)双素材组合，覆盖图片+视频。素材数量仅3个但质量高。9月主推合成3D(跑量)+宠物合成(测留存)。',
    '宠物进化（核心方向）': '进化方向仅P-宠物展示-二阶进化1个素材(CPI=$3.68最低, R1=33.3%接近达标, DNU=27差3个)。素材储备严重不足。Task 2新素材脚本需重点补进化方向。二阶进化优先重测但需观察R3=7.4%在扩量后是否恶化。',
    '战斗': '战斗方向素材最多(16个)但质量分化大。V-帕萌战斗-群殴打鸡R1=52.9%全场最高但R3=5.9%断崖。帕萌战斗应重归类为捉宠子方向。杀宠复刻(CPI=$11,R1=21.5%)明确放弃。雪地竞品/出狱打鸡可小额辅测。',
    '模拟经营/建造': '模拟经营方向素材多(15个)但CPI普遍高。P-灾后重建(R1=44%+R3=33%+CPI=$4.13三冠王,仅4月DNU=9)是最大黑马。帕基玩法寒霜(R1=36%+R3=18%)留存好但CPI=$19.27不可接受,方向有效但需重做素材降CPI。其余模拟经营CPI全部>$8,不推荐。',
    '宠物展示': '宠物展示多为辅标签(与其他方向交叉)。独立宠物展示素材如P-宠物展示-巨物(R1=40%但DNU=5)数据不足。P-幻想宠物3(4月仅DNU=43+6,R1=16%)不推荐。',
    '探索/冒险/收集': '无独立强数据素材。V-帕基世界冒险(CPI=$5.41,R1=10%)不推荐。探索/收集多为辅标签。',
    '天灾/生存': 'P-海啸双月数据完整但非核心方向。4月AEO R1=37.9%曾达标,6月全面退化(R1=28%,CPI=$8.38)。Install版R1常年不达标。预算有限情况下放弃。',
}

# Count tiers per direction
for dir_name, dir_tags in directions.items():
    dir_data = []
    for _, rec in df_all.iterrows():
        if any(t in rec['tags'] for t in dir_tags):
            dir_data.append(rec)

    if not dir_data:
        continue

    df_dir = pd.DataFrame(dir_data)
    june = df_dir[df_dir['month'] == '6月']

    # Count
    total = len(df_dir)
    # Count recommended (June, DNU >= 10, CPI < 12 or R1 > 0.28)
    strong = len(june[(june['dnu'] >= 10) & ((june['cpi'] < 8) | (june['r1'] >= 0.30))])
    optional = len(june[(june['dnu'] >= 10) & (june['cpi'] < 12) & (june['cpi'] >= 8) & (june['r1'] < 0.30)])
    bad = total - strong - optional

    # Best picks
    best = june[june['dnu'] >= 5].nsmallest(3, 'cpi')
    picks = [f'{r["name"]}({r["bid"]},CPI${r["cpi"]:.2f},R1={r["r1"]:.0%})' for _, r in best.iterrows()]

    row += 1
    ws2.cell(row=row, column=1, value=dir_name).font = bf
    ws2.cell(row=row, column=1).alignment = la; ws2.cell(row=row, column=1).border = thin_b

    ws2.cell(row=row, column=2, value=total).font = nf
    ws2.cell(row=row, column=2).alignment = ca; ws2.cell(row=row, column=2).border = thin_b

    ws2.cell(row=row, column=3, value=strong).font = nf
    ws2.cell(row=row, column=3).alignment = ca; ws2.cell(row=row, column=3).border = thin_b
    ws2.cell(row=row, column=3).fill = core_fill

    ws2.cell(row=row, column=4, value=optional).font = nf
    ws2.cell(row=row, column=4).alignment = ca; ws2.cell(row=row, column=4).border = thin_b
    ws2.cell(row=row, column=4).fill = ok_fill

    ws2.cell(row=row, column=5, value=bad).font = nf
    ws2.cell(row=row, column=5).alignment = ca; ws2.cell(row=row, column=5).border = thin_b
    ws2.cell(row=row, column=5).fill = bad_fill

    ws2.cell(row=row, column=6, value=direction_summary.get(dir_name, '')).font = nf
    ws2.cell(row=row, column=6).alignment = la; ws2.cell(row=row, column=6).border = thin_b

    ws2.cell(row=row, column=7, value='\n'.join(picks[:3])).font = nf
    ws2.cell(row=row, column=7).alignment = la; ws2.cell(row=row, column=7).border = thin_b

    is_core = '核心方向' in dir_name
    ws2.cell(row=row, column=8, value='30-40%' if is_core else '5-15%').font = nf
    ws2.cell(row=row, column=8).alignment = ca; ws2.cell(row=row, column=8).border = thin_b

    ws2.row_dimensions[row].height = 65

# Final recommendation
row += 2
ws2.merge_cells(f'A{row}:H{row}')
ws2[f'A{row}'] = '▌9月测试最终推荐（按方向汇总后的结论）'
ws2[f'A{row}'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
ws2[f'A{row}'].fill = sec_fill

row += 1
recommendations = [
    '【捉宠方向 - 主推】V-抓宠战斗AEO(双月R1≥35%底牌)+ P-抓宠经营-幽飘Install(R3=18.2%长留王)+ V-高效抓宠Install(4月三项全能,需6月重验)+ P-抓宠经营-超梦Install(CPI=$4.74均衡基底)。共4个素材组。',
    '【融合方向 - 主推】P-宠物展示-合成3D Install(CPI=$4.36跑量主力)+ V-宠物展示-宠物合成Install(R3=14.6%大样本最高)。双素材覆盖图片+视频。共2个素材组。',
    '【进化方向 - 补强】P-宠物展示-二阶进化Install(CPI=$3.68最低,R1=33.3%)。唯一素材,小预算验证后决定是否加量。Task 2新素材优先补进化。',
    '【辅推 - 黑马观察】P-灾后重建Install(R1=44%+R3=33%三冠王,需6月重验)+ V-帕萌战斗-群殴打鸡Install(R1=52.9%天花板,关注R3暴跌原因)。各小预算验证。',
    '【明确放弃】海啸(CPI无优势+6月退化)、模拟经营除灾后重建外全部(CPI普遍>$10)、帕萌战斗-杀宠复刻(CPI=$11+R1=21.5%)。',
    '【预算分配】捉宠40%+融合30%+进化10%+辅推/机动20%。总计$40-50K。机动预算$5-8K留给Task 2竞品素材脚本。',
    '【P:V比例】维持现有比例(图片14:视频28≈1:2)。图片CPI优势显著(双月低$0.5-1.9),但视频覆盖更多方向。不做人为调整。',
]
for r in recommendations:
    ws2.merge_cells(f'A{row}:H{row}')
    ws2[f'A{row}'] = r
    ws2[f'A{row}'].font = Font(name='Arial', size=10)
    ws2[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws2.row_dimensions[row].height = 32
    row += 1

# Column widths for summary sheet
for i, w in enumerate([28, 10, 10, 10, 10, 55, 45, 15], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save('creative_final.xlsx')
print('Done! Sheets added: 9月推荐-按方向, 9月推荐-方向汇总')
