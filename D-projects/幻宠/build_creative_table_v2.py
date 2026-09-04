#!/usr/bin/env python3
"""Build creative performance table V2 - fixed with all review findings."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule, CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "素材表现汇总"

# ============================================================
# Styles
# ============================================================
header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2972F4", end_color="2972F4", fill_type="solid")
april_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
june_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
summary_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
small_sample_fill = PatternFill(start_color="F4F4F4", end_color="F4F4F4", fill_type="solid")
aeo_font = Font(name="Microsoft YaHei", size=10, color="C00000", bold=False)
install_font = Font(name="Microsoft YaHei", size=10, color="1F4E79", bold=False)
both_font = Font(name="Microsoft YaHei", size=10, color="7030A0", bold=False)
summary_font = Font(name="Microsoft YaHei", size=10, bold=True, color="333333")
small_font = Font(name="Microsoft YaHei", size=9, color="999999", italic=True)
warn_font = Font(name="Microsoft YaHei", size=10, color="FF0000", bold=True)
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0")
)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ============================================================
# Headers
# ============================================================
headers = ["月份", "素材名称", "素材类型", "出价方式", "安装数", "花费(USD)", "CPI", "R1", "R2", "R3", "R7", "备注"]
col_widths = [7, 30, 8, 12, 9, 12, 9, 9, 9, 9, 9, 58]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[1].height = 26

# ============================================================
# Data builder
# ============================================================
rows = []

def add_sep(month):
    rows.append({"month": month, "sep": True})

def add_row(month, name, ctype, bidding, installs, spend, cpi, r1, r2, r3, r7, note=""):
    rows.append({
        "month": month, "sep": False,
        "name": name, "ctype": ctype, "bidding": bidding,
        "installs": installs, "spend": spend, "cpi": cpi,
        "r1": r1, "r2": r2, "r3": r3, "r7": r7,
        "note": note
    })

# ============================================================
# April AEO per-creative (from sheet 8HCs2U, FB platform data)
# ============================================================
add_sep("4月")
april_aeo_data = [
    ("V-抓宠战斗",     "V", 80,  459.97, 5.75, 0.3871, 0.1900, "-", "-", "CVR 40.61%, CTR 0.85%"),
    ("P-海啸",        "P", 81,  441.36, 5.45, 0.4000, 0.1833, "-", "-", "CVR 23.48%, CTR 2.04%"),
    ("V-场景展示",     "V", 11,   93.31, 8.48, 0.2000, 0.0909, "-", "-", "CVR 34.38%, CTR 1.04%; n<20小样本"),
    ("P-幻想宠物3",    "P",  9,   50.98, 5.66, 0.2000, 0.0000, "-", "-", "CVR 21.95%, CTR 1.45%; n<10小样本"),
    ("V-核心玩法",     "V",  7,   42.51, 6.07, 0.2000, 0.0000, "-", "-", "CVR 33.33%, CTR 1.09%; n<10小样本"),
    ("V-天灾重建-长版", "V",  4,   34.48, 8.62, 0.5000, 0.2500, "-", "-", "n=4极小样本, CVR 23.53%"),
]
for name, ctype, inst, spend, cpi, r1, r2, r3, r7, note in april_aeo_data:
    add_row("4月", name, ctype, "AEO", inst, spend, cpi, r1, r2, r3, r7, note)

# April Install summary (no per-creative data in source)
add_row("4月", "【Install整体】", "-", "Install", 541, 2117.82, 3.91, 0.2890, 0.0950, "-", "-",
        "541安装(FB), 436激活(Singular); "
        "留存最佳(>35%): V-核心玩法/V-高效抓宠/P-灾后重建; "
        "留存最差(~15%): P-幻想宠物3/V-帕基世界冒险(像素风); "
        "吸量最佳($3.3-3.5): V-场景展示/V-核心玩法/V-抓宠战斗; 源数据无每素材明细")

# April AEO summary
add_row("4月", "【AEO整体】", "-", "AEO", 192, 1122.61, 5.85, 0.3699, 0.1960, "-", "-",
        "AEO-tur7(~6min): 次留39.69% CPI$5.58; AEO-强制引导(~4min): 次留13.33%(已关停)")

# ============================================================
# June - 5 dual-bidding creatives (AEO data from z89zqR)
# ============================================================
add_sep("6月")

# AEO campaign data (from z89zqR Section 2, rows 36-40)
june_aeo_5 = [
    ("P-海啸",           "P", 166, 1407.25, 8.48, "-", "-", "-", "-", "AEO专场; CVR 25.00%, CTR 1.23%"),
    ("V-抓宠战斗",       "V", 101,  850.65, 8.42, "-", "-", "-", "-", "AEO专场; CVR 40.08%, CTR 0.70%"),
    ("V-天灾重建-长版",   "V",  31,  260.86, 8.41, "-", "-", "-", "-", "AEO专场; CVR 43.06%, CTR 0.96%"),
    ("V-场景展示",       "V",  16,  142.27, 8.89, "-", "-", "-", "-", "AEO专场; CVR 47.06%, CTR 1.06%"),
    ("V-核心玩法",       "V",   8,   76.73, 9.59, "-", "-", "-", "-", "AEO专场; CVR 32.00%, CTR 0.85%; n<10小样本"),
]
for name, ctype, inst, spend, cpi, r1, r2, r3, r7, note in june_aeo_5:
    add_row("6月", name, ctype, "AEO", inst, spend, cpi, r1, r2, r3, r7, note)

# Install campaign data for same 5 creatives (from z89zqR Section 3, Install campaigns)
june_install_5 = [
    ("P-海啸",           "P", 146, 1270.76, 8.70, "-", "-", "-", "-", "Install专场(新旧商店页合并)"),
    ("V-抓宠战斗",       "V",  85,  423.55, 4.98, "-", "-", "-", "-", "Install专场(新旧商店页合并); CTR 3.01%"),
    ("V-场景展示",       "V",  28,  156.84, 5.60, "-", "-", "-", "-", "Install专场(新旧商店页合并)"),
    ("V-天灾重建-长版",   "V",  17,  262.42, 15.44, "-", "-", "-", "-", "Install专场(新旧商店页合并); n<20小样本"),
    ("V-核心玩法",       "V",   3,   32.40, 10.80, "-", "-", "-", "-", "Install专场; n=3极小样本"),
]
for name, ctype, inst, spend, cpi, r1, r2, r3, r7, note in june_install_5:
    add_row("6月", name, ctype, "Install", inst, spend, cpi, r1, r2, r3, r7, note)

# ============================================================
# June Install creative-direction creatives (from 详细的 FfBRvO)
# These are mainly from 素材方向测试 (creative direction test)
# ============================================================
june_detailed = [
    ("P-宠物展示-合成 3D",       "P", 292, 1277.41, 4.37, 0.2572, "-", "-", "-", "CTR 1.65%, CVR 19.76%"),
    ("V-宠物展示-宠物合成",       "V", 142,  976.48, 6.88, 0.2522, "-", "-", "-", "CTR 1.02%, CVR 26.01%"),
    ("P-模拟经营-冬日写实",       "P", 130, 1351.30, 10.39, 0.1852, "-", "-", "-", "CTR 1.27%, CVR 14.67%"),
    ("V-帕萌战斗-杀宠复刻",       "V",  83,  912.48, 10.99, 0.2500, "-", "-", "-", "CTR 1.58%, CVR 18.32%"),
    ("P-抓宠经营-超梦",          "P",  79,  374.38,  4.74, 0.2813, "-", "-", "-", "CTR 1.42%, CVR 25.65%"),
    ("V-抓宠经营-捕捞竞品",       "V",  70,  272.44,  3.89, 0.2182, "-", "-", "-", "CTR 1.59%, CVR 32.26%"),
    ("P-抓宠经营-幽飘",          "P",  54,  473.00,  8.76, 0.3409, "-", "-", "-", "CTR 0.94%, CVR 21.77%"),
    ("V-抓宠经营-售卖帕基",       "V",  43,  383.06,  8.91, 0.3077, "-", "-", "-", "CTR 1.23%, CVR 23.89%"),
    ("V-帕萌战斗-雪地竞品",       "V",  37,  276.61,  7.48, 0.3333, "-", "-", "-", "CTR 3.06%, CVR 12.17%"),
    ("P-宠物展示-二阶进化",       "P",  29,  106.65,  3.68, 0.3103, "-", "-", "-", "CTR 2.40%, CVR 15.85%"),
    ("P-抓宠经营-虐待",          "P",  20,  150.37,  7.52, 0.3750, "-", "-", "-", "CTR 1.81%, CVR 13.61%"),
    ("V-帕萌战斗-出狱打鸡",       "V",  17,  153.22,  9.01, 0.4000, "-", "-", "-", "CTR 1.07%, CVR 25.37%; n<20小样本"),
    ("V-帕萌战斗-群殴打鸡",       "V",  14,  114.96,  8.21, 0.5294, "-", "-", "-", "CTR 3.44%, CVR 11.86%; n<20, R1高但样本小"),
    ("V-模拟经营-温馨",          "V",  12,  100.78,  8.40, 0.3750, "-", "-", "-", "CTR 1.23%, CVR 23.08%; n<20小样本"),
    ("P-宠物展示-帕基战斗",       "P",  11,   76.26,  6.93, 0.1000, "-", "-", "-", "CTR 1.12%, CVR 18.33%; n<20, R1极低⚠"),
    ("P-模拟经营-建造成长",       "P",  11,  156.68, 14.24, 0.2500, "-", "-", "-", "CTR 1.34%, CVR 10.78%; n<20, CPI高⚠"),
    ("V-模拟经营-帕基玩法寒霜",    "V",  10,  192.73, 19.27, 0.4000, "-", "-", "-", "CTR 1.90%, CVR 7.58%; n<10, CPI极高⚠"),
    ("V-抓宠经营-拯救可达鸭",      "V",   9,  168.85, 18.76, 0.2000, "-", "-", "-", "CTR 3.41%, CVR 6.08%; n<10, CPI极高⚠"),
    ("P-宠物展示-巨物",          "P",   7,   77.95, 11.14, 0.3333, "-", "-", "-", "CTR 1.51%, CVR 13.21%; n<10小样本"),
    ("V-宠物展示-二阶合成",       "V",   6,   32.92,  5.49, 0.1667, "-", "-", "-", "CTR 1.26%, CVR 27.27%; n<10小样本"),
    ("V-模拟经营-七日建造",       "V",   6,   84.70, 14.12, 0.0000, "-", "-", "-", "⚠ R1=0%; n<10; CTR 1.63%"),
    ("V-模拟经营-砍树",          "V",   5,  115.65, 23.13, 0.0000, "-", "-", "-", "⚠ R1=0%, CPI极高$23; n<10, 建议关停"),
    ("V-抓宠经营-狐狸",          "V",   5,   77.60, 15.52, 0.2000, "-", "-", "-", "n<10小样本"),
    ("V-抓宠经营-虐待",          "V",   5,   19.89,  3.98, 0.5000, "-", "-", "-", "n<10, R1高但样本极小"),
    ("V-帕萌战斗-合成狙击",       "V",   4,   24.95,  6.24, 0.0000, "-", "-", "-", "⚠ R1=0%; n<10极小样本"),
    ("P-抓宠经营-血腥",          "P",   2,   42.37, 21.19, 0.0000, "-", "-", "-", "⚠ R1=0%, CPI极高$21; n=2, 排除"),
]
for name, ctype, inst, spend, cpi, r1, r2, r3, r7, note in june_detailed:
    add_row("6月", name, ctype, "Install", inst, spend, cpi, r1, r2, r3, r7, note)

# ============================================================
# June segment summaries
# ============================================================
add_row("6月", "【Install-素材方向测试(新商店页)】", "-", "Install", 964, 7991.96, 7.25, 0.2635, 0.1504, 0.0934, "-",
        "素材方向测试-0611广告组; 安装=964为广告组级DNU, 与上方明细合计可能不一致")
add_row("6月", "【Install-留存素材(旧商店页)】", "-", "Install", 250, 2180.07, 7.70, 0.2320, 0.0840, 0.0840, "-",
        "US-AAA-install广告组; 安装=250为广告组级DNU")
add_row("6月", "【AEO整体】", "-", "AEO", 273, 2737.81, 8.50, 0.2967, 0.1648, 0.1209, "-",
        "AEO广告组整体; 上方AEO素材合计322安装, 差异因广告组级vs素材级归因口径不同")

# ============================================================
# Write data to worksheet
# ============================================================
current_month = None
excel_row = 2

for r in rows:
    if r.get("sep"):
        month = r["month"]
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=excel_row, column=c)
            cell.fill = april_fill if "4月" in month else june_fill
            cell.border = thin_border
        sep_cell = ws.cell(row=excel_row, column=1)
        sep_cell.value = f"◆ {month}"
        sep_cell.font = Font(name="Microsoft YaHei", size=11, bold=True, color="1F4E79")
        sep_cell.alignment = left_align
        ws.merge_cells(start_row=excel_row, start_column=1, end_row=excel_row, end_column=len(headers))
        ws.row_dimensions[excel_row].height = 22
        excel_row += 1
        continue

    is_summary = "【" in str(r["name"])
    is_april = "4月" in str(r["month"])
    bidding = r["bidding"]
    installs_val = r["installs"]
    is_small = isinstance(installs_val, (int, float)) and installs_val < 20 and not is_summary

    # Font & fill
    if is_summary:
        cell_font = summary_font
        cell_fill = summary_fill
    elif bidding == "AEO":
        cell_font = aeo_font
        cell_fill = april_fill if is_april else june_fill
    elif bidding == "Install":
        cell_font = install_font
        cell_fill = april_fill if is_april else june_fill
    else:
        cell_font = both_font
        cell_fill = june_fill

    if is_small and not is_summary:
        cell_font = Font(name="Microsoft YaHei", size=9, color="888888", italic=True)

    # Build values
    def fmt_pct(v):
        if v is None or v == "-" or v == '':
            return "-"
        if isinstance(v, str):
            return v
        return v  # Return as float for number format

    def fmt_dollar(v):
        if v is None or v == "-" or v == '':
            return "-"
        if isinstance(v, str):
            return v.strip()
        return v

    values = [
        r["month"],
        r["name"],
        r["ctype"],
        r["bidding"],
        r["installs"] if r["installs"] != "-" else "-",
        fmt_dollar(r["spend"]),
        r["cpi"] if r["cpi"] != "-" else "-",
        fmt_pct(r["r1"]),
        fmt_pct(r["r2"]),
        fmt_pct(r["r3"]),
        fmt_pct(r["r7"]),
        r["note"]
    ]

    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=excel_row, column=col_idx, value=val)
        cell.font = cell_font
        cell.fill = cell_fill
        cell.border = thin_border

        # Alignment
        if col_idx in (1, 3, 4, 5):
            cell.alignment = center_align
        elif col_idx in (6, 7):
            cell.alignment = center_align
        elif col_idx in (8, 9, 10, 11):
            cell.alignment = center_align
        else:
            cell.alignment = left_align

        # Number format
        if col_idx in (6, 7) and isinstance(val, (int, float)):
            cell.number_format = '$#,##0.00'
        elif col_idx in (8, 9, 10, 11) and isinstance(val, (int, float)):
            cell.number_format = '0.00%'
        elif col_idx == 5 and isinstance(val, (int, float)):
            cell.number_format = '#,##0'

    ws.row_dimensions[excel_row].height = 22
    excel_row += 1

# ============================================================
# Freeze pane & auto-filter
# ============================================================
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{excel_row - 1}"

# ============================================================
# Conditional formatting for R1 column (H)
# ============================================================
r1_col = "H"
r1_range = f"{r1_col}2:{r1_col}{excel_row - 1}"
ws.conditional_formatting.add(r1_range,
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"))

# CPI column (G) - data bar
cpi_col = "G"
cpi_range = f"{cpi_col}2:{cpi_col}{excel_row - 1}"
ws.conditional_formatting.add(cpi_range,
    DataBarRule(start_type="min", end_type="max", color="5B9BD5", showValue=True))

# ============================================================
# Sheet 2: 说明
# ============================================================
ws2 = wb.create_sheet("说明")
legend = [
    ("幻宠帝国 — 素材表现汇总表", True, 14),
    ("", False, 10),
    ("数据范围", True, 11),
    ("测试", "4月留存测试 (4/2-4/4) + 6月测试 (5/16-6/15)，仅美国地区", False, 10),
    ("行定义", "每素材 × 出价方式 = 一行，同素材跑AEO+Install的分两行", False, 10),
    ("更新时间", "2026-07-24", False, 10),
    ("", False, 10),
    ("颜色说明", True, 11),
    ("黄色背景", "4月数据", False, 10),
    ("绿色背景", "6月数据", False, 10),
    ("红色字体", "AEO出价", False, 10),
    ("蓝色字体", "Install出价", False, 10),
    ("灰色斜体", "样本量<20，统计意义有限", False, 10),
    ("灰色底+粗体", "汇总行", False, 10),
    ("", False, 10),
    ("数据源", True, 11),
    ("4月AEO每素材", "飞书Wiki「幻想宠物留存测试报告」→ 嵌入表格 Z0l2s2bs1he1TotM2ERcgBNHnEf → Sheet「8HCs2U」(美国-AEO)", False, 10),
    ("4月Install", "仅汇总层级(541安装, CPI $3.91, R1 28.9%)，源数据无每素材拆分。素材方向结论来自文档文字总结。", False, 10),
    ("6月AEO每素材", "飞书表格「6月11日幻宠测试」Bnkpsh602hZL49t39cTcT1f4nBq → Sheet「z89zqR」(整理表) AEO专区(行35-40)", False, 10),
    ("6月Install每素材", "同上 → Sheet「FfBRvO」(详细的) + 「z89zqR」Install广告组数据", False, 10),
    ("6月留存", "同上 → Sheet「azRg8b」(宏观数据-留存) 广告组级数据", False, 10),
    ("", False, 10),
    ("已知限制", True, 11),
    ("1. 4月Install无每素材数据", "源数据为按天汇总，每素材安装/CPI/R1不可拆分。表中4月Install每素材行已省略，仅保留汇总行。", False, 10),
    ("2. 6月AEO素材R1缺失", "5个AEO素材(P-海啸/V-抓宠战斗/V-场景展示/V-核心玩法/V-天灾重建-长版)有安装和CPI，但源数据无每素材R1。仅AEO整体R1=29.67%可参考。", False, 10),
    ("3. R3/R7严重缺失", "两个测试的源数据均无每素材R3/R7。6月仅3个广告组级汇总行有R3，R7完全不存在。", False, 10),
    ("4. 安装数口径差异", "6月【详细的】表26素材安装合计约1104(含2个异常值)，【整体行】素材方向测试DNU=964。差异(140)来自：①部分素材属于旧商店页(留存素材)而非素材方向测试 ②FB/Singular归因口径不同。", False, 10),
    ("5. 小样本警告", "n<10的素材(灰色斜体行)R1和CPI统计意义有限，不应用于决策。", False, 10),
    ("6. CPI列", "来源为Facebook平台数据(除以FB安装数)。6月Install素材来自【详细的】表(已含所有归因)，AEO素材来自【整理表】AEO专区。", False, 10),
    ("", False, 10),
    ("6月广告组级留存（参考）", True, 11),
    ("素材方向测试-0611 (新商店页)", "R1=26.35%, R2=15.04%, R3=9.34%", False, 10),
    ("US-AAA-install (旧商店页/留存素材)", "R1=23.20%, R2=8.40%, R3=8.40%", False, 10),
    ("AEO (核心用户-AEO)", "R1=29.67%, R2=16.48%, R3=12.09%", False, 10),
    ("", False, 10),
    ("素材方向-留存表现排名(6月Install, n>=10)", True, 11),
    ("TOP5 R1", "V-帕萌战斗-群殴打鸡 52.94% > V-抓宠经营-虐待 50.00% > V-帕萌战斗-出狱打鸡 40.00% = V-模拟经营-帕基玩法寒霜 40.00% > V-模拟经营-温馨/P-抓宠经营-虐待 37.50%", False, 10),
    ("BOTTOM5 R1", "V-模拟经营-砍树 0% = V-模拟经营-七日建造 0% = V-帕萌战斗-合成狙击 0% = P-抓宠经营-血腥 0% < P-宠物展示-帕基战斗 10.00%", False, 10),
    ("TOP5 吸量(CPI最低)", "P-宠物展示-二阶进化 $3.68 < V-抓宠经营-捕捞竞品 $3.89 < V-抓宠经营-虐待 $3.98 < P-宠物展示-合成3D $4.37 < P-抓宠经营-超梦 $4.74", False, 10),
]

for row_idx, row_data in enumerate(legend, 1):
    if len(row_data) == 3:
        text, is_bold, font_size = row_data
        cell = ws2.cell(row=row_idx, column=1, value=text)
        cell.font = Font(name="Microsoft YaHei", size=font_size, bold=is_bold)
    elif len(row_data) == 4:
        label, value, is_bold, font_size = row_data
        cell_a = ws2.cell(row=row_idx, column=1, value=label)
        cell_a.font = Font(name="Microsoft YaHei", size=font_size, bold=is_bold)
        cell_b = ws2.cell(row=row_idx, column=2, value=value)
        cell_b.font = Font(name="Microsoft YaHei", size=font_size)
    elif len(row_data) == 2:
        text, is_bold = row_data[:2]
        font_size = row_data[1] if len(row_data) > 1 else 10

ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 85

# ============================================================
# Save
# ============================================================
output_path = r"D:\claude-projects\projects\幻宠\幻宠素材表现汇总_4月6月_US.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Data rows: {excel_row - 2}")
