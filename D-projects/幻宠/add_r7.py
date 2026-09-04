# -*- coding: utf-8 -*-
"""把 r7 留存数据补充到 creative_final.xlsx 的 素材表现汇总 + 素材全量数据-按素材 两个 sheet"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import shutil

SRC_BI = r"D:\claude-projects\projects\幻宠\fb媒体-源数据.xlsx"
CF = r"D:\claude-projects\projects\幻宠\creative_final.xlsx"

# ---------- 1. 从 BI 计算 r7 映射 ----------
map_rows = []
for sn, mon, bid in [('4月install美国幻宠', '4月', 'Install'), ('4月AEO美国幻宠', '4月', 'AEO'),
                     ('6月install美国', '6月', 'Install'), ('6月AEO美国', '6月', 'AEO')]:
    df = pd.read_excel(SRC_BI, sheet_name=sn)
    for _, r in df.iterrows():
        map_rows.append((str(r['广告编号']).strip(), r['广告名称'], mon, bid))
mapping = pd.DataFrame(map_rows, columns=['素材ID', '素材名称', '月份', '出价方式'])
bi = pd.read_excel(SRC_BI, sheet_name='BI汇总')
bi['素材ID'] = bi['素材ID'].astype(str).str.strip()
merged = bi.merge(mapping, on='素材ID', how='inner')
g = merged.groupby(['月份', '出价方式', '素材名称']).agg(
    dnu=('dnu', 'sum'), r1=('r1_cnt', 'sum'), r2=('r2_cnt', 'sum'),
    r3=('r3_cnt', 'sum'), r7=('r7_cnt', 'sum')).reset_index()

# 素材级 r7 映射：key=(月份,出价方式,素材名称)
r7map = {(r['月份'], r['出价方式'], r['素材名称']): r for _, r in g.iterrows()}

# 出价方式级合计（4个）
tot = merged.groupby(['月份', '出价方式']).agg(
    dnu=('dnu', 'sum'), r1=('r1_cnt', 'sum'), r2=('r2_cnt', 'sum'),
    r3=('r3_cnt', 'sum'), r7=('r7_cnt', 'sum')).reset_index()
totmap = {(r['月份'], r['出价方式']): r for _, r in tot.iterrows()}

# ---------- 2. 备份 ----------
BAK = CF.replace('.xlsx', '_备份.xlsx')
shutil.copy(CF, BAK)

wb = openpyxl.load_workbook(CF)

# ================= Sheet 1: 素材表现汇总 =================
ws = wb['素材表现汇总']
# 列索引（1-based）：14=DNU 15=r1_cnt 16=r2_cnt 17=r3_cnt 18=R1 19=R2 20=R3 21=次留成本
ws.insert_cols(18)   # 在 r3_cnt(17) 后插入 r7_cnt
ws.insert_cols(22)   # 在 R3(21) 后插入 R7
# 表头
ws.cell(row=1, column=18, value='r7_cnt')
ws.cell(row=1, column=22, value='R7')

# 逐行填数据
for r in range(2, ws.max_row + 1):
    mon = ws.cell(row=r, column=1).value
    name = ws.cell(row=r, column=2).value
    bid = ws.cell(row=r, column=4).value
    if mon not in ('4月', '6月'):
        continue
    if name is None:
        continue
    name = str(name).strip()
    if name.startswith('【'):  # 合计行
        key = (mon, bid)
        if key in totmap:
            t = totmap[key]
            ws.cell(row=r, column=14, value=int(t['dnu']))   # DNU
            ws.cell(row=r, column=15, value=int(t['r1']))    # r1_cnt
            ws.cell(row=r, column=16, value=int(t['r2']))    # r2_cnt
            ws.cell(row=r, column=17, value=int(t['r3']))    # r3_cnt
            ws.cell(row=r, column=18, value=int(t['r7']))    # r7_cnt
            ws.cell(row=r, column=19, value=round(t['r1'] / t['dnu'], 6))  # R1
            ws.cell(row=r, column=20, value=round(t['r2'] / t['dnu'], 6))  # R2
            ws.cell(row=r, column=21, value=round(t['r3'] / t['dnu'], 6))  # R3
            ws.cell(row=r, column=22, value=round(t['r7'] / t['dnu'], 6))  # R7
    else:
        key = (mon, bid, name)
        if key in r7map:
            m = r7map[key]
            ws.cell(row=r, column=18, value=int(m['r7']))               # r7_cnt
            ws.cell(row=r, column=22, value=round(m['r7'] / m['dnu'], 6))  # R7

# ================= Sheet 2: 素材全量数据-按素材 =================
ws2 = wb['素材全量数据-按素材']
# 列：10=R3 11=次留成本 ... 在 R3(10) 后插入 R7
# 正确处理所有合并单元格：先记录 -> unmerge -> insert -> 按旧坐标重新 merge
from openpyxl.utils import range_boundaries
old_merges = [(str(m), range_boundaries(str(m))) for m in ws2.merged_cells.ranges]
for m in list(ws2.merged_cells.ranges):
    ws2.unmerge_cells(str(m))
ws2.insert_cols(11)
for old_str, (min_col, min_row, max_col, max_row) in old_merges:
    if max_col < 11:  # 完全在插入点左边：不变
        new_str = old_str
    elif min_col >= 11:  # 完全在插入点右边：右移1
        new_str = f"{get_column_letter(min_col + 1)}{min_row}:{get_column_letter(max_col + 1)}{max_row}"
    else:  # 跨越插入点：扩展1列
        new_str = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col + 1)}{max_row}"
    ws2.merge_cells(new_str)
ws2.cell(row=3, column=11, value='R7')

# 逐行填数据（处理合并单元格：素材名继承）
cur_name = None
for r in range(4, ws2.max_row + 1):
    name = ws2.cell(row=r, column=1).value
    mon = ws2.cell(row=r, column=4).value
    bid = ws2.cell(row=r, column=5).value
    dnu = ws2.cell(row=r, column=6).value
    if name is not None:
        cur_name = str(name).strip()
    if mon is None or bid is None or dnu is None:
        continue
    key = (str(mon).strip(), str(bid).strip(), cur_name)
    if key in r7map:
        m = r7map[key]
        pct = round(m['r7'] / m['dnu'] * 100, 1)
        ws2.cell(row=r, column=11, value=f"{pct}%")

wb.save(CF)
print("保存完成")
print("备份文件:", BAK)
