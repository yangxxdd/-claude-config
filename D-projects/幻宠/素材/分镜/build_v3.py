# -*- coding: utf-8 -*-
"""PET COURT v3 分镜表——五维审查修订版"""

import openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============ Sheet 1: 项目信息 ============
ws1 = wb.active
ws1.title = "项目信息"
info = [
    ["素材名称", "PET COURT（宠物抚养权法庭）v3"],
    ["方向", "抓马 · 无厘头恶搞"],
    ["钩子公式", "美式法庭真人秀格式寄生 + 三层反转 + 机制punchline（五维审查全面修订版）"],
    ["时长", "约50s（可剪30s/15s短版）"],
    ["规格", "竖屏 9:16，英文配音+英文字幕，美国市场"],
    ["主角帕基", "幽飘（10111 WoodGril，4月R3=18.2%素材明星）。注意：幽飘为游戏内新手线宠物，本版控诉逻辑已从'它多稀有'改为'我付出多少'（孵化+进化+融合狗粮），避免核心玩家出戏"],
    ["咒语/Slogan", "\"Everyone's catchable.\"（法警deadpan说出 +\"Court's adjourned.\"）"],
    ["CTA", "A版: Who Will You Catch First? / B版: Your SSR Is Waiting — Play Free（两版AB测试）"],
    ["核心机制露出", "孵化（audio_sfx_born存在）、进化（真实三阶链）、融合（真实机制）、捕捉球收服（吸入+读条，游戏真实表现）"],
    ["收服仪式 (v3重做)", "弃用'晃三下+叮'（任天堂trade dress风险）。改用游戏真实表现：光束卷入(buzhuo1吸入音)→读条UI闪现(buzhuo2)→球体闭合震动一次+原创确认音。全程卡通光效，不做痛苦挣扎"],
    ["制作总策略", "人类角色全AI生成 | 帕基用FBX模型DCC渲染（UE5/Blender，S2做pipeline test定管线） | AE合成/特效/字幕 | 真实游戏画面仅在结尾1.5s露出捕捉界面"],
    ["合规清单 (必执行)", "1) 投放端勾选AIGC标注(TikTok/Meta强制) 2) 结尾卡加小字'Dramatization. Not actual gameplay footage.' 3) 法官定妆避开Judy Sheindlin签名特征(深色波波头/蕾丝领/尖细嗓音)，出图后做相似度对比留档 4) AI提示词及内部文档禁止出现'Judge Judy''宝可梦'字样 5) 片头法徽/字幕条/音乐全部原创(法徽可嵌捕捉球剪影)，音乐用罐头授权杜绝sound-alike"],
    ["文档规范", "内部文档统一表述：'参考法庭真人秀类型通用范式(courtroom reality TV genre tropes)'，不点名任何具体节目/竞品IP"],
    ["资产来源警示", "模型目录名为'宝可梦资产汇总'，进制作前必须向资产团队核实幽飘等模型来源(自研/授权/改模)。若来源有问题，本项目停做升级处理"],
    ["三层反转", "1 离婚争的不是房子是宠物 -> 2 法官让宠物自己选(期待拉满) -> 3 宠物把法官收了(punchline)+ 法警补刀(世界观记忆点)"],
    ["风险预案", "AI人脸一致性: 每角色定妆参考图锁定+每镜3-5抽 | 多角色同框拆单人快切 | 法官消散: AI冻结帧+AE粒子，空景背景在场景生成阶段同步输出 | S11掏球改遮挡剪辑"],
    ["短版剪辑点", "30s版: S3/S5台词各保留一半。15s版重排: S0(2s)->S2(2s)->S7压3s(只留creature CHOOSES)->S11+S12合并4s->S13压1.5s(保Everyone's catchable)->S15(2.5s)，咒语不可砍"],
    ["工期估算", "15-22个工作日。最大瓶颈: AI人脸一致性抽卡返工、FBX渲染管线材质重建(一次性)、S12空景依赖链"],
    ["版本记录", "v1 初版 | v2 S3台词消除指代歧义 | v3 五维审查修订(买量/本地化/制作/法务/机制) -> 4个Sheet含19项修改记录"],
]
ws1.append(["项目", "内容"])
for row in info:
    ws1.append(row)

# ============ Sheet 2: 分镜表 ============
ws2 = wb.create_sheet("分镜表")
hdrs = ["镜头","时间","时长","景别/运镜","画面内容","英文台词/字幕","中文对照","音效/BGM","制作方式","资产/参考","制作原因"]
ws2.append(hdrs)

shots = [
    ["S0","0:00-0:03","3s","全景直接起，法槌特写快切插入","(v3: 去黑屏首帧) 首帧即法庭全景+原创片头包装（金色法徽嵌捕捉球剪影+字幕条），中央大字钩子直接砸出",
     "大字钩子: They agree on everything... except ONE. | 字幕条: CASE #4,847 — Thompson v. Thompson",
     "他们什么都谈妥了……除了一样。 | 案件4847: 汤普森夫妇",
     "法槌x3+原创铜管stinger(罐头授权，禁sound-alike)",
     "AI(法庭场景+法官定妆) + AE(原创片头包装)",
     "参考: 法庭真人秀类型通用范式(不点名具体节目)",
     "75%用户静音刷feed，黑屏+纯音效钩子无效；首帧必须有画面+文字双钩子。法徽嵌捕捉球=原创且服务品牌"],

    ["S1","0:03-0:06","3s","中景，法官正面缓推","法官(50+女性，灰白短发，犀利眼神，法官袍——主动避开Sheindlin签名特征)看着卷宗抬眼",
     'Judge: "The Thompsons are divorcing. House, car, money — settled. They agree on EVERYTHING... except one thing."',
     "法官: 汤普森夫妇要离婚了。房子、车子、钱——都分完了。什么都谈妥了……除了一样。",
     "环境音压低，法官低沉沙哑嗓音(避开尖细高频)，尾音留白",
     "AI(法官表演)",
     "法官定妆参考图(出图后与Judy实际形象做相似度对比并留档)",
     "排比句强化真人秀开场腔; except one thing悬念钩子"],

    ["S2","0:06-0:09","3s","慢推+低角度仰拍","镜头缓缓推向当事人席——幽飘端坐在椅子上，戴小领结，眼神无辜，耳朵轻抖(人类夫妻站立场边讲台，符合法庭秀格式)",
     "(无台词)",
     "(无台词)",
     "观众席窃语渐起+一声滑稽高音弦乐",
     "DCC渲染(幽飘FBX+领结AE贴合) + AE(合成进AI场景，加接触阴影+统一胶片LUT)",
     "模型: [宝可梦资产汇总]/宝可梦动作文件汇总(1)/宝可梦动作文件汇总/帕基/木少女一阶10111/mod_10111_high.fbx | 立绘参考: 帕基截图/11【幽飘】10111WoodGril.png",
     "反转1落地: 争夺对象是宠。领结制造荒诞感。本镜头为pipeline test镜头，先做通再批量"],

    ["S3","0:09-0:15","6s","中近景，妻子起立","(v3: 表演改ice-cold冷怒毒舌) 妻子(30+，妆容精致，不哭闹，慢条斯理每个字从牙缝挤出)。首次开口砸名牌字幕条",
     '字幕条: KAREN THOMPSON — Plaintiff | Wife: "Your Honor — he used her as FUSION FODDER. I HATCHED her. I was there for her first evolution... and he tried to feed her to a FUSION MACHINE."',
     "字幕条: 凯伦·汤普森——原告 | 妻子: 法官大人——他拿她当融合狗粮。她是我孵化的，她第一次进化我在场……而他想把她喂给融合机。",
     "冷静但压抑的怒火，背景观众低声惊呼",
     "AI(妻子表演，无声表演+配音+口型同步两步走)",
     "妻子定妆参考图",
     "(v3重写台词) 原COMMON berries+0.01% spawn rate虚构了喂食系统和刷新率数值(游戏均无)，改为孵化/进化/融合三个真实机制; 融合狗粮是玩家真实痛点更好笑。冷怒表演更像法庭秀、规避刻板印象、AI一致性更高"],

    ["S4","0:15-0:16","1s","观众席特写快切","观众席大妈捂嘴倒吸凉气",
     "(无台词)",
     "(无台词)",
     "整齐倒吸气声",
     "AI(群演)",
     "群演参考图",
     "反应镜头是真人秀节奏的呼吸点"],

    ["S5","0:16-0:21","5s","中近景，丈夫反击","丈夫(30+，格子衫，委屈爆发，拍桌而起)。首次开口砸名牌字幕条",
     '字幕条: DOUG THOMPSON — Defendant | Husband: "I was THERE when she evolved! THREE A.M.! Where were YOU?! Oh, right — BOOK CLUB."',
     "字幕条: 道格·汤普森——被告 | 丈夫: 她进化的时候我在场！凌晨三点！你在哪？！哦对——读书会。",
     "拍桌+观众爆笑",
     "AI(丈夫表演)",
     "丈夫定妆参考图",
     "全片最地道一句(本地化agent原话)，人称改her(美国争宠官司不用it)，3 A.M.更口语"],

    ["S6","0:21-0:22","1s","观众席快切","(v3微调) 观众爆笑，一人拍腿(法庭秀观众以笑为主，不起哄)",
     "(无台词)",
     "(无台词)",
     "观众爆笑+拍腿声",
     "AI(群演)",
     "群演参考图",
     "OHHHH起哄偏另一档节目风格，爆笑更贴法庭秀"],

    ["S7","0:22-0:27","5s","法官特写->拉镜","法官连敲法槌，全场安静，法官看向下方幽飘",
     'Judge: "ENOUGH. I have ONE ruling... the creature CHOOSES."',
     "法官: 够了。我只有一个判法……让它自己选。",
     "槌声后全场骤静，低音提琴紧张铺底起",
     "AI(法官)",
     "法官定妆参考图",
     "(v3改法官腔) 原This court has不是法官说话方式，改第一人称I have。本句是全片枢纽，口型单独精修"],

    ["S8","0:27-0:29","2s","三方特写快切(妻->夫->幽飘)","(v3压缩: 3s->2s) 妻子期待脸->丈夫咽口水->幽飘面无表情。叠屏幕文字",
     "屏幕文字: Who will it choose?",
     "屏幕文字: 它会选谁？",
     "心跳声渐强，每切一次加重",
     "AI(人类特写) + DCC(幽飘特写) + AE(屏幕文字)",
     "同前资产",
     "(v3) S8-S10原连续10s无台词是全片最大划走点，整体压至6s并加屏幕文字给静音用户留驻理由"],

    ["S9","0:29-0:31","2s","低角度跟拍","(v3压缩: 4s->2s) 幽飘跳下椅子，小短腿走向法庭中央，正常速度+一声滑稽配乐(不用慢镜头)",
     "(无台词)",
     "(无台词)",
     "放大脚步声+滑稽配乐点",
     "DCC(幽飘move动作) + AE(合成)",
     "动作: 同目录move.fbx(先确认是走是跑，速度可调)",
     "慢镜头+无台词是付费流量大忌，压缩保信息密度"],

    ["S10","0:31-0:33","2s","特写x2","(v3压缩: 3s->2s) 幽飘抬头看妻子(特写)->转头看丈夫(特写)->停顿半秒",
     "(无台词)",
     "(无台词)",
     "音乐戛然抽空，只剩环境底噪",
     "DCC(幽飘头部动画，idle基础上手K转头)",
     "同前资产",
     "二选一抉择构图，观众代入'会选谁'"],

    ["S11","0:33-0:37","4s","中景微仰拍，幽飘背后视角","(v3改遮挡剪辑) 幽飘转身——0.3s身体遮挡帧中捕捉球出现于爪中——继续转身对准法官席。全场凝固拆为2-3个单人反应特写快切(规避多角色同框脸漂移)",
     "(无台词)",
     "(无台词)",
     "所有音乐抽空+球体展开机械轻响(承担50%信息量)",
     "DCC(幽飘转身，手K举球转身而非掏球) + AE(球道具+合成) + AI(单人反应特写)",
     "捕捉球: UI/C_宠物捕捉.psd提取(2.08GB，提前预留提取时间)",
     "无掏球动作资产且小短腿掏背后必穿模，遮挡剪辑是影视魔术标准cheat。球体须游戏自有设计，非红白配色"],

    ["S12","0:37-0:41","4s","全景+特效镜头","(v3重做收服仪式) 光束从球中卷出包裹法官(卡通光效，不做痛苦挣扎)->法官被吸入球中(AI冻结帧+AE粒子消散)->读条UI在球上方闪现->球体闭合，震动一次",
     "(无台词)",
     "(无台词)",
     "buzhuo1(捕捉吸入·游戏原声)->buzhuo2(捕捉读条·游戏原声)->闭合震动+原创确认音(用battle_win变调合成或音效库)",
     "AE(吸入光效+法官消散粒子+读条UI+球体震动)",
     "音效: Audios/audio_sfx_buzhuo1.mp3、audio_sfx_buzhuo2.mp3 | 场景生成阶段必须同步输出无法官空景版本",
     "(v3核心修改) 弃用晃三下+叮(任天堂trade dress高风险)，改用游戏真实收服表现(吸入+读条)，既合规又机制真实。消散必须有空景，否则擦人返工"],

    ["S13","0:41-0:45","4s","中景缓推","死寂中法警缓缓站起，面无表情，从腰带解下自己的捕捉球",
     'Bailiff: "...Everyone\'s catchable." (停顿一拍) "Court\'s adjourned."',
     "法警: ……万物皆可收。(停顿) 现在休庭。",
     "法警低音独白+尾音低沉鼓点",
     "AI(法警) + AE(球道具)",
     "法警定妆参考图",
     "deadpan补刀=全片记忆点; (v3新增) Court's adjourned回收法庭格式再叠一层笑点，零成本"],

    ["S14","0:45-0:47","2s","黑屏字幕","硬切黑屏，白色打字机字体逐行砸出",
     "字幕: Catch pets. Catch feelings. Catch... everyone.",
     "抓宠物。抓走心。抓……所有人。",
     "每行字幕配一声打字机敲击",
     "AE(字幕动效)",
     "—",
     "(v3改) 原Catch bosses/everything是无机制支撑的功能承诺(抓Boss/抓一切游戏均不支持)，改玩笑语气并callback法官被收+与咒语押韵"],

    ["S15","0:47-0:50","3s","结尾卡两段式","(v3改: UI露出0.5s->1.5s) 0:47-0:48.5游戏捕捉界面UI定格(看得清这是手游)->0:48.5-0:50 Logo+下载按钮+CTA。底部全程小字免责声明",
     "CTA A版: Who Will You Catch First? | B版: Your SSR Is Waiting — Play Free | 小字: Dramatization. Not actual gameplay footage.",
     "A版: 你会先抓谁？ | B版: 你的SSR在等你——免费开玩 | 小字: 戏剧演绎，非真实游戏画面",
     "轻快上扬收尾音",
     "AE(结尾卡) + 游戏UI截图",
     "UI: UI/C_宠物捕捉.psd | Logo: 图标发行PNG12.3里无Logo，需向发行/美术索取正式文件",
     "0.5s低于有意识识别阈值，1.5s才能承接转化。CTA双版AB测试。免责声明是FTC减责因素+平台申诉筹码"],
]
for s in shots:
    ws2.append(s)

# ============ Sheet 3: 资产清单 ============
ws3 = wb.create_sheet("资产清单")
ws3.append(["资产类型","名称","路径/来源","用途","制作方式"])
assets = [
    ["角色模型(v3修正)","幽飘10111 mod_10111_high.fbx(+同名.max)",
     "Y:/市场运营部/友蜜/幻想宠物/物料&资产/UE资产/角色模型和动作/宝可梦资产汇总/宝可梦动作文件汇总(1)/宝可梦动作文件汇总/帕基/木少女一阶10111/",
     "S2/S8/S9/S10/S11 帕基全部镜头",
     "DCC渲染(FBX导入UE5或Blender，材质用tex_10111_high_d/m/n.png三张重建，S2先做pipeline test定管线) 进制作前必须核实模型来源(目录名含宝可梦)"],

    ["角色贴图(v3新增)","tex_10111_high_d/m/n.png(diffuse/mask/normal)",
     "Y:/市场运营部/友蜜/幻想宠物/物料&资产/UE资产/角色模型和动作/宝可梦资产汇总/角色模型max+贴图/角色模型max+贴图/英雄模型/木少女1阶10111/tex/",
     "模型材质重建",
     "随模型导入"],

    ["动作(v3修正路径)","move.fbx(S9行走)、idle(S10转头基础)、anger/appear/attack/die/vertigo备用",
     "与模型同目录(宝可梦动作文件汇总(1)/宝可梦动作文件汇总/帕基/木少女一阶10111/)",
     "S9/S10/S11",
     "move先确认走/跑; S10转头手K; S11举球转身手K(无掏球动作，遮挡剪辑方案)"],

    ["立绘参考","11【幽飘】10111WoodGril.png",
     "Y:/市场运营部/友蜜/幻想宠物/物料&资产/平面物料/帕基截图/",
     "合成时角色光影/配色参照",
     "参考用"],

    ["UI素材","C_宠物捕捉.psd(2.08GB，提前预留提取时间)",
     "Y:/市场运营部/友蜜/幻想宠物/物料&资产/平面物料/UI/",
     "捕捉球设计提取(须游戏自有设计，非红白球)、S15结尾UI露出",
     "AE提取球体分层"],

    ["音效(v3修正)","audio_sfx_buzhuo1.mp3(捕捉吸入)、audio_sfx_buzhuo2.mp3(捕捉读条)、audio_sfx_born(孵化/出生)、audio_sfx_battle_win.mp3(变调合成确认音备用)",
     "Y:/市场运营部/友蜜/幻想宠物/物料&资产/平面物料/Audios/",
     "S11/S12收服仪式; born佐证S3 I HATCHED her机制真实",
     "直接调用。不存在收服成功音，S12确认音需battle_win变调合成/游戏内实录/音效库三选一，立项第一天定方案"],

    ["Logo(v3修正)","游戏Logo正式文件",
     "图标发行PNG12.3里只有道具/技能图标，无Logo——需向发行/美术索取",
     "S15结尾卡",
     "AE排版"],

    ["AI人物定妆","法官(避开Sheindlin特征: 灰白短发/非蕾丝领/低沉嗓音)/妻子/丈夫/法警/群演x3-5",
     "AI生成(角色一致性模式，提示词禁出现真人姓名和节目名)",
     "全部人类镜头",
     "定妆图锁定->每镜3-5抽; 法官戏份最多投入最大抽卡预算; 法官定妆出图后与Judy实际形象对比留档"],

    ["AI场景(v3新增要求)","美式法庭全景+各角度+【无法官空景版本】",
     "AI生成",
     "S0-S13背景; 空景为S12消散必备",
     "一个法庭多角度复用，空景必须在场景生成阶段同步输出，否则S12返工"],

    ["小道具","幽飘领结、名牌字幕条(KAREN THOMPSON—Plaintiff / DOUG THOMPSON—Defendant)",
     "AE绘制",
     "领结S2起全程; 字幕条S3/S5",
     "AE跟踪贴合/字幕动效"],

    ["风格统一(v3新增)","胶片颗粒+暗角+统一LUT",
     "AE预设",
     "帕基渲染端与AI场景统一显影，解决画风撕裂",
     "渲染端加颗粒暗角，AE端统一LUT+接触阴影"],
]
for a in assets:
    ws3.append(a)

# ============ Sheet 4: V3修改记录 ============
ws4 = wb.create_sheet("V3修改记录")
ws4.append(["#","级别","镜头/位置","问题(审查发现)","修改内容"])
changes = [
    [1,"阻断","S3台词","虚构喂食系统+虚构0.01%刷新率数值(游戏均无)，FTC事实声明风险",
     "重写为融合狗粮+孵化+进化三真实机制: he used her as FUSION FODDER. I HATCHED her..."],

    [2,"阻断","项目信息","幽飘是新手线宠物(10111)，0.01%神宠设定与游戏定位矛盾",
     "控诉逻辑从'它多稀有'改为'我付出多少'"],

    [3,"阻断","S12收服仪式","晃三下+叮为任天堂trade dress级表达，高风险",
     "重做: 光束卷入(buzhuo1)->读条UI(buzhuo2)->闭合震动一次+原创确认音(=游戏真实收服表现)"],

    [4,"阻断","资产清单","3处事实错误: 动作路径层级错、UE资产实为Max/FBX、收服成功音和Logo不存在",
     "全部修正为实测路径; 确认音给3个备选方案; Logo标注向发行索取"],

    [5,"效果","S0","黑屏首帧=无钩子，静音用户无感知",
     "首帧直接上法庭全景+大字钩子: They agree on everything... except ONE."],

    [6,"效果","S8-S10","连续10s无台词，全片最大划走点",
     "压至6s(S9砍慢镜头)，叠屏幕文字Who will it choose?"],

    [7,"效果","S15","0.5s UI露出低于识别阈值，转化最弱环",
     "延长到1.5s定格; CTA双版AB; 加Dramatization免责声明"],

    [8,"效果","全片","人称it削弱抚养权parody前提",
     "幽飘全片改her"],

    [9,"效果","S3/S5","缺法庭真人秀最强格式签名: 当事人名牌字幕条",
     "新增chyron: KAREN THOMPSON—Plaintiff / DOUG THOMPSON—Defendant"],

    [10,"效果","S3表演","声泪俱下像肥皂剧不像法庭秀，踩刻板印象边缘",
     "改ice-cold冷怒毒舌，台词不变"],

    [11,"效果","S7台词","This court has不是法官腔",
     "改ENOUGH. I have ONE ruling... the creature CHOOSES."],

    [12,"效果","S14字幕","Catch bosses/everything是无机制支撑的功能承诺",
     "改Catch pets. Catch feelings. Catch... everyone."],

    [13,"效果","S13","可再回收格式叠一层笑点",
     "法警补Court's adjourned."],

    [14,"效果","合规动作","AIGC标注强制; 文档禁出现Judge Judy/宝可梦字样(故意侵权证据)",
     "写入项目信息合规清单5条"],

    [15,"制作","S11","掏球动作不可行(无资产+穿模)",
     "改遮挡剪辑: 0.3s身体遮挡帧球出现"],

    [16,"制作","S12","法官消散需空景否则返工; 全场凝固多角色同框脸漂移",
     "空景列入场景生成清单; 凝固拆单人反应快切"],

    [17,"制作","全片","卡通帕基+写实AI法庭画风撕裂",
     "渲染端颗粒暗角+AE统一LUT+接触阴影; S2做pipeline test"],

    [18,"制作","15s短版","原方案砍丢S13咒语",
     "重排保Everyone's catchable"],

    [19,"阻断","资产来源","模型目录名含宝可梦，来源存疑",
     "进制作前必须向资产团队核实，写入项目信息警示"],
]
for c in changes:
    ws4.append(c)

# ============ 样式 ============
thin = Border(*[Side(style='thin', color='CCCCCC')]*4)
hdr_fill = PatternFill('solid', fgColor='2F5597')
hdr_font = Font(bold=True, color='FFFFFF', size=11)
red_fill = PatternFill('solid', fgColor='FFD9D9')
warn_fill = PatternFill('solid', fgColor='FFF3CD')

sheet_widths = [
    (ws1, [16, 120]),
    (ws2, [6,11,6,18,42,44,34,26,28,42,44]),
    (ws3, [16,32,54,32,42]),
    (ws4, [4,9,12,44,52]),
]
for ws, widths in sheet_widths:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for c in row:
            c.border = thin
            if c.row > 1:
                c.alignment = Alignment(vertical='top', wrap_text=True)
    ws.freeze_panes = 'A2'

# 修改记录sheet: 阻断级红色、效果级黄色
for r in range(2, ws4.max_row + 1):
    level = str(ws4.cell(r, 2).value or '')
    if '阻断' in level:
        for cc in range(1, 6):
            ws4.cell(r, cc).fill = red_fill
    elif '效果' in level:
        for cc in range(1, 6):
            ws4.cell(r, cc).fill = warn_fill

out = r"D:/claude-projects/projects/幻宠/素材/分镜/PET_COURT_宠物法庭_分镜_v3.xlsx"
wb.save(out)
print("OK:", out)
print("Sheets:", wb.sheetnames)
print("分镜镜头:", ws2.max_row - 1)
print("资产条目:", ws3.max_row - 1)
print("修改记录:", ws4.max_row - 1)
