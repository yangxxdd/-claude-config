# -*- coding: utf-8 -*-
"""为 PET COURT v3 追加「竞品脚本分析」Sheet"""

import openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"D:\claude-projects\projects\幻宠\素材\分镜\PET_COURT_宠物法庭_分镜_v3.xlsx"
OUT = r"D:\claude-projects\projects\幻宠\素材\分镜\PET_COURT_宠物法庭_分镜_v4.xlsx"  # v3可能被Excel锁住，保存为v4

wb = openpyxl.load_workbook(SRC)

# ========== 通用样式 ==========
thin = Border(*[Side(style='thin', color='CCCCCC')]*4)
hdr_fill = PatternFill('solid', fgColor='2F5597')
hdr_font = Font(bold=True, color='FFFFFF', size=11)
section_fill = PatternFill('solid', fgColor='E2EFDA')
section_font = Font(bold=True, size=12, color='375623')
highlight_fill = PatternFill('solid', fgColor='FFF2CC')
note_font = Font(italic=True, color='666666', size=10)

def style_sheet(ws, col_widths):
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for c in row:
            c.border = thin
            if c.row > 1:
                c.alignment = Alignment(vertical='top', wrap_text=True)
    ws.freeze_panes = 'A2'

# ==========================================
# Sheet 5: 竞品脚本概览
# ==========================================
ws5 = wb.create_sheet("竞品脚本概览")
ws5.append(["脚本名称","作者/角色","核心卖点","时长","制作方式","关键特征","PET COURT可借鉴点"])

overview = [
    ["薇拉的最后三声钟",
     "孟帅（发行制作人）",
     "阶段性强钩子剧情：玩家救薇拉，每完成一段游戏她离死亡更近",
     "—",
     "AI+UE",
     "6分镜 · 连续剧式钩子 · '刀将落下时切断'悬念 · 深游戏世界观融合",
     "⚠️ 过于叙事化: 更像游戏剧情设计文档而非广告脚本，不适合直接交付美术。但'切断悬念'技法值得借鉴→PET COURT S7 creature CHOOSES即同类悬念点"],

    ["莉莉丝的断头台",
     "孟帅（发行制作人）",
     "玩家发现百年前处决母亲的人可能是自己",
     "—",
     "AI+UE",
     "7分镜 · 母子情感线 · '真相递进'反转 · 玩家→处刑者身份悬念",
     "⚠️ 同样过于复杂。可借鉴: '切断点'设计——PET COURT每个反转镜头末帧也是天然切断点，可做连续剧多集投放测试"],

    ["怪物新娘",
     "孟帅（发行制作人）",
     "救下一个外表畸变但保留理智的吸血鬼新娘",
     "—",
     "AI+UE",
     "7分镜 · 道德困境 · '外表vs内心'反转 · 被咬出寄生虫机制露出",
     "⚠️ 三案中最具广告潜力: 怪物外表→温柔内心的反差结构，对应PET COURT幽飘(可爱外表→收服法官的反转)。'反常救援'钩子可迁移"],

    ["吸血鬼追逐",
     "陈永杰（运营）",
     "血月下贵族逃亡→被神秘人拉进酒馆",
     "~13s",
     "AI+UE",
     "5分镜(未完) · 快节奏 · 低机位 · 氛围强 · 少台词",
     "✅ 节奏感: 每镜2-3s快切，适合竖屏短视频。但缺CTA和完整收尾。PET COURT S8-S10已采用类似快切节奏"],

    ["吸血鬼城镇 ⭐",
     "孔祥建（投放/UA）",
     "人类误入永夜镇→被当Special Guest→食物是人体",
     "30-35s",
     "AI+UE(分镜标注)",
     "5分镜 · 参考视频3个 · BGM指定 · 每镜标注AI/UE · 场景有PSD文件 · 尺寸时长明确",
     "⭐⭐⭐ 领导肯定。核心优势见下方专项分析"],

    ["救男吸血鬼(AI) ⭐",
     "孔祥建（投放/UA）",
     "女店主救助吸血鬼→获吸血鬼领主回报升级餐馆",
     "~35s",
     "AI(全AI)",
     "5分镜 · 交互选项覆盖 · 反转(救→被救→回报) · 变身特效 · 金币数字飞涨 · 女性视角",
     "⭐⭐⭐ 领导肯定。核心优势见下方专项分析"],
]

for row in overview:
    ws5.append(row)

style_sheet(ws5, [18, 18, 36, 12, 16, 52, 52])

# ==========================================
# Sheet 6: UA脚本优势深度分析
# ==========================================
ws6 = wb.create_sheet("UA脚本优势分析")
ws6.append(["维度","优势点","脚本实例","PET COURT现状","差距/改进建议"])

analysis = [
    # ---- 交付可执行性 ----
    ["🎯 交付可执行性",
     "分镜表=美术工单，可直接下发无需口头补充",
     "吸血鬼城镇: 每个分镜标注AI/UE，场景(A4)写'游戏内永夜镇全景(有psd文件)'，资产路径已有",
     "PET COURT v3已做到制作方式标注+资产路径，但场景/角色的视觉参考仍以文字为主",
     "🟡 建议: 每个角色和场景配参考图(可AI生成)，见「Gemini生图提示词」Sheet"],

    ["🎯 交付可执行性",
     "BGM/音效/特效具体到文件名或游戏名",
     "吸血鬼城镇: BGM='中世纪阴森，参考Last Asylum_ Plague游戏'；救男: BGM='先阴暗，后轻松'",
     "v3已有音效文件名(buzhuo1/2.mp3)和风格描述，但BGM仅说'罐头授权'未给具体参考",
     "🟡 建议: 指定BGM参考曲风+tempo范围，便于音效师快速定位"],

    # ---- 参考驱动 ----
    ["🎯 参考驱动",
     "每个脚本都附了参考视频/图片文件名",
     "吸血鬼城镇: 附3个mp4参考视频(Guns of Glory, Last Asylum)；救男: 附Game of Vampires视频",
     "v3资产清单有路径，但分镜表里'参考: 法庭真人秀类型通用范式'这种引用太抽象",
     "🔴 关键差距! v3缺少具体视觉参考。建议每个关键镜头附竞品截图或AI生成概念图→见「参考素材清单」Sheet"],

    ["🎯 参考驱动",
     "参考物具体到文件名，美术可立刻找到并打开",
     "Last Asylum_ Plague-2026-05-28-4ac6ebb8...mp4 带完整文件名和时间戳，说明实际看过",
     "v3资产路径已做到文件名级别，但参考视频未附",
     "🟡 建议: 搜集3-5个竞品法庭/反转类广告视频作为参考→见「参考素材清单」Sheet"],

    # ---- 交互/参与感 ----
    ["🎯 交互/参与感",
     "覆盖交互选项UI，给观众'我在玩游戏'的错觉",
     "救男: S6出现'Ignore Him / Help Him'选项+手指点击；S8出现'汉堡/鲜血'选项+手指点击",
     "PET COURT v3无交互UI覆盖",
     "🔴 可借鉴! 可在S7(CREATURE CHOOSES)叠加选项按钮'Choose Wife / Choose Husband'然后被幽飘无视→增强参与感+多一层笑点"],

    ["🎯 交互/参与感",
     "选项制造'选择幻觉'——观众代入决策者角色",
     "救男: 每个选项都用手指标注点击，视觉上像玩家在操作游戏",
     "PET COURT纯线性叙事",
     "🟡 非必须但加分。短版(15s)不加，长版(50s)可在1处加入"],

    # ---- 视觉对比/变身 ----
    ["🎯 视觉对比/变身",
     "强Before/After对比是短视频最高效的满足感来源",
     "救男S8: 吸血→特效→华丽服装变身+The Vampire Lord字幕；S10: 破旧餐馆→豪华酒馆+金币飞涨",
     "PET COURT无视觉对比/变身设计",
     "🔴 关键! 法庭包豪斯→被收服后可加1s法庭外观'被球砸出大洞'或法官席空了的视觉反差→0成本增加satisfaction"],

    # ---- 简单钩子 ----
    ["🎯 简单钩子",
     "不需要了解任何游戏世界观即可理解剧情",
     "吸血鬼城镇: 人类进吸血鬼小镇→被当食物；救男: 女救男→男变吸血鬼→回报",
     "PET COURT: 离婚争宠→法官让宠选→宠收法官。理解门槛: 需知道离婚+法庭+养宠三个概念",
     "✅ PET COURT理解门槛合理(美国家庭法+养宠=大众认知)，但S0钩子可进一步简化: 首帧直接'They're divorcing. Over a PET.'比'They agree on everything... except ONE'更直给"],

    # ---- 进度/数字增长 ----
    ["🎯 进度/数字增长",
     "金币/数字增长是IAP游戏核心满足感视觉化",
     "救男S10: '+1000 +5000 +10000金币从桌上飞向收银台'",
     "PET COURT无语病增长",
     "🟢 不适合法庭题材，不强行加入。但S15结尾卡可考虑加入'1000+ Creatures Waiting'类数字钩子"],

    # ---- 女性主角 ----
    ["🎯 女性主角",
     "女性POV提升女性用户代入感和转化",
     "救男: 女主角全程POV，叹气→救助→被回报→最后抱住吸血鬼",
     "PET COURT: 男女双方均有戏份，幽飘(her)是情感中心",
     "✅ PET COURT避免了纯男性视角，妻子S3/S5戏份相当。幽飘用her人称强化情感连接"],

    # ---- 制作标注 ----
    ["🎯 制作标注",
     "每个镜头明确标注制作方式，美术不用猜",
     "吸血鬼城镇: S4标注'AI', S6标注'AI', S8标注'UE', S10标注'UE/AI'",
     "v3分镜表已有'制作方式'列",
     "✅ 已对齐。但可更精细: v3的AI镜头未区分'静态AI生图+AE动效'vs'AI视频生成'"],

    # ---- 简单反转 ----
    ["🎯 简单反转",
     "反转只需1步，观众3秒内理解",
     "救男: 救→他是吸血鬼(1步)；吸血鬼城镇: 进餐厅→上的是人肉(1步)",
     "PET COURT: 3层反转(争宠→宠选→宠收法官)，结构复杂但每层独立清晰",
     "✅ 3层反转在50s长版里节奏合理。30s短版砍为2层(争宠→宠收法官)，15s只保punchline"],

    # ---- 情感升级 ----
    ["🎯 情感升级",
     "从problem → solution → reward有完整情感曲线",
     "救男: 穷(problem)→救吸血鬼(solution)→变身回报(reward)→抱在一起(emotional payoff)",
     "PET COURT: 离婚(problem)→宠选(solution?)→被收服(punchline≠reward)",
     "✅ PET COURT走的是'反转punchline'路线非reward路线，差异化定位。但S13法警Everyone's catchable补了幽默reward感"],
]

for row in analysis:
    ws6.append(row)

style_sheet(ws6, [18, 40, 52, 48, 48])

# 高亮标记行
for r in range(2, ws6.max_row + 1):
    dim = str(ws6.cell(r, 1).value or '')
    gap = str(ws6.cell(r, 5).value or '')
    if gap.startswith('🔴'):
        for cc in range(1, 6):
            ws6.cell(r, cc).fill = PatternFill('solid', fgColor='FFD9D9')
    elif gap.startswith('🟡'):
        for cc in range(1, 6):
            ws6.cell(r, cc).fill = PatternFill('solid', fgColor='FFF3CD')

# ==========================================
# Sheet 7: 参考素材清单
# ==========================================
ws7 = wb.create_sheet("参考素材清单")
ws7.append(["类别","素材名称/描述","来源/路径","对应镜头","备注"])

refs = [
    # ---- 竞品广告参考 ----
    ["竞品视频","Last Asylum: Plague — 中世纪阴森氛围+建筑参考",
     "feishu: 吸血鬼城镇脚本L4 (Last Asylum_ Plague-2026-05-28-4ac6ebb8....mp4)",
     "S0/S1 法庭场景氛围",
     "吸血鬼城镇脚本引用的BGM+场景参考，中世纪哥特建筑风格可迁移至法庭外景延展"],

    ["竞品视频","Guns of Glory: Lost Island — 吸血鬼角色造型+动作参考",
     "feishu: 吸血鬼城镇脚本L4 (Guns of Glory_ Lost Island-2026-05-28-9e279004....mp4 ×2)",
     "S4/S6 观众反应、角色造型",
     "角色设计参考: 吸血鬼造型如何在广告中呈现而不致恐怖"],

    ["竞品视频","Game of Vampires: Twilight Sun — 吸血鬼+言情+选择互动格式",
     "feishu: 救男脚本L4 (Game of Vampires_ Twilight Sun-2026-05-15-6d2f3e45....mp4)",
     "全片互动设计",
     "交互选项覆盖格式的完整参考，宠物法庭如加入互动元素可参考此片节奏"],

    ["竞品视频","Palmon: Survival 广告 — 幻宠直接竞品",
     "WebSearch: Palmon Survival ads 2025/2026 (Lilith)",
     "全片素材策略",
     "我们的直接竞品! 需搜集其最新素材分析钩子公式→列入后续行动"],

    # ---- 法庭真人秀参考 ----
    ["类型参考","美式法庭真人秀通用格式特征",
     "公开知识: courtroom reality TV genre tropes — 名牌字幕条/法官居中/当事人站讲台/观众反应镜头/法槌节奏",
     "S3/S5/S7/S13",
     "v3已内化这些特征。注意: 参考的是格式范式(genre tropes)，不是任何具体节目"],

    # ---- 游戏资产 ----
    ["游戏资产","幽飘(10111 WoodGril) — 立绘+模型+动作",
     "Y:/市场运营部/友蜜/幻想宠物/物料&资产/UE资产/角色模型和动作/.../木少女一阶10111/",
     "S2/S8/S9/S10/S11",
     "⚠️ 来源目录含'宝可梦'字样，进制作前必须核实。立绘见 帕基截图/11【幽飘】10111WoodGril.png"],

    ["游戏资产","捕捉球UI — C_宠物捕捉.psd (2.08GB)",
     "Y:/市场运营部/友蜜/幻想宠物/物料&资产/平面物料/UI/",
     "S11/S12/S15",
     "提取球体设计做道具参考，须游戏自有设计(非红白球)"],

    ["游戏资产","捕捉音效 — buzhuo1.mp3 / buzhuo2.mp3 / battle_win.mp3",
     "Y:/市场运营部/友蜜/幻想宠物/物料&资产/平面物料/Audios/",
     "S11/S12",
     "buzhuo1=吸入 buzhuo2=读条 battle_win=变调合成确认音 不存在独立收服成功音"],

    # ---- PET COURT已搜集参考 ----
    ["素材数据","幻宠42素材全量分析 — P:V对比、方向标签、9月推荐",
     "memory: [[huanchong-creative-full-analysis]]",
     "策略决策",
     "捉宠方向CPI最优、融合/进化方向黑马。PET COURT以捉宠为核心露出+进化/孵化/融合口播"],

    ["素材数据","6/11测试KPI — 留存/漏斗/渠道数据",
     "memory: [[huanchong-611-test]]",
     "转化预期",
     "为PET COURT投放效果预估提供基线"],

    ["素材数据","素材方向与钩子公式 — 5种已验证钩子",
     "memory: [[huanchong-creative-materials]]",
     "钩子验证",
     "PET COURT对应'沙雕搞笑+反差打脸'双钩子公式"],

    ["制作参考","灵画师AI对话类广告 — 熟悉面孔+求救+反转格式",
     "本地: C:\\Users\\yangxd\\Downloads\\20260806-183510.mp4 / 20260806-183518.mp4",
     "全片格式启发",
     "用户6天前提供的国内参考——AI对话+结尾反转，启发了'格式寄生'策略但不直接模仿"],
]

for row in refs:
    ws7.append(row)

style_sheet(ws7, [14, 42, 52, 22, 44])

# ==========================================
# Sheet 8: Gemini生图提示词
# ==========================================
ws8 = wb.create_sheet("Gemini生图提示词")
ws8.append(["编号","用途","对应镜头","Gemini (Imagen) Prompt","中文说明","备注"])

prompts = [
    # ---- 法官 ----
    ["G01","法官定妆参考图",
     "S0/S1/S7",
     "A 55-year-old female judge with sharp gray eyes and short silver-white hair in a clean modern bob cut, wearing a simple black judge's robe with a plain white collar (no lace, no ruffles). She has an authoritative but dry-humored expression, one eyebrow slightly raised. Professional headshot style against a dark wood-paneled courtroom background. Medium shot, soft dramatic lighting from the side, photorealistic, 9:16 vertical composition, cinematic color grading with subtle warm tones.",
     "法官: 55岁女性，银白短发(非波波头)，黑法官袍+白领(无蕾丝)，挑眉微讽表情。⚠️已主动规避Sheindlin的深色波波头/蕾丝领/尖细嗓音等签名特征。出图后与Judy实际照片对比留档。"],

    ["G02","法官正面特写(法槌时刻)",
     "S7",
     "Same female judge as described, now leaning forward with her gavel raised mid-air. Close-up on her face and hand, intense but controlled expression, mouth open mid-speech saying 'ENOUGH'. The wooden gavel is sharp in foreground. Courtroom blurred background with warm amber lighting. Slight motion blur on the gavel to suggest it's about to strike. Cinematic, photorealistic, 9:16 vertical.",
     "法官举槌瞬间: 身体前倾，槌在半空，嘴微张说'ENOUGH'。前景槌清晰+背景虚化。"],

    # ---- 幽飘 ----
    ["G03","幽飘(可爱生物)参考",
     "S2/S8/S9/S10",
     "A cute small round fantasy creature with large innocent eyes, tiny stubby arms and legs, leaf-like ears, and a soft green-brown woodland color palette. It wears a tiny black bow tie. It sits on a wooden defendant's chair inside a courtroom, looking up with a blank innocent expression. The chair is too big for it, making it look comically small. Soft diffused lighting, Pixar-meets-Ghibli art style, 3D rendered look, 9:16 vertical, shallow depth of field.",
     "幽飘(圆润小精灵): 大眼+小短腿+叶耳+绿棕色系。戴小领结，坐大木椅上。皮克斯×吉卜力画风，3D渲染感。⚠️不能像皮卡丘/任何宝可梦。"],

    ["G04","幽飘正面面部特写",
     "S10",
     "Extreme close-up of the same small round fantasy creature's face. Large expressive eyes looking slightly upward, tiny nose, leaf-like ears slightly drooping. Blank, unreadable expression — the 'who will it choose?' suspense moment. The bow tie is visible at the bottom edge. Soft studio lighting with catchlights in the eyes, shallow depth of field blurring the courtroom background, photorealistic 3D style, 9:16 vertical.",
     "幽飘面部大特写: 眼神向上看，表情空白不可读(制造S10选择悬念)。眼神catchlight。"],

    # ---- 妻子 ----
    ["G05","妻子(Karen Thompson)定妆",
     "S3",
     "A polished American woman in her early 30s, sharp features, wearing a tailored navy blue blazer over a cream blouse. She has medium-length dark hair pulled back in a low ponytail. Her expression is ice-cold controlled fury — lips pressed thin, eyes narrowed but not crying, one hand resting on the courtroom lectern. She looks like she could verbally destroy someone. Professional courtroom setting background. Photorealistic, 9:16 vertical, cool color temperature.",
     "妻子(凯伦): 30出头美国白人女性，精致海军蓝西装+米色衬衫。深色中长发低马尾。表情=ice-cold冷怒毒舌(不哭!)。⚠️避免'歇斯底里妻子'刻板印象。"],

    # ---- 丈夫 ----
    ["G06","丈夫(Doug Thompson)定妆",
     "S5",
     "An American man in his early 30s, slightly disheveled plaid shirt with rolled-up sleeves, tie loosened. He has short brown hair, 5 o'clock shadow, defensive posture with both hands on the courtroom lectern. His expression is indignant and wounded — mouth open mid-protest, eyebrows furrowed. He looks like a guy who has been wronged and is about to explode with his side of the story. Photorealistic, 9:16 vertical, warm lighting.",
     "丈夫(道格): 30出头美国男性，格子衫+卷袖+松开领带。棕短发+胡茬。双手撑讲台，张嘴抗议，委屈+愤怒交织。"],

    # ---- 法警 ----
    ["G07","法警(Bailiff)定妆",
     "S13",
     "A tall African American male bailiff in his 40s, wearing a crisp tan sheriff's deputy uniform with a badge. He has a completely deadpan expression — stone-faced, unreadable, the ultimate straight man. He stands at the side of the courtroom, one hand resting on his duty belt. Medium-full shot. The humor comes from how utterly serious he looks in this absurd situation. Slight low angle to give him authority. Photorealistic, 9:16 vertical, neutral lighting.",
     "法警: 40多岁非裔美国男性，穿米色警长制服+警徽。完全deadpan石头表情=本片喜剧担当。低角度增加权威感。"],

    # ---- 法庭场景 ----
    ["G08","美式法庭全景(带法徽)",
     "S0",
     "A full American courtroom interior viewed from the back. Dark wood paneling, raised judge's bench center-back with a custom round emblem above it (the emblem features a stylized capture-ball silhouette instead of a traditional seal — subtle, elegant, fantasy-legal fusion). Two lecterns facing the bench, wooden spectator benches. Warm amber lighting from wall sconces, slightly dramatic shadows. No people visible. Photorealistic architectural photography style, wide angle, 9:16 vertical.",
     "法庭全景: 深木色调+法官席居中。法徽原创设计(嵌捕捉球剪影替代传统鹰徽)。暖琥珀色调+壁灯。无人空景。⚠️避免像任何真实法庭/电视剧场景。"],

    ["G09","法庭空景(无法官)",
     "S12",
     "Same courtroom interior but the judge's bench is conspicuously empty. The large leather judge's chair is vacant, slightly askew. The gavel lies alone on the bench. Same warm amber lighting but now with a slight eerie quality — dust motes visible in the light beams. The capture-ball emblem above the bench is subtly glowing. The emptiness should feel both funny and slightly magical. Wide shot, photorealistic, 9:16 vertical.",
     "法庭空景(法官消散后): 法官椅空着微歪+法槌孤立。光束中可见灰尘。法徽微发光。氛围=又好笑又魔幻。⚠️v3要求场景生成阶段必须同步输出空景版!"],

    # ---- 观众群演 ----
    ["G10","法庭观众席反应(多样性)",
     "S4/S6",
     "A row of diverse American courtroom spectators in the gallery bench — a mix of ages (30s-60s), ethnicities, and genders. Currently reacting with shock: one older woman has her hand over her mouth, eyes wide; a middle-aged man next to her is mid-gasp; a younger person is leaning forward with eyebrows raised. All wearing casual to business-casual attire. Shallow depth of field focused on the two most expressive faces. Photorealistic, 9:16 vertical, warm natural light.",
     "观众席群演: 年龄/种族/性别混合。当前反应=震惊(捂嘴/倒吸气/前倾)。浅景深聚焦2-3张最生动的脸。⚠️多样性不是贴标签，是真实法庭观众构成。"],

    # ---- 捕捉球道具 ----
    ["G11","捕捉球道具设计参考",
     "S11/S12",
     "A fantasy capture device — a metallic sphere about the size of a baseball, with intricate mechanical engravings and a subtle blue-green energy glow emanating from the seam lines. It looks like a fusion of steampunk craftsmanship and magical technology. The sphere is partially open with a soft light beam emerging from the center. NOT red and white color scheme. NOT a Poké Ball. Clean product-shot style on a dark gradient background, 3D rendered, photorealistic materials (brushed metal, glass, energy light).",
     "捕捉球: 棒球大小金属球，精密机械纹路+蓝绿能量光从缝隙溢出。半开状态+光束。⚠️绝不能用红白配色=任天堂trade dress。须游戏自有设计或原创。"],

    # ---- 情景镜头 ----
    ["G12","幽飘走向法庭中央(全身)",
     "S9",
     "The same small round fantasy creature (with bow tie) walking on tiny stubby legs down the center aisle of a grand courtroom. The camera is at floor level, making the creature look both comically small and oddly determined. The wooden floor is polished and reflective. Warm amber courtroom lighting creates long shadows. The empty judge's bench looms in the distant background. Low angle, cinematic composition, 9:16 vertical, slight motion blur on the tiny feet.",
     "幽飘小短腿走向法庭中央: 地板机位+超低角度。木地板反射+温暖长影+远处法官席。关键=又好笑又莫名有气势。"],

    # ---- 片头包装 ----
    ["G13","法庭片头包装(法徽+字幕条)",
     "S0",
     "A broadcast-style lower-third graphic and court seal combination. Top section: a stylized gold circular emblem featuring a simple capture-ball icon in the center, surrounded by elegant geometric patterns (not traditional eagle/lion motifs). Bottom section: a dark gradient text bar with clean white sans-serif text reading 'CASE #4847 — Thompson v. Thompson'. The overall look is premium reality TV court show opening. Clean vector/graphic design style, dark navy and gold color palette, 9:16 vertical format. The design should feel like a premium streaming show, not a cable TV court show.",
     "片头包装: 金色法徽(中嵌捕捉球)+深蓝渐变字幕条+CASE编号。风格=流媒体真人秀质感，不是有线电视法庭秀。纯平面设计，AE可执行。"],

    # ---- S14黑屏字幕 ----
    ["G14","结尾黑屏字幕风格参考",
     "S14",
     "A completely black background with three lines of white typewriter-font text appearing one by one. The text reads: 'Catch pets.' / 'Catch feelings.' / 'Catch... everyone.' The last line has a slight pause and the word 'everyone' is slightly larger than the rest. The typewriter font should feel vintage but clean — think classic American literature paperback, not horror. Minimalist, high contrast, elegant. 9:16 vertical. This will be animated in AE; the image is just the final frame of the sequence.",
     "黑屏+打字机体三行字幕。末行'everyone'稍大。风格=经典美式平装书质感，非恐怖片。最终帧参考，动画在AE中实现。"],

    # ---- 综合场景 ----
    ["G15","法庭全景(含所有人类角色+幽飘)",
     "S7-S10 参考",
     "A wide shot of a courtroom during an active hearing. At center: the female judge (silver-white short hair, black robe) looks down from her raised bench with an expression of barely-suppressed amusement. In the foreground: a small round fantasy creature with a bow tie sits on a large wooden defendant's chair between two lecterns. At the left lectern: a polished woman in navy blazer (Karen). At the right lectern: a man in plaid shirt (Doug). In the corner: a stone-faced bailiff in tan uniform. All attention is on the small creature. The composition should feel like a Renaissance painting but in a modern courtroom. Photorealistic, 9:16 vertical, dramatic chiaroscuro lighting.",
     "法庭全员全景: 法官+幽飘+妻+夫+法警。构图像文艺复兴画作的现代法庭版。明暗对照法(chiaroscuro)打光。⚠️多角色同框AI极易脸漂移→仅作概念参考，实际制片拆单人快切。"],
]

for row in prompts:
    ws8.append(row)

style_sheet(ws8, [6, 16, 10, 72, 52, 44])

# 高亮提示词中的关键警告
for r in range(2, ws8.max_row + 1):
    notes = str(ws8.cell(r, 6).value or '')
    if '⚠️' in notes:
        ws8.cell(r, 6).fill = PatternFill('solid', fgColor='FFF3CD')

# ==========================================
# 保存
# ==========================================
wb.save(OUT)
print("Done:", OUT)
print("Sheets:", wb.sheetnames)
print(f"Total {len(wb.sheetnames)} sheets")

# 验证
wb2 = openpyxl.load_workbook(OUT)
for name in wb2.sheetnames:
    ws = wb2[name]
    print(f"  {name}: {ws.max_row}rows x {ws.max_column}cols")
