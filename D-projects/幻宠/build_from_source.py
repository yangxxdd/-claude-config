"""Build creative table from FB source data — v3 clean."""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

SRC = r"D:\claude-projects\projects\幻宠\fb媒体-源数据.xlsx"
OUT = r"D:\claude-projects\projects\幻宠\幻宠素材表现汇总_4月6月_US.xlsx"

# Column maps (0-indexed). 4月AEO has extra col at idx 3 (投放状态)
MAP_NORMAL = {"name": 2, "imp": 3, "click": 4, "cpm": 5, "ctr_pct": 6, "cvr": 7, "ipm": 8, "inst": 9, "cpi": 10, "spend": 16, "campaign": 26}
MAP_AEO4 =  {"name": 2, "status": 3, "imp": 4, "click": 5, "cpm": 6, "ctr_pct": 7, "cvr": 8, "ipm": 9, "inst": 10, "cpi": 11, "spend": 17, "campaign": 27}

# 6月 sheets don't have campaign col at 26; campaign info at col 18 (smaller format)
# Actually 6月 sheets have 24 cols total, so campaign is at col 18 (idx 17)

def sf(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0

def si(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0

def read_sheet(ws, m, is_6month=False):
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[m["name"]]).strip() if row[m["name"]] else ""
        if not name or name == "None": continue
        inst = si(row[m["inst"]])
        spend = sf(row[m["spend"]])
        cpi = sf(row[m["cpi"]])
        imp = si(row[m["imp"]])
        click = si(row[m["click"]])
        ctr = sf(row[m["ctr_pct"]])
        cpm = sf(row[m["cpm"]])
        cvr = sf(row[m["cvr"]])
        if cpi == 0 and inst > 0: cpi = round(spend / inst, 2)
        rows.append({"name": name, "inst": inst, "spend": spend, "cpi": cpi,
                     "imp": imp, "click": click, "ctr": ctr, "cpm": cpm, "cvr": cvr})
    return rows

def aggregate(rows, bidding):
    g = {}
    for r in rows:
        k = r["name"]
        if k not in g: g[k] = {"name": k, "inst": 0, "spend": 0, "imp": 0, "click": 0}
        g[k]["inst"] += r["inst"]; g[k]["spend"] += r["spend"]
        g[k]["imp"] += r["imp"]; g[k]["click"] += r["click"]

    out = []
    for k, v in g.items():
        inst, spend, imp, click = v["inst"], v["spend"], v["imp"], v["click"]
        cpi = round(spend / inst, 2) if inst > 0 else 0
        ctr = round(click / imp * 100, 2) if imp > 0 else 0
        cpm = round(spend / imp * 1000, 2) if imp > 0 else 0
        cvr = round(inst / click * 100, 2) if click > 0 else 0
        ctype = k[0] if k and k[0] in "VP" else "-"
        out.append({"name": k, "type": ctype, "bidding": bidding,
                    "inst": inst, "spend": round(spend, 2), "cpi": cpi,
                    "imp": imp, "click": click, "ctr": ctr, "cpm": cpm, "cvr": cvr})
    return sorted(out, key=lambda x: x["inst"], reverse=True)

wb_src = load_workbook(SRC, data_only=True)
sns = wb_src.sheetnames

# Read all 4 sheets (by index to avoid encoding issues)
a_install = read_sheet(wb_src[sns[0]], MAP_NORMAL)
a_aeo     = read_sheet(wb_src[sns[1]], MAP_AEO4)
j_install = read_sheet(wb_src[sns[2]], MAP_NORMAL, is_6month=True)
j_aeo     = read_sheet(wb_src[sns[3]], MAP_NORMAL, is_6month=True)

# Aggregate and tag
final = []
for r in aggregate(a_aeo, "AEO"): r["month"] = "4月"; final.append(r)
for r in aggregate(a_install, "Install"): r["month"] = "4月"; final.append(r)
for r in aggregate(j_aeo, "AEO"): r["month"] = "6月"; final.append(r)
for r in aggregate(j_install, "Install"): r["month"] = "6月"; final.append(r)

# Sort: month then bidding then installs desc
order = {("4月","AEO"):0, ("4月","Install"):1, ("6月","AEO"):2, ("6月","Install"):3}
final.sort(key=lambda r: (order.get((r["month"], r["bidding"]), 9), -r["inst"]))

# ============================================================
# Write Excel
# ============================================================
wb = Workbook()
ws = wb.active; ws.title = "素材表现汇总"

hfont = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
hfill = PatternFill(start_color="2972F4", end_color="2972F4", fill_type="solid")
afill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
jfill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
sfill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
afont = Font(name="Microsoft YaHei", size=10, color="C00000")
ifont = Font(name="Microsoft YaHei", size=10, color="1F4E79")
sfont = Font(name="Microsoft YaHei", size=10, bold=True)
wfont = Font(name="Microsoft YaHei", size=9, color="999999", italic=True)
bd = Border(left=Side("thin","D0D0D0"), right=Side("thin","D0D0D0"),
            top=Side("thin","D0D0D0"), bottom=Side("thin","D0D0D0"))
ca = Alignment(horizontal="center", vertical="center")
la = Alignment(horizontal="left", vertical="center", wrap_text=True)

hdrs = ["月份", "素材名称", "类型", "出价方式", "安装数", "花费(USD)", "CPI",
        "展示次数", "点击量", "CTR", "CPM", "CVR", "备注"]
wids = [7, 30, 6, 10, 9, 12, 9, 12, 9, 8, 9, 8, 40]
nfmts = {5: '#,##0', 6: '$#,##0.00', 7: '$#,##0.00', 8: '#,##0', 9: '#,##0',
         10: '0.00%', 11: '$#,##0.00', 12: '0.00%'}

for ci, (h, w) in enumerate(zip(hdrs, wids), 1):
    c = ws.cell(1, ci, h); c.font = hfont; c.fill = hfill; c.alignment = ca; c.border = bd
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 24

cur_m, erow = None, 2
for r in final:
    m = r["month"]
    if m != cur_m:
        cur_m = m
        for c in range(1, len(hdrs)+1):
            ws.cell(erow, c).fill = afill if "4月" in m else jfill; ws.cell(erow, c).border = bd
        ws.cell(erow, 1).value = f"◆ {m}"
        ws.cell(erow, 1).font = Font(name="Microsoft YaHei", size=11, bold=True, color="1F4E79")
        ws.cell(erow, 1).alignment = la
        ws.merge_cells(start_row=erow, start_column=1, end_row=erow, end_column=len(hdrs))
        ws.row_dimensions[erow].height = 20; erow += 1

    is_a = "4月" in m; is_aeo = r["bidding"] == "AEO"
    small = r["inst"] < 20
    rf = (afill if is_a else jfill)
    fnt = afont if is_aeo else ifont
    if small: fnt = wfont
    note_parts = []
    if small: note_parts.append("小样本")
    # Flag CPI spike for cross-month materials
    if not is_a and not is_aeo and r["name"] == "V-天灾重建-长版":
        note_parts.append("CPI较4月$3.77暴涨4x至$15.44, 关注")
    note = "; ".join(note_parts) if note_parts else ""

    # ctr/cvr from source are already percentages (e.g. 2.37 = 2.37%)
    ctr_val = r["ctr"] / 100.0
    cvr_val = r["cvr"] / 100.0

    vals = [m, r["name"], r["type"], r["bidding"],
            r["inst"], r["spend"], r["cpi"],
            r["imp"], r["click"], ctr_val, r["cpm"], cvr_val, note]

    for ci, v in enumerate(vals, 1):
        c = ws.cell(erow, ci, v); c.font = fnt; c.fill = rf; c.border = bd
        c.alignment = ca if ci != len(vals) else la
        if ci in nfmts and isinstance(v, (int, float)): c.number_format = nfmts[ci]
    ws.row_dimensions[erow].height = 20; erow += 1

# Summary rows
for c in range(1, len(hdrs)+1):
    ws.cell(erow, c).fill = sfill; ws.cell(erow, c).border = bd
ws.merge_cells(start_row=erow, start_column=1, end_row=erow, end_column=len(hdrs))
ws.row_dimensions[erow].height = 4; erow += 1

for m in ["4月", "6月"]:
    for b in ["AEO", "Install"]:
        sub = [r for r in final if r["month"]==m and r["bidding"]==b]
        if not sub: continue
        ti = sum(r["inst"] for r in sub)
        ts = round(sum(r["spend"] for r in sub), 2)
        timp = sum(r["imp"] for r in sub)
        tcl = sum(r["click"] for r in sub)
        tcpi = round(ts/ti, 2) if ti>0 else 0
        tctr = round(tcl/timp*100, 2) if timp>0 else 0
        tcpm = round(ts/timp*1000, 2) if timp>0 else 0
        tcvr = round(ti/tcl*100, 2) if tcl>0 else 0

        sv = [m, f"【{b}合计】", "-", b, ti, ts, tcpi, timp, tcl, tctr/100, tcpm, tcvr/100, f"{len(sub)}个素材"]
        for ci, v in enumerate(sv, 1):
            c = ws.cell(erow, ci, v); c.font = sfont; c.fill = sfill; c.border = bd
            c.alignment = ca if ci != len(sv) else la
            if ci in nfmts and isinstance(v, (int, float)): c.number_format = nfmts[ci]
        ws.row_dimensions[erow].height = 20; erow += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}{erow-1}"
ws.conditional_formatting.add(f"G2:G{erow-1}", DataBarRule(start_type="min", end_type="max", color="5B9BD5"))

# Sheet 2
ws2 = wb.create_sheet("数据概览")
items = [
    ("数据源", "fb媒体-源数据.xlsx (FB Ads Manager导出)"),
    ("时间范围", "4月: 4/8-4/13(仅6天,学习期); 6月: 6/1-6/30(全月)"),
    ("", ""),
    ("⚠ 跨月对比注意", "4月仅6天数据,处于campaign学习期,CPI可能偏低; 6月为全月数据包含更多竞争时段。直接对比CPI需谨慎。"),
    ("地区", "仅美国"),
    ("聚合方式", "按素材名称+出价方式合并(同素材多日期/多广告组合并)"),
    ("更新时间", "2026-07-24"),
    ("", ""),
]
for m in ["4月", "6月"]:
    for b in ["AEO", "Install"]:
        sub = [r for r in final if r["month"]==m and r["bidding"]==b]
        if sub:
            ti = sum(r["inst"] for r in sub)
            ts = sum(r["spend"] for r in sub)
            items.append((f"{m} {b}", f"{len(sub)}素材, {ti}安装, ${ts:,.2f}花费"))
items.append(("总计", f"{len(final)}素材行(不含合计)"))

for ri, (l, v) in enumerate(items, 1):
    ws2.cell(ri,1,l).font = Font(name="Microsoft YaHei", size=10, bold=True)
    ws2.cell(ri,2,v).font = Font(name="Microsoft YaHei", size=10)
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 60

wb.save(OUT)
print(f"Done: {len(final)} rows + {4} summaries")
for m in ["4月","6月"]:
    for b in ["AEO","Install"]:
        n = len([r for r in final if r["month"]==m and r["bidding"]==b])
        if n: print(f"  {m} {b}: {n} materials")
