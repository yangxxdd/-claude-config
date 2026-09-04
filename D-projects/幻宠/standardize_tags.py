import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import os, shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r'D:\claude-projects\projects\幻宠\素材'

# =====================================================
# STANDARDIZED TAG MAPPING (compiled from 3 agent outputs + self-review)
# Format: (relative_path, [core_tags], [style_tags], notes)
# =====================================================
files_4月 = [
    ('4月/P-海啸.jpg',           ['天灾/生存'], ['宠物伙伴'], ''),
    ('4月/P-幻想宠物3.png',      ['宠物展示', '收集'], [], '150+帕基图鉴展示'),
    ('4月/P-灾后重建.jpg',       ['模拟经营', '建造'], [], '废墟vs重建对比'),
    ('4月/V-场景展示.mp4',       ['模拟经营', '建造'], [], '天灾后重建场景全览'),
    ('4月/V-高效抓宠.mp4',       ['抓宠', '战斗', '宠物展示'], [], '野外探索+回合战斗+社交拍照'),
    ('4月/V-核心玩法.mp4',       ['战斗', '抓宠', '探索/冒险'], [], '魔法宝石→回合战斗→世界探索'),
    ('4月/V-帕基世界冒险.mp4',   ['探索/冒险', '模拟经营'], ['像素风'], '多场景切换冒险'),
    ('4月/V-收集升级.mp4',       ['模拟经营', '建造', '收集'], [], '砍树收集+建筑升级'),
    ('4月/V-天灾重建-长版.mp4',  ['模拟经营', '建造', '收集'], [], '废墟→繁荣完整重建过程'),
    ('4月/V-抓宠战斗.mp4',       ['抓宠', '战斗'], [], '野外捕捉+回合战斗+Boss战'),
]

files_6月 = [
    ('6月/P-海啸.jpg',                    ['天灾/生存'], ['宠物伙伴'], '与4月版相同'),
    ('6月/P-宠物展示-二阶进化.jpg',        ['宠物进化', '宠物展示'], ['搞笑/趣味'], 'LV5→LV50进化展示'),
    ('6月/P-宠物展示-合成 3D.jpg',         ['宠物融合/合成', '宠物进化', '宠物展示'], [], 'SSR龙合成进化3D'),
    ('6月/P-宠物展示-巨物.jpg',            ['宠物展示'], [], '巨型龙vs小人剪影'),
    ('6月/P-宠物展示-帕基战斗.png',        ['战斗'], ['竞品素材'], '⚠ Pokemon UNITE截图,非幻宠素材'),
    ('6月/P-模拟经营-冬日写实.png',        ['模拟经营', '建造'], [], '雪地村庄建设'),
    ('6月/P-模拟经营-建造成长.jpg',        ['模拟经营', '建造', '收集'], [], '俯视田园农场建造'),
    ('6月/P-抓宠经营-超梦.png',            ['抓宠'], ['暗黑/暴力'], '暗黑工业风驯服/释放选择'),
    ('6月/P-抓宠经营-虐待.jpg',            ['抓宠'], ['暗黑/暴力'], '手术台捆绑虐待,血腥'),
    ('6月/P-抓宠经营-血腥.jpg',            ['抓宠'], ['暗黑/暴力'], '工厂传送带加工宠物,血腥'),
    ('6月/P-抓宠经营-幽飘.png',            ['抓宠', '探索/冒险'], [], '森林中偶遇传说SSR生物'),
    ('6月/V-场景展示.mp4',                 ['模拟经营', '建造'], [], '与4月版相同'),
    ('6月/V-宠物展示-宠物合成.mp4',        ['宠物展示', '宠物融合/合成'], [], '白紫巨龙→狐兔少女→法老战士'),
    ('6月/V-宠物展示-二阶合成.mp4',        ['宠物展示', '宠物融合/合成', '宠物进化'], [], '神圣翼龙→粉色萌兔→绿色狐精'),
    ('6月/V-高效抓宠.mp4',                 ['抓宠', '战斗', '宠物展示'], [], '与4月版相同'),
    ('6月/V-核心玩法.mp4',                 ['战斗', '抓宠', '探索/冒险'], [], '与4月版相同'),
    ('6月/V-模拟经营-砍树.mp4',            ['模拟经营', '建造', '收集', '战斗'], [], '砍树→基地防御塔战'),
    ('6月/V-模拟经营-帕基玩法寒霜.mp4',    ['模拟经营', '探索/冒险', '收集'], [], '冰原/海洋/沙漠多场景帕基互动'),
    ('6月/V-模拟经营-七日建造.mp4',        ['模拟经营', '建造'], [], 'DAY2→DAY7建造进度展示'),
    ('6月/V-模拟经营-温馨.mp4',            ['模拟经营'], ['温馨/治愈'], '锻造+浆果收集+烹饪经营'),
    ('6月/V-帕萌战斗-出狱打鸡.mp4',        ['战斗'], ['搞笑/趣味'], '餐厅烹饪+鸡群+喷火混乱'),
    ('6月/V-帕萌战斗-合成狙击.mp4',        ['战斗', '抓宠'], [], '雪山抓宠+Boss战+技能系统'),
    ('6月/V-帕萌战斗-群殴打鸡.mp4',        ['战斗'], ['搞笑/趣味'], '蛋黄飞溅+农场鸡群+金色光束'),
    ('6月/V-帕萌战斗-杀宠复刻.mp4',        ['战斗'], ['剧情/故事'], '小熊玩偶情感铺垫→机械Boss大战'),
    ('6月/V-帕萌战斗-雪地竞品.mp4',        ['战斗'], [], '雪地Boss战+39600伤害数值展示'),
    ('6月/V-天灾重建-长版.mp4',            ['模拟经营', '建造', '收集'], [], '与4月版相同'),
    ('6月/V-抓宠经营-捕捞竞品.mp4',        ['抓宠', '模拟经营'], [], '市集交易+海洋捕捞+收集'),
    ('6月/V-抓宠经营-狐狸.mp4',            ['战斗'], ['搞笑/趣味', '雪地'], '雪地互动+跳绳社交+战斗'),
    ('6月/V-抓宠经营-虐待.mp4',            ['战斗', '抓宠'], [], '俯视角砖头攻击+怪兽战斗'),
    ('6月/V-抓宠经营-售卖帕基.mp4',        ['模拟经营', '建造'], [], '传送带资源流转+设施升级'),
    ('6月/V-抓宠经营-拯救可达鸭.mp4',      ['战斗', '建造'], [], '回合制战斗→喷气背包→村庄'),
    ('6月/V-抓宠战斗.mp4',                 ['抓宠', '战斗', '宠物进化'], [], '与4月版相同'),
]

# Combine
all_files = files_4月 + files_6月

# =====================================================
# Tag stats
# =====================================================
from collections import Counter
core_counter = Counter()
style_counter = Counter()
for _, cores, styles, _ in all_files:
    for c in cores: core_counter[c] += 1
    for s in styles: style_counter[s] += 1

print('=== 核心方向标签分布 ===')
for tag, count in core_counter.most_common():
    print(f'  {tag}: {count}')
print(f'\n=== 风格标签分布 ===')
for tag, count in style_counter.most_common():
    print(f'  {tag}: {count}')
print(f'\n总文件数: {len(all_files)}, 其中重复文件: 6对')

# =====================================================
# Create Excel per folder
# =====================================================
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
dup_fill = PatternFill('solid', fgColor='FFF2CC')
norm_font = Font(name='Arial', size=11)
thin_b = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
c_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
l_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

for month in ['4月', '6月']:
    wb = Workbook()
    ws = wb.active
    ws.title = f'{month}素材标签'

    # Filter files for this month
    month_files = [(p, c, s, n) for p, c, s, n in all_files if p.startswith(month)]

    # Column widths
    for i, w in enumerate([8, 40, 15, 45, 30, 45], 1):
        ws.column_dimensions[chr(64+i)].width = w

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f'幻宠 {month} 素材内容标签（基于完整视觉审查）'
    ws['A1'].font = Font(name='Arial', bold=True, size=13, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Headers
    headers = ['序号', '原文件名', '类型', '核心方向标签', '风格/辅助标签', '备注']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = c_align; c.border = thin_b
    ws.row_dimensions[3].height = 30

    for idx, (path, cores, styles, notes) in enumerate(month_files, 1):
        row = 3 + idx
        filename = os.path.basename(path)
        # Detect type
        ext = os.path.splitext(filename)[1].lower()
        ftype = '图片' if ext in ['.jpg','.jpeg','.png','.webp'] else '视频'

        ws.cell(row=row, column=1, value=idx).font = norm_font
        ws.cell(row=row, column=1).alignment = c_align
        ws.cell(row=row, column=1).border = thin_b

        ws.cell(row=row, column=2, value=filename).font = norm_font
        ws.cell(row=row, column=2).alignment = l_align
        ws.cell(row=row, column=2).border = thin_b

        ws.cell(row=row, column=3, value=ftype).font = norm_font
        ws.cell(row=row, column=3).alignment = c_align
        ws.cell(row=row, column=3).border = thin_b

        ws.cell(row=row, column=4, value=', '.join(cores)).font = norm_font
        ws.cell(row=row, column=4).alignment = l_align
        ws.cell(row=row, column=4).border = thin_b

        ws.cell(row=row, column=5, value=', '.join(styles) if styles else '—').font = norm_font
        ws.cell(row=row, column=5).alignment = l_align
        ws.cell(row=row, column=5).border = thin_b

        ws.cell(row=row, column=6, value=notes if notes else '—').font = norm_font
        ws.cell(row=row, column=6).alignment = l_align
        ws.cell(row=row, column=6).border = thin_b

        # Highlight duplicates
        if '与4月版相同' in notes or '与6月版相同' in notes:
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = dup_fill

        ws.row_dimensions[row].height = 22

    # Tag legend below data
    last_row = 3 + len(month_files) + 1
    ws.merge_cells(f'A{last_row}:F{last_row}')
    ws[f'A{last_row}'] = '黄色底色 = 与另一月份文件内容完全相同（字节级重复），非独立新素材'
    ws[f'A{last_row}'].font = Font(name='Arial', size=9, color='996600')

    out_path = os.path.join(BASE, month, f'{month}素材内容标签.xlsx')
    wb.save(out_path)
    print(f'Saved: {out_path}')

# =====================================================
# Rename files with tags appended
# =====================================================
print('\n=== 文件重命名 ===')
renamed = 0
for path, cores, styles, notes in all_files:
    full_path = os.path.join(BASE, path)
    if not os.path.exists(full_path):
        print(f'  SKIP (not found): {path}')
        continue

    dir_name = os.path.dirname(full_path)
    base_name = os.path.splitext(os.path.basename(full_path))[0]
    ext = os.path.splitext(os.path.basename(full_path))[1]

    # Build tag string: core tags + style tags
    # Windows forbids * / \ : < > ? " | in filenames, replace / with -
    all_tags = [t.replace('/', '-') for t in cores + styles]
    tag_str = '#'.join(all_tags)

    # Check if already has tags in name
    if '#' in base_name:
        print(f'  SKIP (已标注): {os.path.basename(full_path)}')
        continue

    new_name = f'{base_name}#{tag_str}{ext}'
    new_path = os.path.join(dir_name, new_name)

    # Handle filename too long
    if len(new_name) > 200:
        # Trim style tags
        tag_str_short = '*'.join(cores)
        new_name = f'{base_name}#{tag_str_short}{ext}'
        new_path = os.path.join(dir_name, new_name)

    try:
        os.rename(full_path, new_path)
        print(f'  {os.path.basename(full_path)} → {os.path.basename(new_path)}')
        renamed += 1
    except Exception as e:
        print(f'  ERROR renaming {path}: {e}')

print(f'\nRenamed: {renamed} files')
print('Done!')
