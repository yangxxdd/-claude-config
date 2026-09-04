# -*- coding: utf-8 -*-
import os
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

BASE = r'D:\claude-projects\projects\幻宠'
path = os.path.join(BASE, '幻宠_V5_商店七图需求与AI提示词.xlsx')
wb = load_workbook(path)
ws = wb['V5_七图需求']

F = lambda **kw: Font(name='微软雅黑', **kw)
mono = Font(name='Consolas', size=10)
title_fill = PatternFill('solid', fgColor='DDEBF7')
label_font = F(bold=True, size=11, color='1F4E79')
body_font = F(size=10)
bold_font = F(bold=True, size=10)
wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
wrap_mono = Alignment(horizontal='left', vertical='top', wrap_text=False)
border = Border(*[Side(style='thin', color='BFBFBF')]*4)

def w_row(r, text, font=None, height=None, fill=None):
    ws.cell(row=r, column=1, value=text)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1)
    c.font = font or body_font
    c.alignment = wrap
    if fill:
        for j in range(1, 10):
            ws.cell(row=r, column=j).fill = fill
    if height:
        ws.row_dimensions[r].height = height

r = 18
w_row(r, '图3 深化方案：卡牌墙 × 进化破格（2026-07-29 追加）', F(bold=True, size=12, color='1F4E79'), 24, title_fill); r += 1
w_row(r, '参考评价：「卡牌墙 + 进化破格」是收集品类的经典商店图打法（Palmon 系验证过的）。与图3结合的思路：卡牌墙承担「收集」，进化主角承担「进化」，一静一动。', body_font, 30); r += 1
w_row(r, '核心创意：背景是一面「帕基图鉴卡牌墙」（=收集），一张小卡通过绿色进化箭头指向破格冲出的 MAX 形态主角（=进化），底部文案 EVOLVE & COLLECT。', bold_font, 30); r += 1
w_row(r, '画面内容（排版结构，1080×1920）：', label_font, 20); r += 1

ascii_art = (
    '┌─────────────────────┐\n'
    '│ 顶部 10%：图鉴标题栏   │  ←「Palkiedex 45/200+」+ 返回箭头+排序按钮\n'
    '│（轻量UI，真实游戏感）   │     一行带过，不抢戏\n'
    '├─────────────────────┤\n'
    '│ ┌卡┬卡┬卡┐            │\n'
    '│ ├卡┼卡┼卡┤   ╭──────╮ │\n'
    '│ ├小卡→箭头→│ MAX主角 │ │  ← 中部 65%：卡牌墙（收集）\n'
    '│ ├卡┼卡┼卡┤   │ 龙火鹿  │ │     左侧3×4卡牌矩阵\n'
    '│ └灰┴灰┴？┘   │ Lv.MAX │ │     右侧进化主角破格冲出\n'
    '│              ╰──────╯ │\n'
    '├─────────────────────┤\n'
    '│ 底部 15%：文案          │  ← EVOLVE & COLLECT\n'
    '│ EVOLVE & COLLECT      │     大号粗体，白字+深色底衬\n'
    '└─────────────────────┘'
)
ws.cell(row=r, column=1, value=ascii_art)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
c = ws.cell(row=r, column=1)
c.font = mono
c.alignment = wrap_mono
ws.row_dimensions[r].height = 240
r += 1

w_row(r, '排版要点（让美术照着排）：', label_font, 20); r += 1
points = (
    '1.卡牌墙：左侧占宽约55%，3列×4排卡牌矩阵，卡牌带等级标识（Lv.30/Lv.27/Lv.15…）和稀有度边框色（蓝/紫/金）；最下面一排2-3张做成灰色剪影+问号（未解锁=收集钩子，替代原来的45/200+面板）。\n'
    '2.进化主角：右侧占宽约45%、高约50%——龙火鹿MAX形态冲出卡牌边框（破格构图），背后火焰/光芒爆发，卡牌上标「Lv.MAX」；破格是整张图的视觉锚点，卡牌墙的「静」衬托它的「动」。\n'
    '3.进化叙事：矩阵中一张「萤火鹿 Lv.20」小卡 → 绿色弯曲箭头 → MAX主角，一眼读懂「小卡能养成大卡」；箭头是视觉动线的起点。\n'
    '4.视觉动线：小卡 → 箭头 → MAX主角 → 底部文案，Z字形一气呵成。\n'
    '5.背景：青蓝渐变UI底色（参考图同款），卡牌墙略微俯视倾斜增加动感，四周用光晕压暗、视线聚焦中央。\n'
    '6.色调：青蓝UI底 + 金/紫稀有度边框 + 主角火焰橙，和图1（田园绿金）、图2（森林明亮）拉开区隔但不跳脱。'
)
w_row(r, points, body_font, 130); r += 1

w_row(r, '可用物料（全部已审计真实存在）：', label_font, 20); r += 1
mats = [
    ('MAX 进化主角', '物料\\进化-龙火鹿-三阶.png（或 首图-龙火鹿.png，同一形象）'),
    ('小卡（进化前）', '物料\\进化-萤火鹿-二阶.png、物料\\进化-幼火鹿-一阶.png'),
    ('卡牌墙填充', '共享盘 帕基截图\\ 52张（火熊/木熊/木少女/水豚/漠砂犬等，含多只SSR级：帝影、翼炎龙、闪电松、嘟王蛟）'),
    ('卡牌头像（最推荐）', '共享盘 角色原画\\帕基\\ui_img_hero_portrait_*.png —— 游戏真实方形头像，天生就是卡牌框素材，直接套'),
    ('UI 参考', '共享盘 UI\\C_宠物列表.psd、UI\\C_宠物进化3.psd（真实图鉴/进化界面）'),
    ('构图参考', '商店图参考\\参考-3b卡牌墙进化.png（竞品图：Palkiedex卡牌墙+破格主角）'),
]
for name, val in mats:
    ws.cell(row=r, column=1, value=name)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=3, value=val)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    for j in range(1, 10):
        c = ws.cell(row=r, column=j)
        c.font = body_font; c.alignment = wrap; c.border = border
    ws.cell(row=r, column=1).font = bold_font
    ws.row_dimensions[r].height = 30
    r += 1

w_row(r, '额外建议：卡牌墙里有意识混 1-2 只SSR（帝影/翼炎龙），金框和普通蓝紫框形成品质差异——这是之前下单文档里「强化不同品质差异」的老需求，正好在这张图上实现。', body_font, 30); r += 1

w_row(r, 'AI 生成提示词 · 中文版（即梦/可灵/通义万相）：', label_font, 20); r += 1
cn = (
    '竖版9:16，3D卡通渲染，手游宠物图鉴界面宣传图，青蓝色渐变UI背景。\n'
    '画面是一面宠物收集卡牌墙：左侧3列×4排卡牌矩阵，每张卡牌是一只原创奇幻宠物（火焰鹿、木甲熊、冰晶企鹅、水豚、精灵狐等），卡牌带等级标识和稀有度颜色边框（蓝色/紫色/金色），最下面一排2-3张是灰色问号剪影（未解锁）。\n'
    '右侧视觉焦点：一张小卡牌通过绿色弯曲箭头，指向一只巨大的火焰鬃毛神鹿MAX形态——它冲破卡牌边框跃出，背后火焰与金色光芒爆发，卡牌上写Lv.MAX，破格构图，充满力量感。\n'
    '顶部：轻量图鉴标题栏（收集进度45/200+）。\n'
    '整体明亮通透，卡牌游戏质感，静动的对比强烈。\n'
    '画面底部15%留干净深色渐变区域供后期加标题。\n'
    '不要文字错误、水印、皮卡丘等已有IP角色。'
)
w_row(r, cn, body_font, 130); r += 1

w_row(r, 'AI Prompt · English（Midjourney/SD/Flux）：', label_font, 20); r += 1
en = (
    'Vertical 9:16, 3D stylized cartoon render, mobile game pet collection dex screen key art, cyan-blue gradient UI background. '
    'A wall of collectible pet cards: left side 3×4 grid of cards, each featuring an original fantasy creature (fire deer, wood-armored bear, ice penguin, capybara, spirit fox), '
    'cards with level badges and rarity-colored borders (blue/purple/gold); bottom row 2-3 cards are grey locked silhouettes with question marks. '
    'Right side focal point: a small card connected by a curved green arrow to a giant MAX-form fire-maned stag bursting out of its card frame, '
    'flames and golden light exploding behind, "Lv.MAX" badge, dramatic frame-breaking composition, full of power. '
    'Top: light dex title bar with collection progress 45/200+. Bright and clean, card-game aesthetic, strong contrast between static grid and dynamic hero. '
    'Bottom 15% clean dark gradient area reserved for title. --no wrong text, watermark, Pikachu, existing IP characters --ar 9:16 --style raw'
)
w_row(r, en, body_font, 120); r += 1

w_row(r, '生成后验收重点：① 卡牌墙 vs 主角的静动对比是否成立；② 绿色进化箭头的动线是否清晰；③ 灰色剪影卡是否有「未解锁」感；④ 底部留白够不够放文案。别看宠物形象（正式图美术用 ui_img_hero_portrait 头像和龙火鹿物料替换）。', bold_font, 30); r += 1

# 重新嵌入参考缩略图（openpyxl 保存会丢失原有图片）
ref_imgs = {11: 'v5_首图参考-2.png', 12: 'v5_参考-1.png', 13: 'v5_参考-1a.png', 15: 'v5_参考-2a.png', 16: 'v5_参考-3.png', 17: 'v5_参考-4.png'}
thumb_dir = os.path.join(BASE, '_thumbs')
for row, fname in ref_imgs.items():
    tp = os.path.join(thumb_dir, fname)
    im = PILImage.open(tp)
    img = XLImage(tp)
    img.width, img.height = im.width, im.height
    ws.add_image(img, f'H{row}')

wb.save(path)
print('saved, last row:', r - 1)
