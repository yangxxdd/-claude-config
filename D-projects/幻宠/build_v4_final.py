# -*- coding: utf-8 -*-
import os
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

BASE = r'D:\claude-projects\projects\幻宠'
REF_DIR = os.path.join(BASE, '商店图参考')
THUMB_DIR = os.path.join(BASE, '_thumbs')
os.makedirs(THUMB_DIR, exist_ok=True)
path = os.path.join(BASE, 'Palkie_V3_Store_Brief_Final.xlsx')

def thumb(fname, max_w=150, max_h=240):
    src = os.path.join(REF_DIR, fname)
    im = PILImage.open(src)
    im.thumbnail((max_w, max_h))
    out = os.path.join(THUMB_DIR, 't_' + fname)
    im.save(out)
    return out, im.width, im.height

wb = load_workbook(path)
for s in ['V4_商店五图需求', 'V4_美术沟通版']:
    if s in wb.sheetnames:
        del wb[s]

F = lambda **kw: Font(name='微软雅黑', **kw)
header_fill = PatternFill('solid', fgColor='305496')
header_font = F(bold=True, color='FFFFFF', size=10)
body_font = F(size=10)
title_font = F(bold=True, size=14)
note_font = F(size=10, color='404040')
wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(*[Side(style='thin', color='BFBFBF')]*4)

# ============ Sheet 1: V4_商店五图需求 ============
ws = wb.create_sheet('V4_商店五图需求')
ws['A1'] = '幻宠帝国 V4 商店五图需求（SLG + 帕基进化方向 · 美术下单版）'
ws['A1'].font = title_font
ws.merge_cells('A1:L1')

notes = [
    '定位：承接 V3 SLG 方向，五图画面全部基于本地物料与参考图落地，无虚构元素。主角帕基严格用自家形象，背景生物/敌方BOSS/联盟单位可泛化设计（无IP风险的原创形象）。',
    '目标用户：SLG/策略玩家为主，帕基收集与进化作为差异化钩子（首图抓 SLG 体量感，后四图用宠物钩子拉回来）。',
    '英文文案：已定稿，锁定在「翻译（图上英文）」列，不得改动。PALKI / EMPIRE / EVOLVE / CONQUER 四个关键词已分配到各图。',
    'ASO原则：前3张截图承担约70%转化权重，90%用户不会滑过第3张；带文案截图比纯UI转化率高15-30%；首图必须在1秒内传达核心卖点；每张图文案大号粗体、占画面≤20%。',
    '输出规格：1080×1920 像素（竖屏 9:16），JPG，RGB，每张≤8MB。命名：01-BUILD-YOUR-PALKI-EMPIRE.jpg ~ 05-ALLY-CONQUER-TOGETHER.jpg。',
    '物料审计（2026-07-28）：本地物料已与共享盘源文件逐一哈希核对，全部真实可靠；原「联盟-大世界地图.png」实为错标（真身=主堡大世界外观），已更名为「建造-主堡大世界外观.png」。⚠ 大世界地图无现成2D资产（大世界为UE场景），图1远景/图5地图需游戏内截图或美术原创重绘。',
    '配套文档：「V4_美术沟通版」sheet 含每张参考图的符合度评估与必须修改清单，下单时一并发给美术。',
]
for i, t in enumerate(notes, start=2):
    c = ws.cell(row=i, column=1, value=t)
    c.font = note_font
    c.alignment = wrap
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=12)
    ws.row_dimensions[i].height = 30

headers = ['图号','主题','素材类型','文案（中文）','翻译（图上英文·定稿）','需求说明','可用物料（本地）','备注（为什么这张图）','比例','尺寸','参考图（主力）','参考图文件']
for j, h in enumerate(headers, 1):
    c = ws.cell(row=9, column=j, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
ws.row_dimensions[9].height = 28

rows1 = [
    ['1','首图·帝国建造','商店图','打造你的帕基帝国','BUILD YOUR\nPALKI EMPIRE',
     '1.三层纵深构图：前景3-5只明星帕基（龙火鹿/豪木熊/帝冰鹅/幽奈娅，严格按物料PNG形象）站在高地/岩石上面向观众，呈主角团阵型；中景繁荣基地（主堡、工坊、资源建筑，帕基劳作）；远景大世界（山脉、浮空岛、远处城堡）。\n2.帕基表情坚毅、战斗姿态，禁止微笑卖萌。\n3.主色调蓝+金：深蓝/藏蓝背景，金色建筑与帕基点缀，高对比。\n4.微俯视角，展现基地规模与三层景深，要电影感，不要扁平截图感。\n5.英文文案置于画面下方15-20%区域，大号粗体无衬线全大写，白字+深色渐变底衬或金色描边，缩略图下可辨认（眯眼测试）。',
     '物料\\首图-龙火鹿.png\n物料\\首图-豪木熊.png\n物料\\首图-帝冰鹅.png\n物料\\首图-幽奈娅.png\n物料\\建造-主堡内城.png\n物料\\建造-主堡Lv1.png / Lv7.png\n物料\\角色-【帕基S】角色资源汇总.jpg\n物料\\角色-【帕基S】场景资源汇总.png\n（远景大世界无现成2D资产，需游戏内/UE截图）',
     '首图承担最大转化权重，用史诗规模感筛选SLG玩家。数据：规模感素材CPI最低（$3.3-3.5）；V2首图只有Logo无文案，是最大转化损失点。',
     '9:16','1080×1920','','首图参考-2.png（主力90%）\n首图参考-1.png（辅助60%）'],
    ['2','抓宠·进化·收集','商店图','抓捕 · 进化 · 收集','CATCH, EVOLVE\n& COLLECT',
     '1.从下到上叙事流：底部——白天森林中发现野生帕基+投掷捕捉道具的抓捕瞬间（训练师手部/道具可泛化设计）；中部——火鹿三阶进化链（幼火鹿→萤火鹿→龙火鹿，严格按物料PNG，光效箭头连接）；顶部——图鉴UI局部（45/200+，未收集为剪影）。\n2.三阶外观变化必须逐级明显（体型/装饰/光效逐级升级）。\n3.色调明亮鲜艳，白天场景，展示帕基多样性但不杂乱。\n4.文案置于顶部或底部，不遮挡进化链。\n5.稀有度标签如出现，只用规范英文 Common / Rare / Legendary。',
     '物料\\进化-幼火鹿-一阶.png\n物料\\进化-萤火鹿-二阶.png\n物料\\进化-龙火鹿-三阶.png\n物料\\角色-训练师弗兰克.png（抓捕场景训练师）\n物料\\角色-【帕基S】角色资源汇总.jpg\nUI参考（共享盘）：UI\\C_宠物捕捉.psd、C_宠物进化3.psd',
     '已验证CVR最高的王牌方向：抓宠+经营联动CVR 26-32%，最高单素材32.26%。竞品无法复制的差异化（WOS没宠物，Palmon没我们的帕基）。',
     '9:16','1080×1920','','参考-1a.png（主力85%）\n参考-1.png（备选65%）'],
    ['3','基地·建造经营','商店图','建造并经营你的庇护所','BUILD & MANAGE\nYOUR SANCTUARY',
     '1.核心视觉：同一建筑升级前后对比（用物料：伐木场/农场/晶矿场 Lv1 vs Lv5，任选1-2组），光效箭头+等级标识 Lv.1→Lv.5 连接，升级变化要明显。\n2.帕基在基地干活：采集/搬运/巡逻，展示「分配帕基做任务」的SLG机制（帕基按物料形象）。\n3.减少军事要塞感（少城墙/兵营），增加庇护所生活设施：训练场/孵化室/温泉/帕基互动区（可泛化设计）。\n4.暖色调：森林绿+木色+丰收金，可点缀少量魔法蓝绿。\n5.角落轻量UI：资源数量（金币/木材/石材）+建造进度条，不抢主视觉。',
     '物料\\建造-主堡内城.png\n物料\\建造-主堡Lv1.png / Lv7.png\n物料\\建造-伐木场Lv1.png / Lv5.png\n物料\\建造-农场Lv1.png / Lv5.png\n物料\\建造-晶矿场Lv1.png / Lv5.png\n物料\\角色-【帕基S】角色资源汇总.jpg\n共享盘更多建筑：最新建筑0409\\（石矿场1-5、温泉、餐厅、研究所、医院、募兵所、指挥所、盾/弓/矛兵营1-5、主堡1-7）',
     '建造经营是SLG付费核心，回答「这游戏能玩多久」。模拟经营CVR 15-23%；4月测试建造类素材留存最好。',
     '9:16','1080×1920','','参考-2a.png（主力85%）\n参考-2.png（辅助75%）\n参考-2b.png（备选70%）'],
    ['4','策略·团队战斗','商店图','训练你的帕基战队','TRAIN YOUR PALKI\n& CONQUER',
     '1.3-5只自家战斗帕基（炽角兽/跃焰虎等，严格按物料PNG）组成战队 vs 一只巨型BOSS（BOSS可泛化设计，必须是无IP风险的原创形象，体型有压迫感）。\n2.暗色背景（深紫/暗蓝）+高亮技能特效（火焰橙/雷电金/冰霜青）：粒子、冲击波、光轨，特效必须华丽，静态图也要有动感。\n3.BOSS带HP血条，显示血量被削减中（如60%）。\n4.轻量战斗UI：底部或侧边队伍头像+技能图标，不遮挡战斗主体。\n5.帕基表情专注战斗，禁止卖萌。\n6.若游戏内实际战斗画面表现力不足，需UE渲染加特效后再截图。',
     '物料\\战斗-炽角兽.png\n物料\\战斗-跃焰虎.png\n物料\\角色-【帕基S】角色资源汇总.jpg\n共享盘：帕基截图\\13【跃焰虎】10132（二阶）、帕基技能\\（各帕基技能特效参考）\n（BOSS无现成资产，泛化原创设计）',
     '战斗CVR方差极大：表现力好的素材CVR 25%，差的仅12%。视觉品质直接=转化率，特效不手软。',
     '9:16','1080×1920','','参考-3.png（主力90%）'],
    ['5','联盟·大世界征服','商店图','结盟一起征服','ALLY & CONQUER\nTOGETHER',
     '1.俯瞰视角大世界地图。⚠ 大世界地图无现成2D资产（大世界为UE场景）：优先游戏内/UE截图，不行则美术按参考-4原创重绘。多地形：山脉/河流/沙漠/森林/火山。\n2.四方联盟用鲜明颜色+旗帜区分（红/蓝/绿/棕），领土边界清晰；联盟名称/旗帜可泛化设计。\n3.中央争议区（Disputed Zone）：战斗/爆炸效果，多方行军箭头交汇于此。\n4.行军路线箭头用不同颜色对应不同联盟，制造「战争正在发生」的紧迫感。\n5.右下角联盟UI弹窗：联盟名+在线成员 145/200（示意数字）+ WAR! 红色按钮 + JOIN 蓝色按钮。\n6.金色+深色地图底色，史诗感；与前4张的斜俯视角形成区隔。',
     '物料\\建造-主堡大世界外观.png（仅主堡外观，非地图）\n物料\\建造-主堡内城.png\n物料\\角色-【帕基S】角色资源汇总.jpg\n⚠ 无真实世界地图2D资产，需游戏内截图或美术原创重绘\nUI参考（共享盘）：UI\\L_联盟.psd、S_世界行军战斗.psd',
     '联盟社交=SLG终极留存与付费发动机，打破「单机游戏」错觉。WOS用户评价反复验证联盟是持续玩下去的核心原因；幻宠联盟系统已实装。',
     '9:16','1080×1920','','参考-4.png（主力95%）\n参考-4b.png（辅助95%）'],
]
for i, row in enumerate(rows1, start=10):
    for j, v in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = body_font; c.alignment = wrap; c.border = border
    ws.row_dimensions[i].height = 245

widths1 = [5, 15, 9, 16, 22, 80, 40, 40, 7, 12, 23, 26]
for j, w in enumerate(widths1, 1):
    ws.column_dimensions[get_column_letter(j)].width = w

main_refs = ['首图参考-2.png','参考-1a.png','参考-2a.png','参考-3.png','参考-4.png']
for idx, fname in enumerate(main_refs):
    tp, w, h = thumb(fname)
    img = XLImage(tp)
    img.width, img.height = w, h
    ws.add_image(img, f'K{10+idx}')

# ============ Sheet 2: V4_美术沟通版 ============
ws2 = wb.create_sheet('V4_美术沟通版')
ws2['A1'] = 'V4 美术沟通版 —— 参考图符合度评估与修改清单（随需求一并下发）'
ws2['A1'].font = title_font
ws2.merge_cells('A1:G1')

rules = [
    '【通用红线 · 适用于全部5张图】',
    '1.角色红线：所有参考图均为AI概念图，含大量宝可梦/幻兽帕鲁形象（皮卡丘/喷火龙/伊布/Pengullet/Incineram等），仅参考构图与氛围，禁止描图或复用角色。主角帕基必须严格按 商店图参考\\物料 文件夹中的PNG形象。',
    '2.文案规范：5张图英文文案已定稿（见需求sheet「翻译」列），大号粗体无衬线全大写，占画面≤20%，缩略图尺寸下必须可辨认。',
    '3.去中文/错字：参考图中的中文、以及「稀说/传有/JOINS UP」等错误文字，一律不得出现在成品中。',
    '4.背景生物可泛化：非主角的生物、敌方BOSS、联盟单位可自由设计，但必须是无IP风险的原创形象。',
    '5.色调统一：5张图均为3D渲染质感，按各图指定色调执行，保持系列感。',
    '6.验收标准：眯眼测试文案可辨认；每张图的「必须改」清单逐条核对后方可交付。',
]
for i, t in enumerate(rules, start=2):
    c = ws2.cell(row=i, column=1, value=t)
    c.font = F(bold=(i == 2), size=10)
    c.alignment = wrap
    ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)
    ws2.row_dimensions[i].height = 30

headers2 = ['图号','主题','主力参考图','辅助参考图1','辅助参考图2','参考图里学什么','必须改 / 替换什么']
for j, h in enumerate(headers2, 1):
    c = ws2.cell(row=9, column=j, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
ws2.row_dimensions[9].height = 28

rows2 = [
    ['1','首图·帝国建造','首图参考-2.png','首图参考-1.png',None,
     '①前景帕基→中景城镇→远景城堡浮岛的三层纵深；\n②军队+城堡+工地的史诗规模感；\n③底部大号英文文案的字体风格与位置。',
     '①前景喷火龙/梦幻/乌鸦等全部换成自家明星帕基（龙火鹿/豪木熊/帝冰鹅/幽奈娅）；\n②底部文案换成 BUILD YOUR PALKI EMPIRE；\n③参考图偏暖金，成品按蓝+金色调执行；\n④首图参考-1 仅参考构图层次，其萌系温馨氛围与中文文案不要学。'],
    ['2','抓宠·进化·收集','参考-1a.png','参考-1.png',None,
     '①白天森林明亮氛围；\n②纵向进化链+发光箭头；\n③顶部收集进度UI（45/200+）。',
     '①所有宝可梦角色换成火鹿三阶（物料PNG），野生帕基与训练师手部/捕捉道具可泛化；\n②参考图没有「抓捕瞬间」，需在底部增加投掷捕捉道具的画面；\n③「稀说/传有」等错误文字一律删除，标签用规范英文；\n④参考-1仅备选，其夜晚暗色调不符合「明亮鲜艳」要求，不要学。'],
    ['3','基地·建造经营','参考-2a.png','参考-2.png','参考-2b.png',
     '①升级前后对比+光效箭头；\n②宠物元素与基地融合的方式；\n③资源面板UI（金币/木材/石材+进度条）。',
     '①建筑必须换成物料里的主堡/伐木场/农场/晶矿场 Lv1→Lv5，不要参考图的欧式城堡；\n②减少城墙/兵营/防御塔比例，增加训练场/孵化室/温泉/帕基互动区；\n③参考-2a偏蓝绿冷调，成品要增加暖色（森林绿+木色+丰收金）；\n④参考-2b的Q版萌系风格不要学；\n⑤中文「升级前后」等文字不要出现。'],
    ['4','策略·团队战斗','参考-3.png',None,None,
     '①4宠团队vs巨型BOSS的构图与包围态势；\n②暗背景+亮技能特效的对比公式；\n③底部队伍UI+BOSS血条的集成方式。',
     '①Pengullet/Lamball/Tanzee/Incineram 换成炽角兽/跃焰虎等自家战斗帕基；\n②BOSS（Ancient Abomination）重新设计为原创巨型魔兽，禁止沿用幻兽帕鲁形象；\n③UI中的宠物名/数值改为示意或删除；\n④底部加文案 TRAIN YOUR PALKI & CONQUER；\n⑤特效华丽度只能加不能减。'],
    ['5','联盟·大世界征服','参考-4.png','参考-4b.png',None,
     '①四方联盟环争议区的地图布局；\n②行军箭头+领土边界的表达；\n③右下角联盟UI弹窗（在线人数+WAR按钮）。',
     '①联盟名称/旗帜换成自家设定（可泛化命名）；\n②地图上的宝可梦式生物换成泛化奇幻生物或自家帕基；\n③「JOINS UP」语法错误，按钮改为 JOIN；\n④参考-4b仅作配色备选；\n⑤大世界地图无现成2D资产：优先游戏内/UE截图，不行则按参考-4原创重绘；主堡外观可用 物料\\建造-主堡大世界外观.png。'],
]
for i, row in enumerate(rows2, start=10):
    for j, v in enumerate(row[:2] + row[5:], 1):
        pass
    # write text columns: A,B then F,G
    ws2.cell(row=i, column=1, value=row[0])
    ws2.cell(row=i, column=2, value=row[1])
    ws2.cell(row=i, column=6, value=row[5])
    ws2.cell(row=i, column=7, value=row[6])
    for j in range(1, 8):
        c = ws2.cell(row=i, column=j)
        c.font = body_font; c.alignment = wrap; c.border = border
    ws2.row_dimensions[i].height = 245

widths2 = [5, 15, 23, 23, 23, 45, 65]
for j, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(j)].width = w

for idx, row in enumerate(rows2):
    r = 10 + idx
    for col_letter, fname in zip(['C', 'D', 'E'], row[2:5]):
        if fname:
            tp, w, h = thumb(fname)
            img = XLImage(tp)
            img.width, img.height = w, h
            ws2.add_image(img, f'{col_letter}{r}')

wb.save(path)
print('saved ok')
print('sheets:', wb.sheetnames)
