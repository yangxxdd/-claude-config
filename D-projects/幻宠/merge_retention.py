import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from copy import copy

# ============================================================
# STEP 1: Load FB + BI, join, aggregate
# ============================================================
fb_sheets_map = {
    ('4月', 'Install'): '4月install美国幻宠',
    ('4月', 'AEO'): '4月AEO美国幻宠',
    ('6月', 'Install'): '6月install美国',
    ('6月', 'AEO'): '6月AEO美国',
}

fb_parts = []
for (month, bid), sheet_name in fb_sheets_map.items():
    df = pd.read_excel('fb媒体-源数据.xlsx', sheet_name=sheet_name)
    df['fb_month'] = month
    df['fb_bid'] = bid
    df['ad_id'] = df['广告编号'].astype(str).str.strip()
    df['ad_name'] = df['广告名称'].astype(str).str.strip()
    fb_parts.append(df[['fb_month', 'fb_bid', 'ad_name', 'ad_id', '移动应用安装量']])

fb_all = pd.concat(fb_parts, ignore_index=True)
print(f'[1] FB combined: {len(fb_all)} rows, {fb_all["ad_id"].nunique()} unique ad IDs')

# BI data
bi = pd.read_excel('fb媒体-源数据.xlsx', sheet_name='BI汇总')
bi['mat_id'] = bi['素材ID'].astype(str).str.strip()
bi = bi[bi['mat_id'] != 'unknown']
bi_agg = bi.groupby('mat_id').agg(
    dnu=('dnu', 'sum'),
    r1_cnt=('r1_cnt', 'sum'),
    r2_cnt=('r2_cnt', 'sum'),
    r3_cnt=('r3_cnt', 'sum'),
).reset_index()
print(f'[2] BI aggregated: {len(bi_agg)} unique 素材IDs')

# Join FB + BI
fb_ret = fb_all.merge(bi_agg, left_on='ad_id', right_on='mat_id', how='left')
fb_ret['R1'] = np.where(fb_ret['dnu'] > 0, fb_ret['r1_cnt'] / fb_ret['dnu'], np.nan)
fb_ret['R2'] = np.where(fb_ret['dnu'] > 0, fb_ret['r2_cnt'] / fb_ret['dnu'], np.nan)
fb_ret['R3'] = np.where(fb_ret['dnu'] > 0, fb_ret['r3_cnt'] / fb_ret['dnu'], np.nan)

matched = fb_ret['dnu'].notna().sum()
print(f'[3] FB+BI matched: {matched}/{len(fb_ret)}')

# Aggregate to creative name + bid + month level (weighted by DNU)
agg = fb_ret.groupby(['fb_month', 'fb_bid', 'ad_name'], dropna=False).agg(
    dnu_total=('dnu', 'sum'),
    r1_total=('r1_cnt', 'sum'),
    r2_total=('r2_cnt', 'sum'),
    r3_total=('r3_cnt', 'sum'),
    ad_count=('ad_id', 'nunique'),
).reset_index()
agg['R1'] = np.where(agg['dnu_total'] > 0, agg['r1_total'] / agg['dnu_total'], np.nan)
agg['R2'] = np.where(agg['dnu_total'] > 0, agg['r2_total'] / agg['dnu_total'], np.nan)
agg['R3'] = np.where(agg['dnu_total'] > 0, agg['r3_total'] / agg['dnu_total'], np.nan)
print(f'[4] Aggregated to {len(agg)} groups')

# ============================================================
# STEP 2: Read creative_final with openpyxl to preserve formatting
# ============================================================
wb = load_workbook('creative_final.xlsx')
ws = wb['素材表现汇总']

# Read all data
rows_data = []
for row in ws.iter_rows(min_row=1, values_only=True):
    rows_data.append(list(row))

header = rows_data[0]
print(f'[5] creative_final header: {header}')
print(f'    {len(rows_data)-1} data rows')

# Build column index map
col_map = {name: i for i, name in enumerate(header)}
print(f'    col_map: {col_map}')

# ============================================================
# STEP 3: Match and compute retention for each data row
# ============================================================
# Find rows to skip (separator/header rows)
results = []  # (row_idx, r1, r2, r3, dnu, r1_cnt, r2_cnt, r3_cnt, ad_count)
unmatched_list = []

for row_idx in range(1, len(rows_data)):
    row = rows_data[row_idx]
    month_raw = str(row[col_map['月份']]).strip() if row[col_map['月份']] is not None else ''
    name_raw = str(row[col_map['素材名称']]).strip() if row[col_map['素材名称']] is not None else ''
    bid_raw = str(row[col_map['出价方式']]).strip() if row[col_map['出价方式']] is not None else ''

    # Skip header/separator rows
    if month_raw.startswith('◆') or name_raw == 'None' or name_raw == '' or name_raw == 'nan':
        results.append((row_idx, None, None, None, None, None, None, None, None))
        continue

    # Extract clean month (first 2 chars)
    month_clean = month_raw[:2]

    # Find matching aggregated data
    match = agg[(agg['fb_month'] == month_clean) &
                (agg['fb_bid'] == bid_raw) &
                (agg['ad_name'] == name_raw)]

    if len(match) == 0:
        results.append((row_idx, None, None, None, None, None, None, None, None))
        unmatched_list.append(f'  {month_raw} | {name_raw} | {bid_raw} | installs={row[col_map["安装数"]]}')
    else:
        m = match.iloc[0]
        results.append((row_idx, m['R1'], m['R2'], m['R3'],
                       m['dnu_total'], m['r1_total'], m['r2_total'], m['r3_total'],
                       m['ad_count']))

matched_count = sum(1 for r in results if r[1] is not None)
print(f'\n[6] Matched: {matched_count}/{len(results)}')

if unmatched_list:
    print('=== UNMATCHED ===')
    for u in unmatched_list:
        print(u)

# ============================================================
# STEP 4: Print full comparison table
# ============================================================
print(f'\n{"="*120}')
print(f'FULL RESULT TABLE')
print(f'{"="*120}')
print(f'{"月份":<8} {"素材名称":<22} {"类型":<4} {"出价":<8} {"安装":>6} {"花费":>10} {"CPI":>7} {"DNU":>6} {"r1_cnt":>7} {"r2_cnt":>7} {"r3_cnt":>7} {"R1":>8} {"R2":>8} {"R3":>8} {"#Ads":>4}')
print(f'{"-"*120}')

for row_idx in range(1, len(rows_data)):
    row = rows_data[row_idx]
    month = str(row[col_map['月份']]) if row[col_map['月份']] is not None else ''
    name = str(row[col_map['素材名称']]) if row[col_map['素材名称']] is not None else ''
    typ = str(row[col_map['类型']]) if row[col_map['类型']] is not None else ''
    bid = str(row[col_map['出价方式']]) if row[col_map['出价方式']] is not None else ''
    installs = row[col_map['安装数']]
    spend = row[col_map['花费(USD)']]
    cpi = row[col_map['CPI']]

    r = results[row_idx - 1]
    if r[1] is not None:
        print(f'{month:<8} {name:<22} {typ:<4} {bid:<8} {installs:>6.0f} {spend:>10.2f} {cpi:>7.2f} {r[4]:>6.0f} {r[5]:>7.0f} {r[6]:>7.0f} {r[7]:>7.0f} {r[1]:>7.1%} {r[2]:>7.1%} {r[3]:>7.1%} {r[8]:>4.0f}')
    else:
        if name and name != 'None' and not month.startswith('◆'):
            print(f'{month:<8} {name:<22} {typ:<4} {bid:<8} {installs:>6.0f} {spend:>10.2f} {cpi:>7.2f} {"N/A":>6} {"N/A":>7} {"N/A":>7} {"N/A":>7} {"N/A":>8} {"N/A":>8} {"N/A":>8}')
        else:
            print(f'{month:<8} {"(separator)":<22}')

print(f'{"="*120}')

# ============================================================
# STEP 5: Write to new Excel file, preserving original formatting
# ============================================================
# Add new column headers
new_header_cols = ['DNU', 'r1_cnt', 'r2_cnt', 'r3_cnt', 'R1', 'R2', 'R3']
header_col_idx = len(header) + 1  # start after 备注

# Find the 备注 column index
notes_col = col_map['备注'] + 1  # 1-indexed for openpyxl

# Add headers in row 1
for i, h in enumerate(new_header_cols):
    cell = ws.cell(row=1, column=notes_col + 1 + i)
    cell.value = h
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='center')
    # Copy style from existing header
    ref_cell = ws.cell(row=1, column=notes_col)
    if ref_cell.fill:
        cell.fill = copy(ref_cell.fill)

# Add data
for row_idx in range(1, len(rows_data)):
    r = results[row_idx - 1]
    excel_row = row_idx + 1

    if r[1] is not None:
        ws.cell(row=excel_row, column=notes_col + 1).value = r[4]  # dnu
        ws.cell(row=excel_row, column=notes_col + 2).value = r[5]  # r1_cnt
        ws.cell(row=excel_row, column=notes_col + 3).value = r[6]  # r2_cnt
        ws.cell(row=excel_row, column=notes_col + 4).value = r[7]  # r3_cnt
        ws.cell(row=excel_row, column=notes_col + 5).value = r[1]  # R1
        ws.cell(row=excel_row, column=notes_col + 6).value = r[2]  # R2
        ws.cell(row=excel_row, column=notes_col + 7).value = r[3]  # R3

        # Format as percentage for R1/R2/R3
        for offset in [5, 6, 7]:
            c = ws.cell(row=excel_row, column=notes_col + offset)
            c.number_format = '0.0%'
    else:
        # For unmatched data rows, mark as N/A
        name_raw = str(rows_data[row_idx][col_map['素材名称']]).strip() if rows_data[row_idx][col_map['素材名称']] is not None else ''
        if name_raw and name_raw != 'None' and name_raw != 'nan':
            mon_raw = str(rows_data[row_idx][col_map['月份']]).strip() if rows_data[row_idx][col_map['月份']] is not None else ''
            if not mon_raw.startswith('◆'):
                ws.cell(row=excel_row, column=notes_col + 1).value = '无BI数据'

# Also update the 数据概览 sheet
ws2 = wb['数据概览']
# Add a note about retention data source
next_row = ws2.max_row + 2
ws2.cell(row=next_row, column=1).value = '留存数据来源'
ws2.cell(row=next_row, column=2).value = 'BI汇总(素材ID↔FB广告编号), 按素材+出价+月份聚合, DNU加权平均'
ws2.cell(row=next_row + 1, column=1).value = 'R1/R2/R3定义'
ws2.cell(row=next_row + 1, column=2).value = 'R1=r1_cnt/dnu (次日留存), R2=r2_cnt/dnu (2日留存), R3=r3_cnt/dnu (3日留存)'
ws2.cell(row=next_row + 2, column=1).value = '匹配率'
ws2.cell(row=next_row + 2, column=2).value = f'{matched_count}/{len(results)} 行匹配 ({matched_count/len(results)*100:.0f}%)'
ws2.cell(row=next_row + 3, column=1).value = '更新时间'
ws2.cell(row=next_row + 3, column=2).value = '2026-07-29'

output_path = 'creative_final.xlsx'
wb.save(output_path)
print(f'\n[7] Saved to {output_path}')
print('Done!')
