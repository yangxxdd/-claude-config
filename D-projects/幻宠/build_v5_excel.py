# -*- coding: utf-8 -*-
import os
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

BASE = r'D:\claude-projects\projects\幻宠'
REF_DIR = os.path.join(BASE, '商店图参考')
THUMB_DIR = os.path.join(BASE, '_thumbs')
os.makedirs(THUMB_DIR, exist_ok=True)
OUT = os.path.join(BASE, '幻宠_V5_商店七图需求与AI提示词.xlsx')

def thumb(fname, max_w=150, max_h=240):
    im = PILImage.open(os.path.join(REF_DIR, fname))
    im.thumbnail((max_w, max_h))
    out = os.path.join(THUMB_DIR, 'v5_' + fname)
    im.save(out)
    return out, im.width, im.height

wb = Workbook()
F = lambda **kw: Font(name='微软雅黑', **kw)
header_fill = PatternFill('solid', fgColor='305496')
header_font = F(bold=True, color='FFFFFF', size=10)
body_font = F(size=10)
title_font = F(bold=True, size=14)
note_font = F(size=10, color='404040')
wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
border = Border(*[Side(style='thin', color='BFBFBF')]*4)

STYLE_DEF = '全部7张统一「明亮版 Palmon Survival / 万龙觉醒式 3D 卡通渲染」：中高饱和度、柔和全局光照、材质细节精致、角色比例偏Q但不低幼。'

# ================= Sheet 1: V5_七图需求 =================
ws = wb.active
ws.title = 'V5_七图需求'
ws['A1'] = '幻宠帝国 商店七图需求 V5（宠物钩子前置 · SLG压轴 · 含画风定义）'
ws['A1'].font = title_font
ws.merge_cells('A1:I1')

notes = [
    '核心策略：宝可梦题材是用来买便宜量的——前4张全部放宠物钩子（家园/抓宠/进化/孵蛋），服务低价休闲受众；SLG深度内容（经营/战斗/联盟）放后3张做留存与付费承诺，不用来导量。',
    '受众漏斗：看得见的宠物乐趣（1-4）→ 玩得到的深度（5-6）→ 留得下的世界（7）。前3张承担约70%转化权重。',
    '画风定义（必须统一）：' + STYLE_DEF,
    '画风必须统一，色调可每张不同（战斗暗色调/地图深金色均可）——渲染风格、材质、光感一致，氛围色各服务主题。',
    '文案已定稿（「英文文案」列锁定不得改动）。必含关键词：PALKI / EMPIRE / EVOLVE / CONQUER。每张文案大号粗体无衬线、占画面≤20%、缩略图可辨认。',
    '输出规格：1080×1920（竖屏9:16），JPG，RGB，每张≤8MB。命名 01~07-英文主题.jpg。',
    '物料已审计（2026-07-28）：本地物料与共享盘源文件哈希核对一致。⚠ 大世界地图无2D资产（UE场景），图7需游戏内截图或美术按参考-4原创重绘。',
    '配套：「AI生图提示词」sheet 含7张图的中英提示词，用于生成构图/氛围参考图（AI画不准帕基形象，正式图由美术按物料执行）。',
]
for i, t in enumerate(notes, start=2):
    c = ws.cell(row=i, column=1, value=t)
    c.font = note_font; c.alignment = wrap
    ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=9)
    ws.row_dimensions[i].height = 30

headers = ['图号','主题','文案（中文）','英文文案（定稿）','画面内容（需求说明）','选用原因（为什么这张图）','可用物料（本地）','主力参考图','参考图文件']
for j, h in enumerate(headers, 1):
    c = ws.cell(row=10, column=j, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
ws.row_dimensions[10].height = 28

rows = [
    ['1','帕基田园家园（首图）','打造你的帕基帝国','BUILD YOUR\nPALKI EMPIRE',
     '1.三层纵深：前景4只明星帕基（龙火鹿/豪木熊/帝冰鹅/幽奈娅，按物料PNG）站草坡上，自信伙伴姿态面向观众（不卖萌也不狰狞）；中景田园家园——金黄农田+伐木场+晶矿场+蓝色小河与小码头，帕基浇水种田、砍树、采矿、搬运；远景山巅魔法城堡+浮空岛+瀑布雪山（规模感保留）。\n2.零军事：不要士兵/军队/兵营/硝烟。\n3.色调明亮自然：蓝天+翠绿+丰收金+湖水蓝，点缀魔法光效。\n4.文案置于底部15-20%，白字+深色渐变底衬或金色描边，大号粗体全大写。',
     '门面图。CVR最高的「宠物+经营」组合直接做门面，服务低价休闲受众；规模感靠远景城堡浮岛保留（不是萌宠小游戏），军事感清零（不筛贵价SLG用户）。与图5分工：图1卖家园氛围「想住进去」，图5卖升级机制「有深度」。',
     '物料\\首图-龙火鹿.png\n物料\\首图-豪木熊.png\n物料\\首图-帝冰鹅.png\n物料\\首图-幽奈娅.png\n物料\\建造-主堡Lv1.png / Lv7.png\n物料\\建造-主堡内城.png\n物料\\角色-【帕基S】角色资源汇总.jpg\n共享盘：最新建筑0409\\（农场/伐木场/晶矿场1-5、餐厅、温泉等）',
     '首图参考-2.png（构图参考，去军事化）\n首图参考-1.png（明亮色调参考）'],
    ['2','野外抓宠','探索并捕捉！','EXPLORE\n& CATCH!',
     '1.主体只有一个动作——抓捕瞬间：明亮野外（森林边缘/草原/湖边），训练师弗兰克背影抛出发光捕捉球，球在半空中拖螺旋光轨即将命中一只受惊的野生帕基（幼火鹿或戏水喵），球不闭合、结局未知的悬念感。\n2.训练师只给背影/手部，不抢帕基的戏；野生帕基要有受惊/警觉动态，与球形成对视。\n3.远景隐约2-3只其他野生帕基剪影（暗示世界到处有帕基可抓）。\n4.色调明亮鲜艳白天，丁达尔光。\n5.文案置底部15-20%，不遮挡抓捕动作。',
     '抓捕是宝可梦品类的第一动作钩子，值得独占一张（旧版把抓捕塞在拼图角落太弱）。抓宠+经营联动是全素材方向CVR最高（26-32%），最强钩子放大。素材齐全：训练师弗兰克、野生帕基、UI\\C_宠物捕捉.psd可参考真实捕捉界面。',
     '物料\\角色-训练师弗兰克.png\n物料\\进化-幼火鹿-一阶.png\n物料\\角色-【帕基S】角色资源汇总.jpg\nUI参考（共享盘）：UI\\C_宠物捕捉.psd',
     '参考-1.png（抓捕球瞬间参考）\n※无完整参考，用AI提示词生成新参考'],
    ['3','进化+收集','进化并收集！','EVOLVE\n& COLLECT',
     '1.画面中部：火鹿三阶进化链（幼火鹿→萤火鹿→龙火鹿，按物料PNG）阶梯式从左下到右上排列，发光箭头连接，每阶站发光石台，体型/装饰/光效逐级明显升级，进化粒子飞舞。\n2.顶部：半透明图鉴UI面板，收集进度45/200+，已解锁明亮、未解锁灰色剪影。\n3.明亮魔法森林场景，色彩鲜艳和谐，蝴蝶发光孢子。\n4.稀有度标签如出现只用 Common / Rare / Legendary。\n5.文案置顶部或底部，不遮挡进化链。',
     '「进化」是宠物品类最强好奇心钩子（它会变成什么样→我也想养）；图鉴进度制造收集目标感（45/200+=未完成任务心理）。原图2拆出抓捕后更聚焦，一张图只讲一个钩子。',
     '物料\\进化-幼火鹿-一阶.png\n物料\\进化-萤火鹿-二阶.png\n物料\\进化-龙火鹿-三阶.png\n共享盘更多进化链：帕基截图\\（火熊/木熊/木少女三阶等52张）\nUI参考（共享盘）：UI\\C_宠物进化3.psd',
     '参考-1a.png（主力，进化链+图鉴）\n参考-1.png（备选，层级结构）'],
    ['4','孵蛋繁育','孵化并繁育！','HATCH\n& BREED!',
     '1.画面中央：一颗巨大发光帕基蛋正在裂开，裂缝透出耀眼金光，一只幼年帕基探出脑袋的破壳瞬间，蛋壳碎片飞溅。\n2.两侧：两只成年帕基（如幽奈娅系+水巴拉系，按物料形象）头碰头，头顶浮起爱心光效（表达繁育）。\n3.背景：温馨孵化室——干草窝、暖黄小灯、一排等待孵化的蛋（暗示还有很多惊喜）。\n4.暖色调：橙+粉+金，全组最软最暖的一张，治愈感。\n5.文案置底部15-20%。',
     '①繁育孵化是宝可梦用户重点追求（之前下单文档明确结论），功能已实装（V2商店图有孵蛋位）。②破壳瞬间=盲盒心理，「蛋里会孵出什么」与抽卡同一多巴胺回路，是休闲受众最强下载冲动之一。③拉宽受众：对偏女性/休闲用户吸引力强于战斗，正好服务拉低单价目标。④补全宠物循环（抓→养→生→收集），漏斗节奏：夹在进化与经营之间色调过渡刚好。',
     '物料\\角色-【帕基S】角色资源汇总.jpg\n共享盘：帕基截图\\（幼年形态帕基多只）\n旧版参考（共享盘）：素材成品\\平面成品\\商店物料\\2026-06\\商店图\\7-孵蛋.jpg\n（宠物蛋无现成资产，可泛化设计）',
     '※无现成参考图\n用AI提示词生成新参考\n旧版7-孵蛋.jpg可给美术看'],
    ['5','基地建造经营','建造并经营你的庇护所','BUILD & MANAGE\nYOUR SANCTUARY',
     '1.核心视觉：主堡升级前后对比（用物料 建造-主堡Lv1 vs Lv7），光效箭头+等级标识连接，升级变化要明显。\n2.帕基在基地干活：采集/搬运/巡逻，展示「分配帕基做任务」机制（帕基按物料形象）。\n3.减少军事要塞感（少城墙/兵营），增加生活设施：训练场/孵化室/温泉/帕基互动区（可泛化）。\n4.暖色调：森林绿+木色+丰收金，点缀少量魔法蓝绿。\n5.角落轻量UI：资源数量（金币/木材/石材）+建造进度条，不抢主视觉。',
     '建造经营是SLG付费核心，回答「这游戏能玩多久」——模拟经营CVR 15-23%，4月测试建造类素材留存最好。放第5张：宠物钩子已完成转化任务，这张开始筛选高LTV用户。',
     '物料\\建造-主堡Lv1.png / Lv7.png\n物料\\建造-伐木场Lv1.png / Lv5.png\n物料\\建造-农场Lv1.png / Lv5.png\n物料\\建造-晶矿场Lv1.png / Lv5.png\n物料\\建造-主堡内城.png\n共享盘：最新建筑0409\\（石矿场1-5、温泉、餐厅、研究所、医院、募兵所、指挥所等）',
     '参考-2a.png（主力85%）\n参考-2.png（辅助75%）\n参考-2b.png（备选70%）'],
    ['6','策略团队战斗','训练你的帕基战队','TRAIN YOUR PALKI\n& CONQUER',
     '1.3-5只自家战斗帕基（炽角兽/跃焰虎等，按物料PNG）组成战队 vs 一只巨型BOSS（泛化原创设计，体型有压迫感，HP血条被削减中如60%）。\n2.暗色背景（深紫/暗蓝）+高亮技能特效（火焰橙/雷电金/冰霜青）：粒子、冲击波、光轨，特效必须华丽，静态图也要有动感。\n3.轻量战斗UI：底部或侧边队伍头像+技能图标，不遮挡战斗主体。\n4.帕基表情专注战斗，禁止卖萌。\n5.若游戏内战斗表现力不足，需UE渲染加特效后再截图。',
     '战斗CVR方差极大（好素材25% vs 差素材12%），视觉品质直接=转化率。放第6张：是全组视觉冲击力最强的一张，为深度内容收尾、给犹豫用户最后一推。',
     '物料\\战斗-炽角兽.png\n物料\\战斗-跃焰虎.png\n共享盘：帕基截图\\13【跃焰虎】10132（二阶）、帕基技能\\（技能特效参考）\n（BOSS无现成资产，泛化原创设计）',
     '参考-3.png（主力90%）'],
    ['7','联盟大世界征服','结盟一起征服','ALLY & CONQUER\nTOGETHER',
     '1.俯瞰视角大世界地图。⚠ 大世界地图无现成2D资产（UE场景）：优先游戏内/UE截图，不行则美术按参考-4原创重绘。多地形：山脉/河流/沙漠/森林/火山。\n2.四方联盟用鲜明颜色+旗帜区分（红/蓝/绿/棕），领土边界清晰；联盟名/旗帜可泛化。\n3.中央争议区（Disputed Zone）战斗爆炸，多色行军箭头交汇。\n4.右下角联盟UI弹窗：联盟名+在线成员145/200（示意）+ WAR! 红按钮 + JOIN 蓝按钮。\n5.金色+深色地图底色，史诗感。',
     '联盟社交=SLG终极留存与付费发动机（WOS用户评价反复验证；幻宠联盟系统已实装）。SLG元素集中在最后一张：不用来导量（SLG受众贵），只做「这是有长期价值的大世界」的承诺，同时打破「单机小游戏」错觉。',
     '物料\\建造-主堡大世界外观.png（仅主堡外观）\n物料\\建造-主堡内城.png\n物料\\角色-【帕基S】角色资源汇总.jpg\n⚠ 无世界地图2D资产，需游戏内截图或美术原创重绘\nUI参考（共享盘）：UI\\L_联盟.psd、S_世界行军战斗.psd',
     '参考-4.png（主力95%）\n参考-4b.png（辅助95%）'],
]
for i, row in enumerate(rows, start=11):
    for j, v in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=v)
        c.font = body_font; c.alignment = wrap; c.border = border
    ws.row_dimensions[i].height = 250

for j, w in enumerate([5, 15, 14, 20, 72, 42, 42, 23, 24], 1):
    ws.column_dimensions[get_column_letter(j)].width = w

ref_imgs = {11: '首图参考-2.png', 12: '参考-1.png', 13: '参考-1a.png', 15: '参考-2a.png', 16: '参考-3.png', 17: '参考-4.png'}
for r, fname in ref_imgs.items():
    tp, w, h = thumb(fname)
    img = XLImage(tp)
    img.width, img.height = w, h
    ws.add_image(img, f'H{r}')

# ================= Sheet 2: AI生图提示词 =================
ws2 = wb.create_sheet('AI生图提示词')
ws2['A1'] = 'AI 生图提示词（7张全套）—— 用于生成构图/氛围参考图'
ws2['A1'].font = title_font
ws2.merge_cells('A1:G1')

notes2 = [
    '使用说明：AI 画不准我们的帕基形象，生成图只看构图/色调/氛围/元素布局；角色形象正式图由美术按物料PNG执行。文案区域留空，后期合成（AI写字不可靠）。',
    '参数建议：即梦/可灵选 9:16 或 1080×1920；Midjourney 在提示词末尾加 --ar 9:16 --style raw。通用负面词：文字、水印、士兵军队（除图7）、已有IP角色（皮卡丘/宝可梦/幻兽帕鲁）、2D平涂、夜景霓虹（除图6）。',
    '验收看什么：构图是否符合需求说明、色调对不对、关键元素齐不齐——不要看角色像不像我们的帕基。',
]
for i, t in enumerate(notes2, start=2):
    c = ws2.cell(row=i, column=1, value=t)
    c.font = note_font; c.alignment = wrap
    ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)
    ws2.row_dimensions[i].height = 30

headers2 = ['图号','主题','中文提示词（即梦/可灵/通义万相）','英文提示词（Midjourney/SD/Flux）','专属负面词','参数建议','生成后验收重点']
for j, h in enumerate(headers2, 1):
    c = ws2.cell(row=5, column=j, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = center; c.border = border
ws2.row_dimensions[5].height = 28

STYLE_CN = '竖版9:16，3D卡通渲染风格，万龙觉醒/Palmon Survival式明亮奇幻风，中高饱和度，柔和全局光照，材质细节精致，角色比例偏Q但不低幼，高品质手游商店宣传图，全部原创奇幻宠物形象'
STYLE_EN = 'vertical 9:16, 3D stylized cartoon render, Call-of-Dragons / Palmon-Survival style bright fantasy, medium-high saturation, soft global illumination, detailed materials, slightly chibi but not childish proportions, premium mobile game store art, all original fantasy creature designs'

prompts = [
    ['1','帕基田园家园',
     STYLE_CN + '。前景：绿草山坡上4只原创奇幻宠物——火焰鬃毛鹿角神鹿、木甲战熊、冰晶企鹅、森林精灵少女，自信伙伴姿态面向观众。中景：繁荣田园家园——金黄色农田、伐木场原木堆、发光晶矿洞、蓝色小河与小码头穿流而过，多只宠物在浇水种田、砍树、采矿、搬运木材，忙碌欢乐。远景：山巅宏伟魔法城堡、漂浮绿岛、瀑布、雪山、柔和云海。明亮白天暖阳光，蓝天+翠绿+丰收金+湖水蓝色调，点缀魔法光效。三层纵深构图，电影感，史诗而治愈。画面底部15%留干净深色渐变区域供后期加标题。',
     STYLE_EN + '. Foreground: 4 original fantasy creatures on a grassy hill — fire-maned antler deer, wood-armored bear, crystal ice penguin, forest spirit girl — confident poses facing viewer. Midground: thriving pastoral homestead — golden farmland, lumber camp with stacked logs, glowing crystal mine, a blue river with small dock running through; cute pets watering crops, chopping wood, mining, hauling logs, busy and joyful. Background: majestic magic castle on mountain peak, floating green islands, waterfalls, snowy mountains, soft clouds. Bright daylight, warm sunlight, palette of sky blue + lush green + harvest gold + lake blue, subtle magic glow. Three-layer depth composition, cinematic, epic yet heartwarming. Bottom 15% clean dark gradient area reserved for title. --ar 9:16 --style raw',
     '士兵、军队、战争、硝烟、兵营、文字、水印、皮卡丘等IP角色',
     '9:16 / 1080×1920',
     '三层纵深是否清晰；农田/伐木/采矿/湖泊四元素是否齐全；有无军事元素混入；底部留白够不够放文案'],
    ['2','野外抓宠',
     STYLE_CN + '。明亮白天的野外草原与森林边缘。画面主体只有一个动作——抓捕瞬间：一名年轻冒险训练师的背影（短发、出行装），正抛出一颗发光的能量捕捉球，球在半空中拖着螺旋光轨，即将命中一只受惊的可爱火焰小鹿幼崽，小鹿瞪大眼睛回头，野草丛生野花点缀，蝴蝶惊飞，悬念感张力十足。远景隐约可见两三只其他奇幻小动物的剪影。阳光透过树冠形成丁达尔光，色彩明亮鲜艳。画面底部15%留干净区域供后期加标题。',
     STYLE_EN + '. Bright daylight meadow at forest edge. Single focal action — capture moment: a young adventurer trainer seen from behind (short hair, travel outfit), throwing a glowing energy capture orb mid-air with spiral light trail, about to catch a startled cute fire-deer fawn looking back with wide eyes, tall grass and wildflowers, startled butterflies, maximum suspense. Silhouettes of 2-3 other fantasy creatures in far background. God rays through tree canopy, vibrant bright colors. Bottom 15% clean area for title. --ar 9:16 --style raw',
     '文字、水印、黑夜、抓住后状态（球要在空中未闭合）、IP角色',
     '9:16 / 1080×1920',
     '球是否在空中未闭合（悬念感）；训练师是否只是背影不抢戏；帕基受惊动态是否到位'],
    ['3','进化+收集',
     STYLE_CN + '。明亮魔法森林空地上的进化仪式场景。画面中部：火焰鹿三阶进化链，从左下到右上阶梯式排列——可爱幼鹿、长出小火焰角的少年鹿、浑身烈焰鬃毛的威严神鹿，体型逐级变大，之间用发光箭头连接，每阶站在发光符文石台上，进化光效粒子飞舞。顶部：半透明游戏图鉴UI面板，显示收集进度45/200+，已解锁宠物图标明亮、未解锁为灰色剪影。周围蝴蝶与发光孢子，色彩鲜艳和谐，充满期待感与成就感。',
     STYLE_EN + '. Evolution ceremony in a bright enchanted forest clearing. Middle: three-stage fire deer evolution chain arranged diagonally upward — adorable fawn, juvenile with small flame antlers, majestic stag with blazing fire mane — growing larger each stage, connected by glowing arrows, each standing on a glowing rune stone platform, evolution light particles flying. Top: semi-transparent game collection UI panel showing "45/200+" progress, unlocked pet icons bright, locked ones grey silhouettes. Butterflies and glowing spores, vibrant harmonious colors, anticipation and achievement. --ar 9:16 --style raw',
     '文字错误、夜景、霓虹、水印、IP角色',
     '9:16 / 1080×1920',
     '三阶变化是否逐级明显；箭头连接是否清晰；顶部UI是否不抢主视觉'],
    ['4','孵蛋繁育',
     STYLE_CN + '。温馨暖色调的室内孵化室场景。画面中央：一颗巨大的发光奇幻宠物蛋正在裂开，裂缝中透出耀眼金色光芒，一只可爱的幼年奇幻宠物探出脑袋的破壳瞬间，蛋壳碎片飞溅。两侧：两只成年奇幻宠物（紫藤精灵狐与蓝色水豚）头碰头，头顶浮起爱心光效。背景：干草窝、暖黄小灯、一排等待孵化的蛋。柔和暖光，橙+粉+金色调，治愈温馨，充满惊喜感。画面底部15%留干净区域供后期加标题。',
     STYLE_EN + '. Cozy warm nursery hatchery interior. Center: a giant glowing fantasy egg cracking open, dazzling golden light bursting from the cracks, a cute baby creature peeking out at the hatching moment, shell fragments flying. Two adult fantasy creatures (a violet spirit fox and a blue capybara-like pet) touching heads on both sides, floating heart glow above them. Background: hay nests, warm yellow lamps, a row of eggs waiting to hatch. Soft warm light, orange + pink + gold palette, heartwarming and full of surprise. Bottom 15% clean area for title. --ar 9:16 --style raw',
     '文字、水印、冷色调、黑暗、IP角色',
     '9:16 / 1080×1920',
     '破壳瞬间的光效是否够惊艳；爱心元素是否表达了繁育；暖色调是否到位'],
    ['5','基地建造经营',
     STYLE_CN + '。奇幻宠物基地建造经营场景。核心视觉：同一栋主堡建筑升级前后对比——左侧简陋小木屋（Lv.1），右侧宏伟蓝金魔法城堡（Lv.7），中间用发光箭头与等级标识连接，升级变化明显。中景：多只原创奇幻宠物在基地干活——砍树、采矿、搬运建材、巡逻。角落：轻量资源UI（金币/木材/石材图标）与建造进度条。暖色调：森林绿+木色+丰收金，点缀少量魔法蓝绿光效，繁荣有序有生活气息。画面底部15%留干净区域供后期加标题。',
     STYLE_EN + '. Fantasy pet base building and management scene. Core visual: before/after upgrade comparison of the same main castle — left a shabby wooden hut (Lv.1), right a magnificent blue-gold magic castle (Lv.7), connected by a glowing arrow with level badges, dramatic upgrade. Midground: original fantasy pets working in the base — chopping trees, mining, hauling materials, patrolling. Corner: light resource UI (gold/wood/stone icons) and construction progress bar. Warm palette: forest green + wood brown + harvest gold with hints of magic teal, prosperous and lively. Bottom 15% clean area for title. --ar 9:16 --style raw',
     '士兵、军队、战争、城墙要塞感过重、文字、水印',
     '9:16 / 1080×1920',
     '升级前后对比是否一眼可辨；帕基干活元素在不在；军事感是否过重'],
    ['6','策略团队战斗',
     STYLE_CN + '。暗色调史诗战斗场景。画面中央：4只原创奇幻宠物组成战队（火焰独角兽、烈焰虎、冰晶企鹅、木甲熊）围攻一只巨型原创魔兽BOSS，BOSS体型庞大有压迫感，头顶HP血条剩余60%。技能特效轰炸：火焰风暴、雷电冲击、冰霜箭雨、能量光束，粒子、冲击波、光轨充满画面，静态图也有强烈动感。深紫暗蓝背景+高亮技能特效形成最强对比。底部轻量战斗UI：队伍头像与技能图标。画面顶部留干净区域供后期加标题。',
     STYLE_EN + '. Dark epic battle scene. Center: a team of 4 original fantasy pets (fire unicorn, flame tiger, ice penguin, wood-armored bear) besieging a giant original monster boss, boss huge and oppressive with an HP bar at 60% above its head. Skill effect barrage: firestorm, lightning strikes, frost arrow rain, energy beams — particles, shockwaves and light trails filling the frame, strong motion even in a static image. Deep purple / dark blue background against high-brightness skill effects for maximum contrast. Light battle UI at bottom: team portraits and skill icons. Top clean area for title. --ar 9:16 --style raw',
     '文字、水印、明亮白天背景、卖萌表情、IP角色',
     '9:16 / 1080×1920',
     '暗背景+亮特效对比是否够强；团队围攻感（不是1v1）；特效华丽度'],
    ['7','联盟大世界征服',
     STYLE_CN + '。俯瞰视角的奇幻世界战略地图。多地形：雪山、喷发的火山、沙漠、森林、河流。四方联盟用红/蓝/绿/棕旗帜与颜色区分领土，边界清晰，各有风格迥异的城堡要塞。中央紫色争议区有战斗与爆炸效果，多条不同颜色的行军箭头交汇于此。地图上分布原创奇幻生物与军队单位。右下角：联盟UI弹窗——联盟名、在线成员145/200、红色WAR!按钮与蓝色JOIN按钮。金色+深色地图底色，阴云与火山光晕，史诗战争氛围。',
     STYLE_EN + '. Top-down fantasy world strategy map. Multiple terrains: snowy mountains, erupting volcano, desert, forests, rivers. Four alliances distinguished by red/blue/green/brown banners and territory colors, clear borders, each with distinct castles. Central purple disputed zone with battles and explosions, multiple colored marching arrows converging. Original fantasy creatures and army units scattered across the map. Bottom-right alliance UI panel: alliance name, 145/200 online members, red WAR! button and blue JOIN button. Gold + dark map palette, storm clouds and volcanic glow, epic war atmosphere. --ar 9:16 --style raw',
     '文字错误（联盟名可后期合成）、皮卡丘等IP角色、水印',
     '9:16 / 1080×1920',
     '四方联盟颜色是否可区分；争议区+行军箭头是否清晰；右下角UI位置是否预留'],
]
for i, row in enumerate(prompts, start=6):
    for j, v in enumerate(row, 1):
        c = ws2.cell(row=i, column=j, value=v)
        c.font = body_font; c.alignment = wrap; c.border = border
    ws2.row_dimensions[i].height = 230

for j, w in enumerate([5, 14, 62, 62, 24, 14, 30], 1):
    ws2.column_dimensions[get_column_letter(j)].width = w

wb.save(OUT)
print('saved:', OUT)
