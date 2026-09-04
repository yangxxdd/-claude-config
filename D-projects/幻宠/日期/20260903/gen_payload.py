# -*- coding: utf-8 -*-
"""解析 6月 数据，生成 lark-cli 写入 payload JSON 文件（不做任何 CLI 调用）"""
import openpyxl, json

BASE = "MPN4b2YpKa9pdAshFcucTIAEnnd"
SRC = '源数据/幻宠fb媒体-源数据.xlsx'

# ---------- 素材方向映射 ----------
DIR_MAP = {
    '宠物展示': '宠物展示', '幻想宠物': '宠物展示',
    '抓宠经营': '抓宠经营',
    '模拟经营': '模拟经营',
    '帕萌战斗': '战斗',
    '抓宠战斗': '抓宠战斗',
    '高效抓宠': '抓宠', '收集升级': '抓宠',
    '天灾重建': '天灾/生存', '海啸': '天灾/生存', '灾后重建': '天灾/生存',
    '场景展示': '场景展示', '帕基世界冒险': '场景展示',
    '核心玩法': '核心玩法',
}
DIRS = ['宠物展示','抓宠经营','模拟经营','战斗','抓宠战斗','天灾/生存','场景展示','核心玩法','抓宠']
HUES = ['Green','Lime','Blue','Carmine','Orange','Yellow','Turquoise','Purple','Green']

def direction(name):
    parts = name.split('-')
    return DIR_MAP.get(parts[1], parts[1]) if len(parts) >= 2 else name

def mat_type(name):
    return 'V视频' if name.startswith('V') else 'P图片'

# ---------- 解析 Meta ----------
wb = openpyxl.load_workbook(SRC, data_only=True)
def col_index(hdr): return {n:i for i,n in enumerate(hdr) if n}

meta, ad2name, ad2bid = {}, {}, {}
for sn, bid in [('6月install美国','install'),('6月AEO美国','AEO')]:
    ws = wb[sn]; rows = list(ws.iter_rows(values_only=True)); idx = col_index(rows[0])
    for r in rows[1:]:
        name = r[idx['广告名称']]
        if not name: continue
        imp = r[idx.get('展示次数')] or 0
        clk = r[idx.get('点击量（全部）')] or 0
        ins = r[idx.get('移动应用安装量')] or 0
        spend = r[idx.get('已花费金额 (USD)')] or 0
        adid = r[idx.get('广告编号')]
        d = meta.setdefault((bid, name), {'展示':0,'点击':0,'安装':0,'花费':0.0})
        d['展示']+=imp; d['点击']+=clk; d['安装']+=ins; d['花费']=round(d['花费']+spend,2)
        if adid:
            ad2name[str(adid)] = name; ad2bid[str(adid)] = bid

# ---------- 解析 BI ----------
ws = wb['BI汇总']; rows = list(ws.iter_rows(values_only=True)); idx = col_index(rows[0])
bi = {}
for r in rows[1:]:
    sid = r[idx['素材ID']]
    if sid is None: continue
    sid = str(sid)
    g = lambda c: int(r[idx[c]]) if r[idx[c]] else 0
    d = bi.setdefault(sid, {'dnu':0,'r1':0,'r3':0,'r7':0})
    d['dnu']+=g('dnu'); d['r1']+=g('r1_cnt'); d['r3']+=g('r3_cnt'); d['r7']+=g('r7_cnt')

# ---------- 构造素材日报 rows ----------
fields = ["日期","出价方式","国家","素材类型","素材名称","素材方向","标记",
          "m_花费","m_展示","m_点击","m_CTR%","m_CPC","m_CVR%","m_CPM","m_安装","m_CPI",
          "b_安装","b_r1_cnt","b_r3_cnt","b_r7_cnt",
          "6月CPI","6月CPM","6月CTR%","6月CVR%","6月D1%","6月D3%"]
rows, summary = [], {'install':{}, 'AEO':{}}
for (bid, name), d in sorted(meta.items()):
    imp, clk, ins, spend = d['展示'], d['点击'], d['安装'], d['花费']
    ctr = round(clk/imp*100, 2) if imp else 0
    cpc = round(spend/clk, 2) if clk else 0
    cvr = round(ins/clk*100, 2) if clk else 0
    cpm = round(spend/imp*1000, 2) if imp else 0
    cpi = round(spend/ins, 2) if ins else 0
    adids = [a for a,n in ad2name.items() if n==name and ad2bid[a]==bid]
    dnu = sum(bi.get(a,{}).get('dnu',0) for a in adids)
    r1 = sum(bi.get(a,{}).get('r1',0) for a in adids)
    r3 = sum(bi.get(a,{}).get('r3',0) for a in adids)
    r7 = sum(bi.get(a,{}).get('r7',0) for a in adids)
    d1 = round(r1/dnu*100,2) if dnu else 0
    d3 = round(r3/dnu*100,2) if dnu else 0
    rows.append(["2026-06-30 00:00:00", bid, "美国", mat_type(name), name, direction(name), "🟡 观察",
                 spend, imp, clk, ctr, cpc, cvr, cpm, ins, cpi,
                 dnu, r1, r3, r7, cpi, cpm, ctr, cvr, d1, d3])
    s = summary[bid]
    for k, v in [('花费',spend),('展示',imp),('点击',clk),('安装',ins),('dnu',dnu),('r1',r1),('r3',r3),('r7',r7)]:
        s[k] = s.get(k,0) + v

# ---------- 写文件 ----------
opt_payload = {"name":"素材方向","type":"select","multiple":False,
               "options":[{"name":d,"hue":HUES[i],"lightness":"Lighter"} for i,d in enumerate(DIRS)]}
json.dump(opt_payload, open('日期/20260903/_field_options.json','w',encoding='utf-8'), ensure_ascii=False)
json.dump({"fields":fields,"rows":rows}, open('日期/20260903/_rows.json','w',encoding='utf-8'), ensure_ascii=False)
json.dump(summary, open('日期/20260903/_summary.json','w',encoding='utf-8'), ensure_ascii=False)
print(f'素材日报 rows: {len(rows)} (install {sum(1 for (b,_) in meta if b=="install")}, AEO {sum(1 for (b,_) in meta if b=="AEO")})')
print('files written: _field_options.json, _rows.json, _summary.json')
