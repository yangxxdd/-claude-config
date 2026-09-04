# -*- coding: utf-8 -*-
"""用 6月 数据干跑填充幻宠日报模板：更新素材方向选项 + 填素材日报/每日汇总/财务汇总"""
import openpyxl, json, subprocess, sys

BASE = "MPN4b2YpKa9pdAshFcucTIAEnnd"
TBL_MATERIAL = "tbl5CUjOIX9OWKbb"   # 素材日报
SRC = '源数据/幻宠fb媒体-源数据.xlsx'

def lark(*args):
    r = subprocess.run(['lark-cli', *args], capture_output=True, text=True, encoding='utf-8')
    return r.stdout, r.stderr

# ---------- 1. 素材方向：从素材名推导 + 更新字段选项 ----------
# 方向映射（素材名中段 -> 归一化方向）
DIR_MAP = {
    '宠物展示': '宠物展示', '幻想宠物': '宠物展示',
    '抓宠经营': '抓宠经营',
    '模拟经营': '模拟经营',
    '帕萌战斗': '战斗',
    '抓宠战斗': '抓宠战斗',
    '高效抓宠': '抓宠', '收集升级': '抓宠',
    '天灾重建': '天灾/生存', '海啸': '天灾/生存', '灾后重建': '天灾/生存',
    '场景展示': '场景展示', '帕基世界冒险': '场景展示',
    '核心玩法': '核心玩法',
}
DIRS = ['宠物展示','抓宠经营','模拟经营','战斗','抓宠战斗','天灾/生存','场景展示','核心玩法','抓宠']
HUES = ['Green','Lime','Blue','Carmine','Orange','Yellow','Turquoise','Purple','Green']
options = [{"name": d, "hue": HUES[i], "lightness": "Lighter"} for i, d in enumerate(DIRS)]
opt_json = json.dumps({"name":"素材方向","type":"select","multiple":False,"options":options}, ensure_ascii=False)
out, err = lark('base','+field-update','--base-token',BASE,'--table-id',TBL_MATERIAL,
                '--field-id','fldtKmrm3v','--json',opt_json,'--yes')
print('更新素材方向:', 'OK' if '"ok": true' in out else 'FAIL', err[:200])

def direction(name):
    # name 形如 'V-抓宠战斗' / 'P-宠物展示-超梦'
    parts = name.split('-')
    if len(parts) >= 2:
        seg = parts[1]
        return DIR_MAP.get(seg, seg)
    return name

def mat_type(name):
    return 'V视频' if name.startswith('V') else 'P图片'

# ---------- 2. 解析 + 聚合 6月 Meta ----------
wb = openpyxl.load_workbook(SRC, data_only=True)
def col_index(hdr):
    return {n:i for i,n in enumerate(hdr) if n}

meta = {}
ad2name = {}
ad2bid = {}
for sn, bid in [('6月install美国','install'),('6月AEO美国','AEO')]:
    ws = wb[sn]; rows = list(ws.iter_rows(values_only=True)); idx = col_index(rows[0])
    for r in rows[1:]:
        name = r[idx['广告名称']]
        if not name: continue
        imp = r[idx.get('展示次数')] or 0
        clk = r[idx.get('点击量（全部）')] or 0
        ins = r[idx.get('移动应用安装量')] or 0
        spend = r[idx.get('已花费金额 (USD)')] or 0
        adid = r[idx.get('广告编号')]
        key = (bid, name)
        d = meta.setdefault(key, {'展示':0,'点击':0,'安装':0,'花费':0.0})
        d['展示']+=imp; d['点击']+=clk; d['安装']+=ins; d['花费']=round(d['花费']+spend,2)
        if adid:
            ad2name[str(adid)] = name; ad2bid[str(adid)] = bid

# BI
ws = wb['BI汇总']; rows = list(ws.iter_rows(values_only=True)); idx = col_index(rows[0])
bi = {}
for r in rows[1:]:
    sid = r[idx['素材ID']]
    if sid is None: continue
    sid = str(sid)
    g = lambda c: int(r[idx[c]]) if r[idx[c]] else 0
    d = bi.setdefault(sid, {'dnu':0,'r1':0,'r3':0,'r7':0})
    d['dnu']+=g('dnu'); d['r1']+=g('r1_cnt'); d['r3']+=g('r3_cnt'); d['r7']+=g('r7_cnt')

# ---------- 3. 构造 素材日报 rows ----------
fields = ["日期","出价方式","国家","素材类型","素材名称","素材方向","标记",
          "m_花费","m_展示","m_点击","m_CTR%","m_CPC","m_CVR%","m_CPM","m_安装","m_CPI",
          "b_安装","b_r1_cnt","b_r3_cnt","b_r7_cnt",
          "6月CPI","6月CPM","6月CTR%","6月CVR%","6月D1%","6月D3%"]
rows = []
summary = {'install':{}, 'AEO':{}}
for (bid, name), d in sorted(meta.items()):
    imp, clk, ins, spend = d['展示'], d['点击'], d['安装'], d['花费']
    ctr = round(clk/imp*100, 2) if imp else 0
    cpc = round(spend/clk, 2) if clk else 0
    cvr = round(ins/clk*100, 2) if clk else 0
    cpm = round(spend/imp*1000, 2) if imp else 0
    cpi = round(spend/ins, 2) if ins else 0
    # BI
    adids = [a for a,n in ad2name.items() if n==name and ad2bid[a]==bid]
    dnu = sum(bi.get(a,{}).get('dnu',0) for a in adids)
    r1 = sum(bi.get(a,{}).get('r1',0) for a in adids)
    r3 = sum(bi.get(a,{}).get('r3',0) for a in adids)
    r7 = sum(bi.get(a,{}).get('r7',0) for a in adids)
    d1 = round(r1/dnu*100,2) if dnu else 0
    d3 = round(r3/dnu*100,2) if dnu else 0
    row = ["2026-06-30 00:00:00", bid, "美国", mat_type(name), name, direction(name), "🟡 观察",
           spend, imp, clk, ctr, cpc, cvr, cpm, ins, cpi,
           dnu, r1, r3, r7,
           cpi, cpm, ctr, cvr, d1, d3]
    rows.append(row)
    # 累计用于每日汇总
    s = summary[bid]
    for k in ['花费','展示','点击','安装','dnu','r1','r3','r7']:
        s[k] = s.get(k,0) + (locals().get(k,0) if k in locals() else 0)
    for k, v in [('花费',spend),('展示',imp),('点击',clk),('安装',ins),('dnu',dnu),('r1',r1),('r3',r3),('r7',r7)]:
        s[k] = s.get(k,0) + v

payload = json.dumps({"fields": fields, "rows": rows}, ensure_ascii=False)
out, err = lark('base','+record-batch-create','--base-token',BASE,'--table-id',TBL_MATERIAL,'--json',payload)
print('填素材日报:', 'OK' if '"ok": true' in out else 'FAIL', 'rows=', len(rows))
if 'FAIL' in out or '"ok": true' not in out:
    print(out[:1500]); print(err[:500])

# ---------- 4. 每日汇总 + 财务汇总 ----------
print('\n每日汇总数据:')
for bid in ['install','AEO']:
    s = summary[bid]
    cpi = round(s['花费']/s['安装'],2) if s['安装'] else 0
    ctr = round(s['点击']/s['展示']*100,2) if s['展示'] else 0
    cpc = round(s['花费']/s['点击'],2) if s['点击'] else 0
    cvr = round(s['安装']/s['点击']*100,2) if s['点击'] else 0
    cpm = round(s['花费']/s['展示']*1000,2) if s['展示'] else 0
    r1p = round(s['r1']/s['dnu']*100,2) if s['dnu'] else 0
    r3p = round(s['r3']/s['dnu']*100,2) if s['dnu'] else 0
    r7p = round(s['r7']/s['dnu']*100,2) if s['dnu'] else 0
    print(f'{bid}: 花费={s["花费"]} 安装={s["安装"]} CPI={cpi} CTR={ctr}% CVR={cvr}% CPM={cpm} dnu={s["dnu"]} r1={r1p}% r3={r3p}% r7={r7p}%')
    # 存供后续填充
    s['cpi'],s['ctr'],s['cpc'],s['cvr'],s['cpm'] = cpi,ctr,cpc,cvr,cpm
    s['r1p'],s['r3p'],s['r7p'] = r1p,r3p,r7p

# 序列化 summary 供下一步
with open('日期/20260903/_summary.json','w',encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False)
print('\nsummary saved')
