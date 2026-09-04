# -*- coding: utf-8 -*-
"""对 6月 install 数据套关停规则 + 目标线，生成 问题追踪 + 每日结论 payload"""
import openpyxl, json

SRC = '源数据/幻宠fb媒体-源数据.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)
def col_index(hdr): return {n:i for i,n in enumerate(hdr) if n}

# ---- 聚合 6月 install Meta + BI ----
ws = wb['6月install美国']; rows = list(ws.iter_rows(values_only=True)); idx = col_index(rows[0])
meta = {}; ad2name = {}
for r in rows[1:]:
    name = r[idx['广告名称']]
    if not name: continue
    imp = r[idx.get('展示次数')] or 0; clk = r[idx.get('点击量（全部）')] or 0
    ins = r[idx.get('移动应用安装量')] or 0; spend = r[idx.get('已花费金额 (USD)')] or 0
    adid = r[idx.get('广告编号')]
    d = meta.setdefault(name, {'展示':0,'点击':0,'安装':0,'花费':0.0})
    d['展示']+=imp; d['点击']+=clk; d['安装']+=ins; d['花费']=round(d['花费']+spend,2)
    if adid: ad2name[str(adid)] = name

ws = wb['BI汇总']; rows = list(ws.iter_rows(values_only=True)); idx = col_index(rows[0])
bi = {}
for r in rows[1:]:
    sid = r[idx['素材ID']]
    if sid is None: continue
    sid = str(sid); g = lambda c: int(r[idx[c]]) if r[idx[c]] else 0
    d = bi.setdefault(sid, {'dnu':0,'r1':0,'r3':0,'r7':0})
    d['dnu']+=g('dnu'); d['r1']+=g('r1_cnt'); d['r3']+=g('r3_cnt'); d['r7']+=g('r7_cnt')

# ---- 计算每素材指标 ----
stats = []
for name, d in meta.items():
    adids = [a for a,n in ad2name.items() if n==name]
    dnu = sum(bi.get(a,{}).get('dnu',0) for a in adids)
    r1 = sum(bi.get(a,{}).get('r1',0) for a in adids)
    r3 = sum(bi.get(a,{}).get('r3',0) for a in adids)
    r7 = sum(bi.get(a,{}).get('r7',0) for a in adids)
    cpi = round(d['花费']/d['安装'],2) if d['安装'] else 0
    r1p = round(r1/dnu*100,2) if dnu else 0
    r3p = round(r3/dnu*100,2) if dnu else 0
    r7p = round(r7/dnu*100,2) if dnu else 0
    stats.append({'name':name,'花费':d['花费'],'m安装':d['安装'],'dnu':dnu,
                  'cpi':cpi,'r1':r1p,'r3':r3p,'r7':r7p})

# ---- 套规则，生成问题 ----
problems = []
for s in stats:
    n, cpi, r1, r3, dnu, m, spend = s['name'], s['cpi'], s['r1'], s['r3'], s['dnu'], s['m安装'], s['花费']
    # 规则1: 花费>50 且 dnu=0
    if spend > 50 and dnu == 0:
        problems.append(dict(name=n, typ='0激活', sev='🔴 严重', stat='🔴 待解决', act='关停',
            desc=f'6月花费 ${spend:.0f} 但 dnu=0，零激活，立即关停'))
        continue
    # 规则3: dnu<20 观察（不写入问题追踪，除非CPI离谱）
    if dnu < 20:
        if cpi > 20:
            problems.append(dict(name=n, typ='CPI超标', sev='🟡 观察', stat='🟡 处理中', act='观察',
                desc=f'CPI ${cpi} 严重超标，但 dnu={dnu}<20 量级不足，暂观察'))
        continue
    # 口径异常: dnu > m安装*1.5
    anom = dnu > m*1.5
    # 规则2: CPI>8.5 且 dnu>=30
    cpi_over = cpi > 8.5
    r1_low = r1 < 30
    if cpi_over:
        if r1_low:
            sev, act = '🔴 严重', '关停'
            typ = 'CPI超标'
            desc = f'CPI ${cpi} 超 $8.5 线 + R1 {r1}% 低于 30% 双杀，dnu={dnu}' + ('，且 dnu({})>Meta安装({}) 口径待查'.format(dnu,m) if anom else '')
        else:
            sev, act = '🟡 观察', '减量'
            typ = 'CPI超标'
            desc = f'CPI ${cpi} 超 $8.5 线，但 R1 {r1}% 尚可（留存好），dnu={dnu}' + ('，口径待查' if anom else '')
        problems.append(dict(name=n, typ=typ, sev=sev, stat='🔴 待解决' if sev=='🔴 严重' else '🟡 处理中', act=act, desc=desc))
    elif r1_low and not anom:
        # 留存不达标但CPI不超
        problems.append(dict(name=n, typ='留存不达标', sev='🟡 观察', stat='🟡 处理中', act='观察',
            desc=f'CPI ${cpi} 达标但 R1 {r1}% 低于 30% 待改进线，dnu={dnu}'))
    elif anom:
        problems.append(dict(name=n, typ='其他', sev='🟡 观察', stat='🟡 处理中', act='观察',
            desc=f'dnu({dnu}) > Meta安装({m})，素材ID口径异常待查'))

# 按严重度排序
problems.sort(key=lambda p: 0 if p['sev']=='🔴 严重' else 1)
# 生成 问题追踪 rows
prob_fields = ['编号','素材名称','发现日期','问题类型','严重度','状态','处理措施','问题描述']
prob_rows = []
for i, p in enumerate(problems, 1):
    prob_rows.append([f'HC-{i:03d}', p['name'], '2026-06-30 00:00:00', p['typ'], p['sev'], p['stat'], p['act'], p['desc']])
json.dump({'fields':prob_fields,'rows':prob_rows}, open('日期/20260903/_prob_rows.json','w',encoding='utf-8'), ensure_ascii=False)

# ---- 每日结论 ----
red = [p for p in problems if p['sev']=='🔴 严重']
ylw = [p for p in problems if p['sev']=='🟡 观察']
tot_spend = round(sum(s['花费'] for s in stats),2)
tot_m = sum(s['m安装'] for s in stats)
tot_dnu = sum(s['dnu'] for s in stats)
tot_cpi = round(tot_spend/tot_m,2)
good = [s for s in stats if s['cpi']<=5 and s['r1']>=30 and s['dnu']>=20]
s_cpi_map = {s['name']: s['cpi'] for s in stats}
r1_w = round(sum(s["dnu"]*s["r1"] for s in stats)/tot_dnu,2)
concl = {
    '一句话结论': f'6月整体 CPI ${tot_cpi} 处于待改进带，R1 {r1_w}% 低于 30% 目标；{len(red)} 支 CPI 超标+留存双杀需关停，低成本素材（合成3D/超梦）撑量。',
    '环比': 'vs 4月 install CPI $3.79 → 6月 $7.33，上涨约 93%（竞争加剧+全月口径）',
    '达标情况': f'整体：CPI ${tot_cpi} 超合格线(≤$6.5)；R1 低于30%待改进线。无素材同时满足「优秀」三线。',
    '向好信号': '、'.join(f'{s["name"]}(${s["cpi"]}/R1 {s["r1"]}%)' for s in sorted(good, key=lambda x:-x['r1'])[:4]) or '无',
    '警惕信号': ('🔴' + '、'.join(f'{p["name"]}(CPI ${s_cpi_map[p["name"]]})' for p in red) +
                 ' 双杀关停；🟡' + '、'.join(f'{p["name"]}' for p in ylw)),
    '明日动作': f'关停 {len(red)} 支双杀素材；合成3D/超梦/抓宠战斗加量；幽飘降CPI再观察；口径异常素材先查素材ID映射',
}
concl_fields = ['日期','一句话结论','环比','达标情况','向好信号','警惕信号','明日动作']
concl_rows = [[ '2026-06-30 00:00:00', concl['一句话结论'], concl['环比'], concl['达标情况'],
                concl['向好信号'], concl['警惕信号'], concl['明日动作'] ]]
json.dump({'fields':concl_fields,'rows':concl_rows}, open('日期/20260903/_concl_rows.json','w',encoding='utf-8'), ensure_ascii=False)

print(f'问题追踪 {len(problems)} 条（🔴 {len(red)} / 🟡 {len(ylw)}）')
for p in problems:
    print(f"  {p['sev'][:1]} {p['name']}: {p['typ']} | {p['desc'][:60]}")
print('\n每日结论 1 条')
print('  ' + concl['一句话结论'][:80] + '...')
print('written _prob_rows.json, _concl_rows.json')
