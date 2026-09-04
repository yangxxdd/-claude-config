# -*- coding: utf-8 -*-
"""解析 幻宠fb媒体-源数据.xlsx：Meta(m_) + BI(b_) 按素材名对齐聚合"""
import openpyxl, json, sys

SRC = '源数据/幻宠fb媒体-源数据.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)

def col_index(hdr):
    return {name: i for i, name in enumerate(hdr) if name}

# ---------- 解析 Meta（4 个 sheet） ----------
meta = {}  # 素材名 -> {m_展示, m_点击, m_安装, m_花费, m_CTR%加权, ...}
ad2name = {}  # 广告编号 -> 素材名
ad2bid = {}   # 广告编号 -> 出价方式(install/AEO)
meta_sheets = [
    ('4月install美国幻宠', 'install'),
    ('4月AEO美国幻宠', 'AEO'),
    ('6月install美国', 'install'),
    ('6月AEO美国', 'AEO'),
]
month_of = {'4月': '4月', '6月': '6月'}
for sn, bid in meta_sheets:
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    idx = col_index(rows[0])
    mon = sn[:2]  # '4月' / '6月'
    for r in rows[1:]:
        name = r[idx['广告名称']]
        if not name:
            continue
        imp = r[idx.get('展示次数')] or 0
        clk = r[idx.get('点击量（全部）')] or 0
        ins = r[idx.get('移动应用安装量')] or 0
        spend = r[idx.get('已花费金额 (USD)')] or 0
        adid = r[idx.get('广告编号')]
        key = (mon, bid, name)
        d = meta.setdefault(key, {'展示': 0, '点击': 0, '安装': 0, '花费': 0.0})
        d['展示'] += imp
        d['点击'] += clk
        d['安装'] += ins
        d['花费'] += round(spend, 2)
        if adid:
            ad2name[str(adid)] = name
            ad2bid[str(adid)] = bid

# ---------- 解析 BI 汇总 ----------
ws = wb['BI汇总']
rows = list(ws.iter_rows(values_only=True))
idx = col_index(rows[0])
bi = {}  # 素材ID -> {dnu, r1, r3, r7, r14}
for r in rows[1:]:
    sid = r[idx['素材ID']]
    if sid is None:
        continue
    sid = str(sid)
    def g(c):
        v = r[idx[c]]
        return int(v) if v else 0
    d = bi.setdefault(sid, {'dnu': 0, 'r1': 0, 'r3': 0, 'r7': 0, 'r14': 0})
    d['dnu'] += g('dnu')
    d['r1'] += g('r1_cnt')
    d['r3'] += g('r3_cnt')
    d['r7'] += g('r7_cnt')
    d['r14'] += g('r14_cnt')

# ---------- join 检查 ----------
meta_adids = set(ad2name.keys())
bi_ids = set(bi.keys())
matched = meta_adids & bi_ids
print(f'Meta 广告编号: {len(meta_adids)} | BI 素材ID: {len(bi_ids)} | 交集: {len(matched)}')
print(f'BI 有但 Meta 无: {len(bi_ids - meta_adids)} 个')
print(f'Meta 有但 BI 无: {len(meta_adids - bi_ids)} 个')

# ---------- 输出 6月 install + AEO 的 join 结果 ----------
print('\n===== 6月 install 按素材聚合（m_ + b_）=====')
print('素材名\tm_展示\tm_点击\tm_CTR%\tm_CVR%\tm_安装\tm_花费\tm_CPI\tb_dnu\tb_r1\tb_r3\tb_r7\tb_R1%\tb_R3%\tb_R7%')
for (mon, bid, name), d in sorted(meta.items()):
    if mon != '6月' or bid != 'install':
        continue
    ctr = d['点击']/d['展示']*100 if d['展示'] else 0
    cvr = d['安装']/d['点击']*100 if d['点击'] else 0
    cpi = d['花费']/d['安装'] if d['安装'] else 0
    # BI: 该素材名对应的所有广告编号的 BI 汇总
    adids = [a for a, n in ad2name.items() if n == name and ad2bid[a] == bid]
    dnu = sum(bi.get(a, {}).get('dnu', 0) for a in adids)
    r1 = sum(bi.get(a, {}).get('r1', 0) for a in adids)
    r3 = sum(bi.get(a, {}).get('r3', 0) for a in adids)
    r7 = sum(bi.get(a, {}).get('r7', 0) for a in adids)
    r1p = r1/dnu*100 if dnu else 0
    r3p = r3/dnu*100 if dnu else 0
    r7p = r7/dnu*100 if dnu else 0
    print(f'{name}\t{d["展示"]}\t{d["点击"]}\t{ctr:.2f}\t{cvr:.2f}\t{d["安装"]}\t{d["花费"]:.2f}\t{cpi:.2f}\t{dnu}\t{r1}\t{r3}\t{r7}\t{r1p:.2f}\t{r3p:.2f}\t{r7p:.2f}')
