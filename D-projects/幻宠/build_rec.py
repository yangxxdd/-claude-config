import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = load_workbook('creative_final.xlsx')

if '9月测试素材推荐' in wb.sheetnames:
    del wb['9月测试素材推荐']

ws = wb.create_sheet('9月测试素材推荐', 0)

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2F5496')
tier1_fill = PatternFill('solid', fgColor='C6EFCE')
tier2_fill = PatternFill('solid', fgColor='BDD7EE')
tier3_fill = PatternFill('solid', fgColor='FCE4D6')
tier4_fill = PatternFill('solid', fgColor='D9D9D9')
section_fill = PatternFill('solid', fgColor='4472C4')
section_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
bold_font = Font(name='Arial', bold=True, size=11)
normal_font = Font(name='Arial', size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

col_widths = [6, 7, 18, 10, 12, 10, 13, 12, 12, 12, 12, 14, 18, 50, 50, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Row 1: Title
ws.merge_cells('A1:P1')
ws['A1'] = '幻宠 9月测试 — 素材推荐（按"筛选 → 跑量 → 进付费Camp"策略）'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 30

# Row 2: Context
ws.merge_cells('A2:P2')
ws['A2'] = '测试时间：9月初 | 预估预算：$40-50K | 量级：~5K | 优化目标：Purchase | 策略：低CPI+高留存素材 → Install/AEO筛选 → 优胜者进Purchase Camp | 核心方向：捉宠/融合/进化（主推）+ 数据好的非核心方向（辅推）'
ws['A2'].font = Font(name='Arial', size=10, color='333333')
ws['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[2].height = 32

headers = ['优先级', '方向', '素材名称', '出价方式', '验证月份', 'DNU', 'CPI(USD)', 'R1', 'R2', 'R3',
           '次留成本', '4月CPI', '4月R1→6月R1', '推荐理由', '建议动作', '建议预算(USD)']

def write_section_header(ws, row, text):
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'] = text
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1, 17):
        ws.cell(row=row, column=c).fill = section_fill

def write_table_header(ws, row):
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[row].height = 35

def write_data_rows(ws, start_row, data, tier_fill):
    for i, d in enumerate(data):
        row = start_row + i
        ws.row_dimensions[row].height = 80
        for j, val in enumerate(d, 1):
            cell = ws.cell(row=row, column=j, value=val)
            cell.font = normal_font
            cell.alignment = left_align if j in [14, 15] else center_align
            cell.border = thin_border
            cell.fill = tier_fill if j == 1 else PatternFill()

write_section_header(ws, 4, 'Tier 1 — 必测素材（核心方向 + 数据过硬 + 样本充足）')
write_table_header(ws, 5)

tier1_data = [
    ['1 (最优先)', '捉宠', 'V-抓宠战斗', 'AEO', '4+6月(双月可靠)', 77, 8.42, '35.1%', '18.2%', '13.0%', '$31.51',
     5.45, '38.6% → 35.1%',
     '捉宠方向标杆。双月R1均>=35%达KPI，R3稳定12-13%。6月DNU=77样本充足。CPI从$5.45涨到$8.42但仍在可接受范围。唯一双月留存稳定达标的核心素材。',
     '直接进Purchase Camp，也跑Install保量。单独Campaign，日预算$300+。', '$8,000-10,000'],
    ['2', '融合', 'P-宠物展示-合成 3D', 'Install', '6月', 232, 4.36, '27.6%', '13.8%', '10.3%', '$19.96',
     '—(6月新增)', '—',
     '融合方向标杆。全场最高DNU=232，数据最可靠。CPI=$4.36极低(6月Install平均$7.33)，性价比全场第一。R1=27.6%虽未达35%但次留成本仅$19.96极具竞争力。',
     '必进Purchase Camp(量大便宜)，也保Install。做主力量素材，日预算$500+。', '$10,000-12,000'],
    ['3', '融合', 'V-宠物展示-宠物合成', 'Install', '6月', 96, 6.88, '27.1%', '24.0%', '14.6%', '$37.56',
     '—(6月新增)', '—',
     '融合方向R3冠军。R3=14.6%为6月大样本中最高，R2=24.0%全场最强中段留存。CPI=$6.88低于6月Install均值。R2→R3衰减控制优异(仅-9.4%)。',
     '与合成3D搭配测试，覆盖图片+视频双形式，日预算$200+。', '$5,000-7,000'],
    ['4', '进化', 'P-宠物展示-二阶进化', 'Install', '6月', 27, 3.68, '33.3%', '7.4%', '7.4%', '$11.85',
     '—(6月新增)', '—',
     '进化方向唯一选手。全场最低CPI=$3.68，次留成本仅$11.85也是最低。R1=33.3%接近35%目标线。DNU=27只差3个达标。R3=7.4%是短板但CPI够低可容忍。',
     '加$200预算跑足DNU>=50验证R3。CPI优势太大值得赌。', '$1,500-2,000'],
    ['5', '捉宠', 'V-抓宠战斗', 'Install', '4+6月(双月可靠)', 77, 4.98, '20.8%', '5.2%', '6.5%', '$26.47',
     3.51, '26.7% → 20.8%',
     '捉宠方向Install版。CPI出色($3.51→$4.98)，6月R1降至20.8%但CPI优势足以覆盖。配合AEO版跑双出价。R1下滑趋势需关注是否延续。',
     '与AEO版同Campaign跑，Install出价保量。', '与AEO合并预算'],
]

write_data_rows(ws, 6, tier1_data, tier1_fill)

write_section_header(ws, 13, 'Tier 2 — 强烈推荐（核心方向 + 数据好但样本需补/需重新验证）')
write_table_header(ws, 14)

tier2_data = [
    ['6', '捉宠', 'V-高效抓宠', 'Install', '4月(6月仅DNU=4)', 32, 5.68, '37.5%', '25.0%', '21.9%', '$18.44',
     '—(6月=$8.54/DNU=4)', '37.5% → 25.0%(6月仅4样本,不可靠)',
     '捉宠方向隐藏宝石。4月R1=37.5%+R3=21.9%+CPI=$5.68三项全能。6月仅DNU=4几乎无数据，无法判断真伪。核心方向+4月极佳，必须重新验证。',
     '重新制作素材(避免素材老化)，单独开Campaign跑Install，目标DNU>=50。', '$2,000-3,000'],
    ['7', '捉宠', 'P-抓宠经营-幽飘', 'Install', '6月', 44, 8.60, '31.8%', '20.5%', '18.2%', '$33.79',
     '—(6月新增)', '—',
     '捉宠经营留存王。R3=18.2%为6月所有素材Top 2，R1=31.8%接近目标。CPI=$8.60略高于均值但留存溢价足以补偿。DNU=44样本可信。',
     '作为捉宠方向Variation测，观察R3能否维持18%+。', '$2,500-3,500'],
    ['8', '捉宠', 'P-抓宠经营-超梦', 'Install', '6月', 60, 4.74, '28.3%', '16.7%', '10.0%', '$22.02',
     '—(6月新增)', '—',
     'CPI=$4.74低，R1=28.3%尚可，R3=10%达标。DNU=60样本充足。整体均衡无短板，适合做捉宠方向量大的基座素材。',
     '作为捉宠方向基底素材保量，不主推但跑量。', '$1,500-2,500'],
    ['9', '融合', 'V-宠物展示-二阶合成', 'Install', '6月', 6, 5.49, '16.7%', '0.0%', '33.3%', '$32.92',
     '—(6月新增)', '—',
     'R3=33.3%异常高(可能统计偏差或真有长尾)，R1仅16.7%。CPI=$5.49低。小样本需验证R3是否真实。融合方向值得花小钱赌。',
     '小预算$100-200验证R3异常是否可复现。', '$200-500'],
]

write_data_rows(ws, 15, tier2_data, tier2_fill)

write_section_header(ws, 21, 'Tier 3 — 值得辅测（非核心方向但数据亮眼，作为补充/黑马）')
write_table_header(ws, 22)

tier3_data = [
    ['10', '灾后重建', 'P-灾后重建', 'Install', '4月(仅,6月未测)', 9, 4.13, '44.4%', '33.3%', '33.3%', '$15.48',
     '—', '—(6月未测)',
     'R1=44%+R3=33%+CPI=$4.13三冠王。4月小样本天花板。非核心方向但在Campaign语境下(灾后/建造)可能是潜在爆点。',
     '小预算验证6月环境，目标DNU>=30。', '$500-1,000'],
    ['11', '帕萌战斗', 'V-帕萌战斗-群殴打鸡', 'Install', '6月', 17, 8.21, '52.9%', '29.4%', '5.9%', '$12.77',
     '—(6月新增)', '—',
     'R1=52.9%全场最高！R2=29.4%也很强。但R3暴跌至5.9%(次日R1→R2再→R3断崖)。用户次日很喜欢但第三天大量流失。CPI=$8.21中等。需理解R3暴跌原因(产品问题?素材误导?)。',
     '加预算跑DNU>=50，观察R3是产品问题还是统计偏差。', '$1,000-1,500'],
    ['12', '场景展示', 'V-场景展示', 'AEO', '6月(4月为辅)', 12, 8.89, '50.0%', '25.0%', '16.7%', '$23.71',
     15.44, '100%(DNU=1)→50.0%(DNU=12)',
     'AEO版R1=50%+R3=16.7%留存极佳。Install版4月DNU=76已验证R1=31.6%+R3=13.2%。但6月Install留存下滑。AEO场景展示是潜力股。',
     'AEO+Install双出价测试，参考场景展示4月Install成功经验。', '$1,500-2,000'],
    ['13', '帕萌战斗', 'V-帕萌战斗-雪地竞品', 'Install', '6月', 41, 7.48, '31.7%', '12.2%', '9.8%', '$21.28',
     '—(6月新增)', '—',
     'R1=31.7%不错+CPI=$7.48可接受。DNU=41样本可信。作为竞品对标素材(雪地竞品)，对理解市场定位有价值。',
     '小额测试，观察是否能复现R1>30%。', '$500-1,000'],
]

write_data_rows(ws, 23, tier3_data, tier3_fill)

# Section 4: NOT recommended
write_section_header(ws, 29, 'Tier 4 — 不推荐 / 建议放弃')
row = 30
skip_headers = ['素材名称', '出价方式', '月份', 'CPI', 'R1', 'R3', 'DNU', '花费(USD)', '不推荐原因']
for i, h in enumerate(skip_headers, 1):
    cell = ws.cell(row=row, column=i, value=h)
    cell.font = header_font
    cell.fill = PatternFill('solid', fgColor='808080')
    cell.alignment = center_align
    cell.border = thin_border

skip_data = [
    ['P-海啸(AEO)', 'AEO', '4+6月', '$8.38', '28.1%', '12.5%', 128, '$1,407', '非核心方向。CPI上涨$3.11无优势。R1从38%跌至28%不达标。AEO里CPI最高($8.38)。'],
    ['P-海啸(Install)', 'Install', '4+6月', '$8.70', '26.3%', '10.5%', 95, '$1,271', '非核心方向。CPI=$8.70无优势，次留成本$50.83太高。R1常年不达标(24%→26%)。'],
    ['P-模拟经营-冬日写实', 'Install', '6月', '$10.40', '17.9%', '4.2%', 95, '$1,352', '非核心方向。高花费低留存典型。CPI=$10.40太贵，R1=17.9%全场倒数。模拟经营方向整体翻车。'],
    ['V-帕萌战斗-杀宠复刻', 'Install', '6月', '$11.00', '21.5%', '6.2%', 65, '$913', '非核心方向。CPI=$11.00全场最贵之一，R1=21.5%低。$913高花费验证方向失败。'],
    ['V-模拟经营-帕基玩法寒霜', 'Install', '6月', '$19.27', '36.4%', '18.2%', 11, '$193', 'CPI=$19.27太贵，即使R1/R3不错也无法承受。需素材重做降CPI，9月来不及。'],
    ['V-模拟经营-砍树', 'Install', '6月', '$23.13', '0%', '0%', 3, '$116', 'CPI=$23.13全场最贵，R1=0%。模拟经营方向验证失败。'],
    ['P-模拟经营-建造成长', 'Install', '6月', '$14.24', '33.3%', '0%', 6, '$157', 'CPI高+R3=0%。R1虽33%但仅6个DNU不可靠。'],
    ['V-模拟经营-七日建造', 'Install', '6月', '$14.12', '0%*', '14.3%', 7, '$85', 'D1跟踪异常，CPI=$14.12太贵。数据不可用。'],
    ['V-抓宠经营-拯救可达鸭', 'Install', '6月', '$18.76', '25.0%', '0%', 8, '$169', 'CPI=$18.76太贵。小样本留存差。'],
    ['V-抓宠经营-狐狸', 'Install', '6月', '$15.52', '40.0%', '0%', 5, '$78', 'CPI=$15.52太贵。R1虽40%但仅5个样本不可靠。'],
    ['P-抓宠经营-血腥', 'Install', '6月', '$21.19', '0%', '0%', 1, '$42', 'CPI=$21.19，DNU=1。彻底失败。'],
    ['V-抓宠经营-售卖帕基', 'Install', '6月', '$8.91', '22.6%', '3.2%', 31, '$383', 'R3=3.2%太低。CPI偏高。'],
    ['V-抓宠经营-虐待(Install)', 'Install', '6月', '$3.98', '50.0%', '0%', 4, '$20', '小样本DNU=4。CPI低但R3=0。可随手测但不重点。'],
]

for i, d in enumerate(skip_data):
    row = 31 + i
    ws.row_dimensions[row].height = 36
    for j, val in enumerate(d, 1):
        cell = ws.cell(row=row, column=j, value=val)
        cell.font = normal_font
        cell.alignment = left_align if j == 9 else center_align
        cell.border = thin_border
        cell.fill = tier4_fill

# Budget summary
row = 46
write_section_header(ws, row, '预算分配概览')
row = 47
sum_headers = ['等级', '素材数', '建议预算', '占比', '说明']
for i, h in enumerate(sum_headers, 1):
    cell = ws.cell(row=row, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

sum_data = [
    ['Tier 1 必测', '5个(4素材组)', '$25,000-31,000', '~60%', '核心方向验证付费，主力量'],
    ['Tier 2 强烈推荐', '4个', '$6,200-9,000', '~18%', '补样本+核心方向Variation'],
    ['Tier 3 辅测', '4个', '$3,500-5,500', '~10%', '非核心黑马+场景补充'],
    ['机动/新素材(Task 2)', '待定', '$5,000-8,000', '~12%', '竞品分析后新方向素材脚本'],
    ['合计', '—', '$39,700-53,500', '100%', '预算弹性取决于CPI实际表现和砍量情况'],
]

for i, d in enumerate(sum_data):
    row = 48 + i
    ws.row_dimensions[row].height = 25
    for j, val in enumerate(d, 1):
        cell = ws.cell(row=row, column=j, value=val)
        cell.font = bold_font if i == 4 else normal_font
        cell.alignment = center_align
        cell.border = thin_border

# Strategy notes
row = 55
write_section_header(ws, row, '核心策略说明')

notes = [
    '1. 核心方向（捉宠、融合、进化）优先级最高。V-抓宠战斗AEO是唯一双月R1均>=35%的素材，是9月测试最确定的底牌。',
    '2. 融合方向以P-宠物展示-合成 3D（CPI最低）和V-宠物展示-宠物合成（R3最高）双素材组合，覆盖图片和视频两种形式。',
    '3. 进化方向仅P-宠物展示-二阶进化一个选手，CPI优势巨大但R3是短板，需小预算验证后决定是否加量。',
    '4. 素材筛选逻辑：先跑Install/AEO（3-5天）→ 挑CPI低+R1/R3好的 → 复制到Purchase Camp测3日付费。',
    '5. 4月数据整体偏好看（R1/R3虚高），以6月数据为决策主锚。4月的高效抓宠/灾后重建需6月环境下重新验证。',
    '6. 海啸/模拟经营方向整体放弃。帕萌战斗方向仅保留群殴打鸡（R1最高）和雪地竞品（均衡）做观察。',
    '7. 机动预算$5-8K留给Task 2的新方向素材（等你竞品素材后产出脚本）。',
    '8. V-抓宠经营-虐待(Install) CPI=$3.98/R1=50%但DNU仅4，跑Tier 1素材时可随手搭一个Ad Set测一下，不单独预算。',
]

for i, note in enumerate(notes):
    row = 56 + i
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'] = note
    ws[f'A{row}'].font = Font(name='Arial', size=10, color='333333')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 22

wb.save('creative_final.xlsx')
print('Done! Sheet "9月测试素材推荐" added to creative_final.xlsx')
