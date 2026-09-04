import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from collections import defaultdict

wb = load_workbook('creative_final.xlsx')
ws1 = wb['素材全量数据-按素材']
ws2 = wb['9月推荐-按方向']

# Collect S1 data with merged cell handling
s1 = {}
current_name = None
for row in range(4, 92):
    raw_name = ws1.cell(row=row, column=1).value
    month = ws1.cell(row=row, column=4).value
    bid = ws1.cell(row=row, column=5).value
    tier = ws1.cell(row=row, column=21).value

    if raw_name and str(raw_name).strip():
        current_name = str(raw_name).strip()
    if not month:
        continue
    if tier:
        s1[(current_name, str(month).strip(), str(bid).strip())] = str(tier).strip()

# Collect S2 data
s2 = {}
s2_name_tiers = defaultdict(set)
for row in range(1, ws2.max_row + 1):
    c2 = str(ws2.cell(row=row, column=2).value or '')
    c3 = str(ws2.cell(row=row, column=3).value or '')
    c4 = str(ws2.cell(row=row, column=4).value or '')
    c5 = str(ws2.cell(row=row, column=5).value or '')
    if ('tier' not in c2) and ('🥇' in c2 or '🥈' in c2 or '🥉' in c2 or '❌' in c2):
        if c3.strip() and c3.strip() != 'None':
            key = (c3.strip(), c4.strip(), c5.strip())
            s2[key] = c2.strip()
            s2_name_tiers[c3.strip()].add(c2.strip())

print(f'S1 combos: {len(s1)}')
print(f'S2 combos: {len(s2)}')

# Cross-sheet mismatches
mismatches = []
for k in s1:
    if k in s2 and s1[k] != s2[k]:
        mismatches.append((k, s1[k], s2[k]))
    elif k not in s2:
        mismatches.append((k, s1[k], 'MISSING'))

print(f'\nCross-sheet mismatches: {len(mismatches)}')
if mismatches:
    for m in mismatches[:10]:
        print(f'  {m[0]}: S1={m[1]}, S2={m[2]}')

# Same creative, multiple tiers in S2
inconsistent = {k:v for k,v in s2_name_tiers.items() if len(v) > 1}
print(f'\nSame creative, multiple tiers in S2: {len(inconsistent)}')
if inconsistent:
    print('  (These are OK if different bidding/months have different data)')
    for name, tiers in sorted(inconsistent.items())[:8]:
        print(f'  {name}: {tiers}')

# Key case check
print('\nKey case checks:')
for name, month, bid in [
    ('P-抓宠经营-幽飘', '6月', 'Install'),
    ('V-抓宠战斗', '6月', 'AEO'),
    ('V-抓宠战斗', '6月', 'Install'),
    ('V-高效抓宠', '4月', 'Install'),
    ('V-高效抓宠', '6月', 'Install'),
    ('P-海啸', '6月', 'AEO'),
    ('P-海啸', '6月', 'Install'),
    ('P-海啸', '4月', 'Install'),
    ('P-海啸', '4月', 'AEO'),
]:
    t1 = s1.get((name,month,bid), 'NOT FOUND')
    t2 = s2.get((name,month,bid), 'NOT FOUND')
    ok = 'OK' if t1 == t2 else 'MISMATCH'
    print(f'  {ok}: {name} {month} {bid} => S1={t1} S2={t2}')

# Also check: in S2, does the same (name,month,bid) appear in multiple direction sections with same tier?
s2_full = defaultdict(list)
for row in range(1, ws2.max_row + 1):
    c2 = str(ws2.cell(row=row, column=2).value or '')
    c3 = str(ws2.cell(row=row, column=3).value or '')
    c4 = str(ws2.cell(row=row, column=4).value or '')
    c5 = str(ws2.cell(row=row, column=5).value or '')
    if ('🥇' in c2 or '🥈' in c2 or '🥉' in c2 or '❌' in c2) and c3.strip() and c3.strip() != 'None':
        s2_full[(c3.strip(), c4.strip(), c5.strip())].append(c2.strip())

# Check multi-appearance entries have consistent tiers
multi = {k: set(v) for k, v in s2_full.items() if len(v) > 1}
print(f'\nEntries appearing in multiple S2 sections: {len(multi)}')
bad_multi = {k: v for k, v in multi.items() if len(v) > 1}
if bad_multi:
    print('  ISSUE - same combo, different tiers in different sections:')
    for k, v in bad_multi.items():
        print(f'    {k}: {v}')
else:
    print('  All consistent (same tier across all sections)')
