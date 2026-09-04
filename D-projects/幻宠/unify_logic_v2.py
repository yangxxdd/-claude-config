import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

wb = load_workbook('creative_final.xlsx')
ws1 = wb['素材全量数据-按素材']

BLUE_FONT = Font(name='Arial', size=10, color='0066CC')
BLUE_BOLD = Font(name='Arial', size=10, color='0066CC', bold=True)
BLACK_FONT = Font(name='Arial', size=10, color='000000')
NF = Font(name='Arial', size=10)

# =====================================================
# PHASE 1: Read all data from Sheet 1
# =====================================================
rows_data = []
current_name = None
for row in range(4, 92):
    raw_name = ws1.cell(row=row, column=1).value
    month = ws1.cell(row=row, column=4).value
    bid = ws1.cell(row=row, column=5).value

    if raw_name and str(raw_name).strip():
        current_name = str(raw_name).strip()
    if not month:
        continue

    try:
        dnu = int(float(ws1.cell(row=row, column=6).value or 0))
        cpi = float(ws1.cell(row=row, column=7).value or 0)
        r1_str = str(ws1.cell(row=row, column=8).value or '0%').replace('%','')
        r1 = float(r1_str) / 100
        r2_str = str(ws1.cell(row=row, column=9).value or '0%').replace('%','').replace('—','0')
        r2 = float(r2_str) / 100 if r2_str and r2_str != '0' else 0  # treat — as 0
        r3_str = str(ws1.cell(row=row, column=10).value or '0%').replace('%','')
        r3 = float(r3_str) / 100
        # 次留成本 from Col11 (formula: =L/(F*H))
        cost22_val = ws1.cell(row=row, column=11).value
        cost22 = None
        if cost22_val and str(cost22_val) != '—':
            try:
                cost22 = float(cost22_val)
            except:
                pass
        # CTR*CVR from Col18 (formula: =Q*P)
        ctr_cvr_val = ws1.cell(row=row, column=18).value
        ctr_cvr = None
        if ctr_cvr_val and str(ctr_cvr_val) != '—':
            try:
                ctr_cvr = float(ctr_cvr_val)
            except:
                pass
    except Exception as e:
        continue

    rows_data.append({
        'row': row, 'name': current_name, 'month': str(month).strip(),
        'bid': str(bid).strip(), 'dnu': dnu, 'cpi': cpi,
        'r1': r1, 'r2': r2, 'r3': r3,
        'cost22': cost22, 'ctr_cvr': ctr_cvr
    })

# =====================================================
# PHASE 2: Tier logic — CPI primary, retention secondary, CTR*CVR tiebreaker
# =====================================================
# CPI score (0-5): highest weight — CPI is the #1 concern
#   <$5: 5pt | $5-6.99: 4pt | $7-8.99: 3pt | $9-10.99: 1pt | $11-14.99: 0pt | $15+: -1pt
# Retention score (0-4): R1+R3 combined
#   R1: >=35% 2pt | >=28% 1.5pt | >=20% 1pt | <20% 0pt
#   R3: >=15% 2pt | >=10% 1.5pt | >=7% 1pt | <7% 0pt
# 次留成本 bonus (0-1):
#   <$20: 1pt | <$30: 0.5pt
# CTR*CVR bonus (0-1): competitiveness tiebreaker
#   >=1.0%: 1pt | >=0.5%: 0.5pt
# 6月 bonus (0-1):
#   6月: 1pt | 4月: 0pt
# DNU penalty:
#   DNU<10: score capped at 待观察 max
#   DNU<30: -0.5pt

def get_tier_and_reason(d):
    dnu, cpi, r1, r2, r3 = d['dnu'], d['cpi'], d['r1'], d['r2'], d['r3']
    cost22 = d.get('cost22')
    ctr_cvr = d.get('ctr_cvr')
    month = d['month']
    is_june = month == '6月'
    noise = dnu < 10
    small = dnu < 30

    # ---- CPI score (0-5) ----
    if cpi < 5:
        cpi_score = 5
        cpi_label = f'CPI=${cpi:.2f}优秀'
    elif cpi < 7:
        cpi_score = 4
        cpi_label = f'CPI=${cpi:.2f}良好'
    elif cpi < 9:
        cpi_score = 3
        cpi_label = f'CPI=${cpi:.2f}适中'
    elif cpi < 11:
        cpi_score = 1
        cpi_label = f'CPI=${cpi:.2f}偏高'
    elif cpi < 15:
        cpi_score = 0
        cpi_label = f'CPI=${cpi:.2f}高'
    else:
        cpi_score = -1
        cpi_label = f'CPI=${cpi:.2f}过高'

    # ---- Retention score (0-4) ----
    if r1 >= 0.35:
        r1_score = 2.0
        r1_label = f'R1={r1:.0%}达标'
    elif r1 >= 0.28:
        r1_score = 1.5
        r1_label = f'R1={r1:.0%}尚可'
    elif r1 >= 0.20:
        r1_score = 1.0
        r1_label = f'R1={r1:.0%}偏低'
    else:
        r1_score = 0
        r1_label = f'R1={r1:.0%}差'

    if r3 >= 0.15:
        r3_score = 2.0
        r3_label = f'R3={r3:.0%}优秀'
    elif r3 >= 0.10:
        r3_score = 1.5
        r3_label = f'R3={r3:.0%}达标'
    elif r3 >= 0.07:
        r3_score = 1.0
        r3_label = f'R3={r3:.0%}一般'
    elif r3 > 0:
        r3_score = 0
        r3_label = f'R3={r3:.0%}偏低'
    else:
        r3_score = 0
        r3_label = 'R3=0%'

    ret_score = r1_score + r3_score

    # ---- 次留成本 bonus (0-1) ----
    cost22_score = 0
    cost22_label = ''
    if cost22 and cost22 > 0:
        if cost22 < 18:
            cost22_score = 1.0
            cost22_label = f'次留成本${cost22:.0f}很优秀'
        elif cost22 < 25:
            cost22_score = 0.5
            cost22_label = f'次留成本${cost22:.0f}良好'
        elif cost22 < 35:
            cost22_label = f'次留成本${cost22:.0f}尚可'
        else:
            cost22_label = f'次留成本${cost22:.0f}偏高'

    # ---- CTR*CVR competitiveness (0-1) ----
    ctrcvr_score = 0
    ctrcvr_label = ''
    if ctr_cvr and ctr_cvr > 0:
        ctr_cvr_pct = ctr_cvr * 100  # CTR*CVR as percentage
        if ctr_cvr >= 0.008:  # >=0.8%, top quartile
            ctrcvr_score = 1.0
            ctrcvr_label = f'CTR*CVR={ctr_cvr_pct:.1f}%竞争力强'
        elif ctr_cvr >= 0.003:  # >=0.3%, median level
            ctrcvr_score = 0.5
            ctrcvr_label = f'CTR*CVR={ctr_cvr_pct:.1f}%竞争力尚可'

    # ---- Month bonus ----
    month_score = 1.0 if is_june else 0
    month_label = '6月数据' if is_june else '4月数据(需6月重验)'

    # ---- DNU adjustment ----
    dnu_penalty = 0.5 if small else 0
    dnu_label = f'DNU={dnu}'
    if noise:
        dnu_label += '样本不足'
    elif small:
        dnu_label += '偏小'
    else:
        dnu_label += '可靠'

    # ---- Total score ----
    total = cpi_score + ret_score + cost22_score + ctrcvr_score + month_score - dnu_penalty

    # ---- Determine tier ----
    if noise:
        # DNU<10: limited to 待观察 at best
        if total >= 7:
            tier = '🥉 待观察'
        elif total >= 5:
            tier = '🥉 待观察'
        else:
            tier = '❌ 不推荐'
    elif not is_june:
        # 4月 only: cap at 可选
        if total >= 8:
            tier = '🥈 可选'
        elif total >= 6:
            tier = '🥉 待观察'
        else:
            tier = '❌ 不推荐'
    else:
        # 6月, DNU>=10: full range
        # V-抓宠战斗 AEO (底牌) = 3+2+1.5+0+0+1 = 7.5 → 🥇
        if total >= 7.5:
            tier = '🥇 强烈推荐'
        elif total >= 6:
            tier = '🥈 可选'
        elif total >= 4.5:
            tier = '🥉 待观察'
        else:
            tier = '❌ 不推荐'

    # ---- Build reason ----
    parts = [cpi_label, r1_label, r3_label]
    if cost22_label:
        parts.append(cost22_label)
    if ctrcvr_label:
        parts.append(ctrcvr_label)
    parts.append(dnu_label)
    parts.append(month_label)

    # Add directional assessment
    if tier == '🥇 强烈推荐':
        parts.append('→ 主推,可进Purchase Camp')
    elif tier == '🥈 可选':
        parts.append('→ 可测,需观察后再进Purchase')
    elif tier == '🥉 待观察':
        parts.append('→ 小预算验证')
    else:
        parts.append('→ 不推荐投放')

    reason = '; '.join(parts)
    return tier, reason, total

# Apply
for d in rows_data:
    tier, reason, score = get_tier_and_reason(d)
    d['tier'] = tier
    d['reason'] = reason
    d['score'] = score

# Print distribution
from collections import Counter
tc = Counter(d['tier'] for d in rows_data)
print('=== Tier Distribution ===')
for t in ['🥇 强烈推荐', '🥈 可选', '🥉 待观察', '❌ 不推荐']:
    print(f'  {t}: {tc.get(t, 0)}')

# =====================================================
# PHASE 3: Update Sheet 1
# =====================================================
tier_col = 21

# Unmerge tier column area
for mc in list(ws1.merged_cells.ranges):
    if mc.min_col <= tier_col <= mc.max_col and mc.min_row >= 4:
        ws1.unmerge_cells(str(mc))

# Write tier header
ws1.cell(row=3, column=tier_col, value='推荐等级').font = Font(name='Arial', bold=True, size=10, color='0066CC')
ws1.cell(row=3, column=tier_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws1.cell(row=3, column=tier_col).border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

for d in rows_data:
    row = d['row']

    # Check if note cell is in a merged range (separator row)
    note_cell = ws1.cell(row=row, column=19)
    is_merged_note = False
    for mc in ws1.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= 19 <= mc.max_col:
            is_merged_note = True
            break
    if is_merged_note:
        continue

    # Update 备注
    note_cell.value = d['reason']
    note_cell.font = BLUE_FONT

    # Write tier
    try:
        tier_cell = ws1.cell(row=row, column=tier_col)
        tier_cell.value = d['tier']
        tier_cell.font = BLUE_BOLD
        tier_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        tier_cell.border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
        if '🥇' in d['tier']:
            tier_cell.fill = PatternFill('solid', fgColor='C6EFCE')
        elif '🥈' in d['tier']:
            tier_cell.fill = PatternFill('solid', fgColor='BDD7EE')
        elif '🥉' in d['tier']:
            tier_cell.fill = PatternFill('solid', fgColor='FCE4D6')
        else:
            tier_cell.fill = PatternFill('solid', fgColor='F4B4C2')
    except AttributeError:
        continue

# =====================================================
# PHASE 4: Rebuild Sheet 2 — derived from Sheet 1
# =====================================================
for s in ['9月推荐-按方向', '9月推荐-方向汇总']:
    if s in wb.sheetnames:
        del wb[s]

ws2 = wb.create_sheet('9月推荐-按方向')

# Group by direction (from Col2, handle merged)
direction_groups = defaultdict(list)
current_dir = None
for d in rows_data:
    row = d['row']
    raw_dir = ws1.cell(row=row, column=2).value
    if raw_dir and str(raw_dir).strip():
        current_dir = str(raw_dir).strip()
    if current_dir:
        direction_groups[current_dir].append(d)

dir_priority = {'捉宠':1,'捉宠/战斗':2,'捉宠/模拟经营/建造':3,'融合':4,'融合/进化':5,'进化':6,
                '战斗':7,'模拟经营/建造':8,'模拟经营/建造/战斗':9,'宠物展示':10,
                '天灾/生存':11,'经营/其他':12,'其他':13}
sorted_dirs = sorted(direction_groups.keys(), key=lambda x: dir_priority.get(x.split('/')[0], 99))

hdr_f = Font(name='Arial', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='2F5496')
sec_fill = PatternFill('solid', fgColor='D6E4F0')
thin_b = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
la = Alignment(horizontal='left', vertical='center', wrap_text=True)

row = 1
ws2.merge_cells('A1:J1')
ws2['A1'] = '幻宠 9月测试素材推荐 — 按方向归类（统一推荐等级 v2，CPI权重最高+留存辅+CTR*CVR参考）'
ws2['A1'].font = Font(name='Arial', bold=True, size=13, color='0066CC')
ws2.row_dimensions[1].height = 30

for i, w in enumerate([5, 11, 24, 8, 8, 8, 10, 10, 10, 55], 1):
    ws2.column_dimensions[chr(64+i)].width = w

row = 3
for dir_name in sorted_dirs:
    dir_data = direction_groups[dir_name]
    tier_order = {'🥇 强烈推荐':0, '🥈 可选':1, '🥉 待观察':2, '❌ 不推荐':3}
    dir_data.sort(key=lambda x: (tier_order.get(x['tier'],99), x['cpi']))

    ws2.merge_cells(f'A{row}:J{row}')
    ws2[f'A{row}'] = f'{dir_name}（{len(dir_data)}个素材-出价组合）'
    ws2[f'A{row}'].font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    for c in range(1, 11):
        ws2.cell(row=row, column=c).fill = sec_fill
    row += 1

    for i, h in enumerate(['序号','推荐等级','素材名称','月份','出价方式','DNU','CPI','次留成本','CTR*CVR','推荐理由'], 1):
        c = ws2.cell(row=row, column=i, value=h)
        c.font = hdr_f; c.fill = hdr_fill; c.alignment = ca; c.border = thin_b
    row += 1

    for idx, d in enumerate(dir_data, 1):
        tier_fill = PatternFill()
        if '🥇' in d['tier']: tier_fill = PatternFill('solid', fgColor='C6EFCE')
        elif '🥈' in d['tier']: tier_fill = PatternFill('solid', fgColor='BDD7EE')
        elif '🥉' in d['tier']: tier_fill = PatternFill('solid', fgColor='FCE4D6')
        else: tier_fill = PatternFill('solid', fgColor='F4B4C2')

        cost_str = f'${d["cost22"]:.0f}' if d.get('cost22') and d['cost22'] > 0 else '—'
        ctrcvr_str = f'{d["ctr_cvr"]*100:.1f}%' if d.get('ctr_cvr') and d['ctr_cvr'] > 0 else '—'

        vals = [idx, d['tier'], d['name'], d['month'], d['bid'], d['dnu'],
                f'${d["cpi"]:.2f}', cost_str, ctrcvr_str, d['reason']]
        for i, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=i, value=v)
            c.font = BLUE_FONT if i in [2, 7, 8, 9, 10] else NF
            c.alignment = la if i == 10 else ca
            c.border = thin_b
            c.fill = tier_fill
        ws2.row_dimensions[row].height = 24
        row += 1
    row += 1

# =====================================================
# PHASE 5: Rebuild Sheet 3
# =====================================================
ws3 = wb.create_sheet('9月推荐-方向汇总')

row = 1
ws3.merge_cells('A1:G1')
ws3['A1'] = '9月测试素材推荐 — 方向级汇总（CPI权重最高+留存辅+CTR*CVR参考）'
ws3['A1'].font = Font(name='Arial', bold=True, size=13, color='0066CC')

for i, w in enumerate([28, 10, 10, 10, 10, 55, 55], 1):
    ws3.column_dimensions[chr(64+i)].width = w

row = 3
for i, h in enumerate(['方向','强烈推荐','可选','待观察','不推荐','核心结论','推荐素材(等级+CPI+R1+R3)'], 1):
    c = ws3.cell(row=row, column=i, value=h)
    c.font = hdr_f; c.fill = hdr_fill; c.alignment = ca; c.border = thin_b
ws3.row_dimensions[row].height = 30
row += 1

for dir_name in sorted_dirs:
    dir_data = direction_groups[dir_name]
    c1 = sum(1 for d in dir_data if '🥇' in d['tier'])
    c2 = sum(1 for d in dir_data if '🥈' in d['tier'])
    c3 = sum(1 for d in dir_data if '🥉' in d['tier'])
    c4 = sum(1 for d in dir_data if '❌' in d['tier'])

    top = [d for d in dir_data if '🥇' in d['tier'] or '🥈' in d['tier']][:4]
    pick_list = '\n'.join([
        f'{d["tier"][:2]} {d["name"]}({d["bid"]},CPI${d["cpi"]:.2f},R1={d["r1"]:.0%},R3={d["r3"]:.0%})'
        for d in top
    ])

    if c1 >= 2:
        conclusion = f'方向储备充足。{c1}个强烈推荐素材做主力量。'
    elif c1 == 1:
        conclusion = f'有1个强烈推荐素材为方向核心。'
    elif c2 >= 1:
        conclusion = f'{c2}个可选素材需进一步验证。'
    else:
        conclusion = '无推荐素材,方向储备不足。'
    if c4 > len(dir_data)*0.5:
        conclusion += f' {c4}/{len(dir_data)}不推荐,方向整体偏弱。'

    vals = [dir_name, c1, c2, c3, c4, conclusion, pick_list]
    for i, v in enumerate(vals, 1):
        c = ws3.cell(row=row, column=i, value=v)
        c.font = BLUE_FONT
        c.alignment = la if i in [6,7] else ca
        c.border = thin_b
    ws3.row_dimensions[row].height = 80
    row += 1

wb.save('creative_final.xlsx')
print('\nDone v2! CPI-weighted tier logic applied.')
print('All changes in BLUE font.')
