import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

wb = load_workbook('creative_final.xlsx')

# Remove old sheets if exist
for s in ['9月测试素材推荐', '9月测试素材推荐v2']:
    if s in wb.sheetnames:
        del wb[s]

ws = wb.create_sheet('9月测试素材推荐v2', 0)

# ===== Styles =====
hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='2F5496')
tier1_fill = PatternFill('solid', fgColor='C6EFCE')
tier2_fill = PatternFill('solid', fgColor='BDD7EE')
tier3_fill = PatternFill('solid', fgColor='FCE4D6')
tier4_fill = PatternFill('solid', fgColor='D9D9D9')
sec_fill = PatternFill('solid', fgColor='4472C4')
sec_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
bold_f = Font(name='Arial', bold=True, size=11)
norm_f = Font(name='Arial', size=11)
red_f = Font(name='Arial', size=10, color='CC0000')
green_f = Font(name='Arial', size=10, color='006100')
thin_b = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
c_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
l_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Column widths
col_w = [6, 8, 22, 10, 14, 10, 14, 14, 14, 14, 14, 14, 20, 55, 50, 20]
for i, w in enumerate(col_w, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ===== Row 1-2: Title & Context =====
ws.merge_cells('A1:P1')
ws['A1'] = '幻宠 9月测试 — 素材推荐 v2（扩增版：更多核心方向素材 + 更宽松的准入）'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 32

ws.merge_cells('A2:P2')
ws['A2'] = '测试：9月初 | 预算：$40-50K | 量级：~5K | 优化目标：Purchase(付费) | 策略：Install/AEO先跑3-5天筛选(看CPI+R1+R3) → 优胜者进Purchase Camp测3日付费 | 核心方向：捉宠/融合/进化'
ws['A2'].font = Font(name='Arial', size=10, color='333333')
ws['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[2].height = 28

# ===== Helpers =====
headers = ['优先级', '方向', '素材名称', '出价方式', '验证月份', 'DNU', 'CPI(USD)', 'R1', 'R2', 'R3',
           '次留成本', '4月CPI', '4月R1→6月R1', '推荐理由', '建议动作', '建议预算(USD)']

def sec_hdr(ws, row, text):
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'] = text
    ws[f'A{row}'].font = sec_font
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    for c in range(1, 17):
        ws.cell(row=row, column=c).fill = sec_fill

def tbl_hdr(ws, row):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = c_align
        c.border = thin_b
    ws.row_dimensions[row].height = 35

def write_rows(ws, start_row, data, fill):
    for i, d in enumerate(data):
        row = start_row + i
        ws.row_dimensions[row].height = max(72, len(d[13]) * 0.6 + 30)
        for j, val in enumerate(d, 1):
            c = ws.cell(row=row, column=j, value=val)
            c.font = norm_f
            c.alignment = l_align if j in [14, 15] else c_align
            c.border = thin_b
            c.fill = fill if j == 1 else PatternFill()

# ======================================================
# Tier 1: 必测 (expanded from 5 to 9 entries)
# ======================================================
sec_hdr(ws, 4, 'Tier 1 — 必测素材（核心方向 + 数据有说服力 + 样本充足或接近充足）')
tbl_hdr(ws, 5)

t1 = [
    ['1 🥇', '捉宠', 'V-抓宠战斗', 'AEO', '4+6月双月可靠', 77, '$8.42', '35.1%', '18.2%', '13.0%', '$31.51',
     '$5.45', '38.6% → 35.1%(持平)',
     '[捉宠方向标杆] 双月R1均≥35%唯一素材，R3稳定12-13%。6月DNU=77样本充足。CPI从$5.45涨至$8.42但仍低于6月AEO均值。唯一双月留存都达KPI的核心素材。9月最确定的底牌。',
     '直接进Purchase Camp(主力)，同时跑Install保量。单独Campaign日预算$400+。', '$10,000-12,000'],
    ['2 🥈', '融合', 'P-宠物展示-合成 3D', 'Install', '6月(大样本)', 232, '$4.36', '27.6%', '13.8%', '10.3%', '$19.96',
     '— (6月新增)', '—',
     '[融合方向标杆+量王] 全场最高DNU=232数据最可靠。CPI=$4.36全场最低(6月Install均值$7.33)，性价比第一。R1=27.6%虽未达35%但次留成本仅$19.96性价比极高。跑量首选。',
     '必进Purchase Camp(量大便宜)，做主力量素材日预算$500+。', '$10,000-12,000'],
    ['3 🥉', '捉宠', 'P-抓宠经营-幽飘', 'Install', '6月(DNU=44)', 44, '$8.60', '31.8%', '20.5%', '18.2%', '$33.79',
     '— (6月新增)', '—',
     '[捉宠方向R3冠军] R3=18.2%为6月大样本中最高，R2=20.5%→R3=18.2%衰减仅2.3%全场最佳。R1=31.8%接近目标线。CPI=$8.60略高于均值但留存溢价完全覆盖。',
     '单独Ad Set跑Install观察R3能否维持，达标后进Purchase。', '$3,000-4,000'],
    ['4', '进化', 'P-宠物展示-二阶进化', 'Install', '6月(DNU=27)', 27, '$3.68', '33.3%', '7.4%', '7.4%', '$11.85',
     '— (6月新增)', '—',
     '[进化方向唯一选手+CPI冠军] 全场最低CPI=$3.68+次留成本$11.85双料最低。R1=33.3%仅差1.7%达35%。DNU=27差3个达标。R3=7.4%是短板但CPI够低可容忍。进化方向必须保。',
     '加$300跑足DNU≥50验证R3稳定性。CPU低值得赌。', '$1,500-2,500'],
    ['5', '捉宠', 'V-抓宠战斗', 'Install', '4+6月双月可靠', 77, '$4.98', '20.8%', '5.2%', '6.5%', '$26.47',
     '$3.51', '26.7% → 20.8%(下滑)',
     '[捉宠Install版] CPI出色($3.51→$4.98)，远低于6月Install均值$7.33。R1虽下滑至20.8%但CPI优势显著。配合AEO版跑双出价互补。R1下滑可能是素材老化或竞争加剧，需关注。',
     '与AEO版同Campaign跑Install出价保量。', '与#1合并预算'],
    ['6', '捉宠', 'V-核心玩法', 'Install', '4月(可靠),6月缺', 54, '$3.52', '40.7%', '18.5%', '7.4%', '$12.65',
     '— (6月DNU=3)', '40.7% → 66.7%(6月DNU=3噪音)',
     '[捉宠相关,4月黑马] 4月R1=40.7%+CPI=$3.52+DNU=54非常扎实。R3=7.4%偏低但次留成本仅$12.65极低。6月仅DNU=3无数据。核心捉宠方向相关(展示核心玩法)，4月表现强必须重验。',
     '重新制作素材(防老化)，单独Campaign跑Install目标DNU≥50。', '$2,000-3,000'],
    ['7', '捉宠', 'V-高效抓宠', 'Install', '4月(可靠),6月缺', 32, '$5.68', '37.5%', '25.0%', '21.9%', '$18.44',
     '— (6月DNU=4)', '37.5% → 25.0%(6月仅4样本)',
     '[捉宠方向隐藏宝石] 4月R1=37.5%+R3=21.9%+CPI=$5.68+均优秀。R3=21.9%为全场第二高(仅次于灾后重建)。6月仅DNU=4几乎无数据。核心方向+4月三好生必须重验。',
     '重新制作素材(避免老化)，单独Campaign跑Install目标DNU≥50。', '$2,000-3,000'],
    ['8', '捉宠', 'P-抓宠经营-超梦', 'Install', '6月(DNU=60)', 60, '$4.74', '28.3%', '16.7%', '10.0%', '$22.02',
     '— (6月新增)', '—',
     '[捉宠方向稳健基底] CPI=$4.74+DNU=60+无短板。R1+R3都达标。适合做捉宠方向跑量的基座素材。不惊艳但稳定可预测。',
     '捉宠方向基底素材保量，不主推但跑量。', '$2,000-3,000'],
    ['9', '融合', 'V-宠物展示-宠物合成', 'Install', '6月(DNU=96)', 96, '$6.88', '27.1%', '24.0%', '14.6%', '$37.56',
     '— (6月新增)', '—',
     '[融合方向R3+中段留存双冠] R3=14.6%大样本最高,R2=24.0%全场最强。CPI=$6.88低于6月Install均值。DNU=96充分可信。与合成3D搭配覆盖图片+视频双形式。',
     '与合成3D搭配测试覆盖不同素材形式，日预算$200+。', '$4,000-5,000'],
]

write_rows(ws, 6, t1, tier1_fill)

# ======================================================
# Tier 2: 强烈推荐
# ======================================================
sec_hdr(ws, 17, 'Tier 2 — 强烈推荐（核心方向 + 数据好但样本需补 / 非核心但数据突出）')
tbl_hdr(ws, 18)

t2 = [
    ['10', '捉宠', 'V-抓宠经营-捕捞竞品', 'Install', '6月(DNU=54)', 54, '$3.89', '18.5%', '9.3%', '7.4%', '$27.24',
     '— (6月新增)', '—',
     'CPI=$3.89极低(6月Install第二低)，跑量能力突出。DNU=54样本充足。R1=18.5%偏低但CPI优势足够大，适合做低CPI跑量补充。竞品对标素材也有市场参考价值。',
     '捉宠方向低CPI跑量补充，作为基底素材之一。', '$1,500-2,000'],
    ['11', '灾后重建', 'P-灾后重建', 'Install', '4月(仅,DNU=9)', 9, '$4.13', '44.4%', '33.3%', '33.3%', '$15.48',
     '— (6月未测)', '— (6月未测)',
     'R1=44%+R3=33%+CPI=$4.13三冠王。4月综合评价全场第一。但仅9个DNU+6月未测。非核心方向(灾后/建造)但数据太强不能忽略。',
     '小预算验证6月环境能否复现，目标DNU≥30。', '$800-1,500'],
    ['12', '场景展示', 'V-场景展示', 'AEO', '6月(4月为辅)', 12, '$8.89', '50.0%', '25.0%', '16.7%', '$23.71',
     '$15.44(DNU=1)', '100%(噪音) → 50.0%',
     'AEO版R1=50%+R3=16.7%留存极佳，但DNU仅12需大样本验证。Install版4月DNU=76验证过R1=31.6%+R3=13.2%。AEO版可能是潜力股。非核心方向但数据质量高。',
     'AEO+Install双出价测试。参考4月Install成功经验。', '$1,500-2,500'],
    ['13', '融合', 'V-宠物展示-二阶合成', 'Install', '6月(DNU=6)', 6, '$5.49', '16.7%', '0.0%', '33.3%', '$32.92',
     '— (6月新增)', '—',
     'R3=33.3%异常高(R1→R2→R3 = 16.7%→0%→33.3%，可能是统计波动或真有长尾价值)。CPI=$5.49低。核心融合方向+低CPI值得花小钱赌R3能否复现。',
     '极小预算$100-200验证R3异常是否复现。不报太大期望。', '$200-500'],
]

write_rows(ws, 19, t2, tier2_fill)

# ======================================================
# Tier 3: 辅测
# ======================================================
sec_hdr(ws, 25, 'Tier 3 — 值得辅测（数据有趣但不确定性大 / 竞品对标 / 黑马）')
tbl_hdr(ws, 26)

t3 = [
    ['14', '帕萌战斗', 'V-帕萌战斗-群殴打鸡', 'Install', '6月(DNU=17)', 17, '$8.21', '52.9%', '29.4%', '5.9%', '$12.77',
     '— (6月新增)', '—',
     'R1=52.9%全场最高！R2=29.4%也很强。但R3断崖式下跌至5.9%(-23.5%衰减)说明用户次日很喜欢但第三天大量流失。疑似素材误导或产品问题。需观察R3是否可修复。非核心方向。',
     '加预算跑DNU≥50观察R3暴跌是否持续。', '$1,000-1,500'],
    ['15', '帕萌战斗', 'V-帕萌战斗-雪地竞品', 'Install', '6月(DNU=41)', 41, '$7.48', '31.7%', '12.2%', '9.8%', '$21.28',
     '— (6月新增)', '—',
     'R1=31.7%不错+CPI=$7.48可接受+DNU=41可信。竞品对标素材(雪地竞品)，对市场定位有参考价值。非核心方向整体均衡。',
     '小额测试观察能否复现R1>30%。', '$500-1,000'],
    ['16', '帕萌战斗', 'V-帕萌战斗-出狱打鸡', 'Install', '6月(DNU=11)', 11, '$9.01', '36.4%', '9.1%', '9.1%', '$38.31',
     '— (6月新增)', '—',
     'R1=36.4%是6月为数不多R1>35%的素材。R2→R3平稳(9.1%→9.1%)无衰减问题。CPI=$9.01偏高+仅11个DNU需扩大验证。非核心方向。',
     '小预算跑DNU≥30验证R1能否维持。', '$500-800'],
    ['17', '抓宠经营', 'V-抓宠经营-虐待', 'Install', '6月(DNU=4)', 4, '$3.98', '50.0%', '25.0%', '0.0%', '$9.95',
     '— (6月新增)', '—',
     'CPI=$3.98极低+R1=50%看起来很香但DNU仅4。R3=0%可能因样本太小。捉宠经营方向但"虐待"主题可能偏离品牌调性。不作为正式推荐但跑Tier 1时可随手搭一个Ad Set。',
     '搭便车测(搭在Tier 1 Campaign里)，不单独预算。', '$100-200'],
    ['18', '捉宠', 'V-帕基世界冒险', 'Install', '4月(DNU=10)', 10, '$5.41', '10.0%', '20.0%', '0.0%', '$97.41',
     '— (6月未测)', '—',
     'CPI=$5.41在4月属中上水平。R1仅10%很低但R2=20%(R1→R2不降反升说明R1跟踪可能有问题)。6月未测。捉宠方向+CPI还行+数据诡异值得花小钱澄清。',
     '极小预算验证数据异常是否为跟踪问题。', '$200-300'],
]

write_rows(ws, 27, t3, tier3_fill)

# ======================================================
# Tier 4: 不推荐
# ======================================================
sec_hdr(ws, 35, 'Tier 4 — 不推荐 / 明确放弃')

row = 36
sk_h = ['素材名称', '出价方式', '月份', 'CPI', 'R1', 'R3', 'DNU', '花费', '核心问题']
for i, h in enumerate(sk_h, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = hdr_font
    c.fill = PatternFill('solid', fgColor='808080')
    c.alignment = c_align
    c.border = thin_b

sk = [
    ['P-海啸(AEO)', 'AEO', '4+6月', '$8.38', '28.1%', '12.5%', 128, '$1,407', '非核心。CPI无优势。R1从38%跌至28%不达标。AEO里CPI最高。'],
    ['P-海啸(Install)', 'Install', '4+6月', '$8.70', '26.3%', '10.5%', 95, '$1,271', '非核心。CPI无优势。次留成本$50.83太高。R1常年不达标。'],
    ['P-模拟经营-冬日写实', 'Install', '6月', '$10.40', '17.9%', '4.2%', 95, '$1,352', '非核心。高花费+低留存典型。CPI=$10.40太贵。模拟经营方向整体翻车。'],
    ['V-模拟经营-帕基玩法寒霜', 'Install', '6月', '$19.27', '36.4%', '18.2%', 11, '$193', 'CPI=$19.27太贵。即使R1/R3好也无法承受。需素材重做降CPI。'],
    ['V-帕萌战斗-杀宠复刻', 'Install', '6月', '$11.00', '21.5%', '6.2%', 65, '$913', '非核心。CPI=$11.00太贵。高花费$913验证方向失败。'],
    ['其余模拟经营系列', 'Install', '6月', '>$14', '<33%', '~0%', '<10', '<$160', 'CPI全部>$14，R3全部≈0%。方向性失败。'],
    ['其余低DNU抓宠变体', 'Install', '6月', '>$15', '<25%', '0%', '<10', '<$170', '狐狸/可达鸭/血腥等CPI>$15太贵+DNU<10无统计意义。'],
]

for i, d in enumerate(sk):
    row = 37 + i
    ws.row_dimensions[row].height = 30
    for j, val in enumerate(d, 1):
        c = ws.cell(row=row, column=j, value=val)
        c.font = norm_f
        c.alignment = l_align if j == 9 else c_align
        c.border = thin_b
        c.fill = tier4_fill

# ======================================================
# Budget Summary
# ======================================================
row = 46
sec_hdr(ws, row, '▌预算分配概览')
row = 47
for i, h in enumerate(['等级', '素材数', '建议预算', '占比', '说明'], 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = c_align; c.border = thin_b

budget = [
    ['Tier 1 必测', '9个(7组独立素材)', '$35,000-43,500', '~72%', '核心方向验证付费，主力量。6月可靠素材优先给预算。'],
    ['Tier 2 强烈推荐', '4个', '$4,000-6,500', '~10%', '补样本+非核心黑马验证。'],
    ['Tier 3 辅测', '5个', '$2,300-3,800', '~5%', '黑马+竞品对标+异常澄清。'],
    ['机动/新素材(Task 2)', '待定', '$5,000-8,000', '~13%', '竞品分析后新方向素材脚本。'],
    ['合计', '18个(不含Task2)', '$46,300-61,800', '100%', '预算上限可能触发砍量，优先保Tier 1。'],
]

for i, d in enumerate(budget):
    row = 48 + i
    ws.row_dimensions[row].height = 25
    for j, val in enumerate(d, 1):
        c = ws.cell(row=row, column=j, value=val)
        c.font = bold_f if i == 4 else norm_f
        c.alignment = c_align; c.border = thin_b

# ======================================================
# Strategy Notes
# ======================================================
row = 55
sec_hdr(ws, row, '▌核心策略说明')

notes = [
    '1. [方向结构] 捉宠方向7个素材(主力)+融合方向3个+进化方向1个+非核心辅测5个。捉宠方向最大，覆盖不同子方向(战斗/经营/核心玩法)。',
    '2. [底牌] V-抓宠战斗AEO是唯一双月R1均≥35%的素材，9月最确定的底牌。P-宠物展示-合成3D是跑量主力(CPI最低+DNU最大)。',
    '3. [缺位的方向] 进化方向仅P-宠物展示-二阶进化一个选手，素材储备不足。Task 2新素材脚本需重点补进化方向。',
    '4. [筛选漏斗] Install/AEO先跑3-5天 → 按CPI<$8 + R1>25% + R3>8%筛选 → 优胜者进Purchase Camp测3日付费。',
    '5. [4月vs6月] 6月数据为决策主锚。4月表现好但6月未测的素材(高效抓宠/核心玩法/灾后重建)需在6月环境下重新验证，不盲目加量。',
    '6. [放弃] 海啸(CPI无优势+非核心)；模拟经营全系列(CPI高+留存差)；帕萌战斗其他变体(杀宠复刻CPI=$11太高)；抓宠经营狐狸/可达鸭/血腥(CPI>$15)。',
    '7. [素材老化] 4月素材(V-高效抓宠/V-核心玩法/P-灾后重建)放置了4-5个月，建议重新制作/翻新再测，避免素材疲劳导致CPI虚高。',
    '8. [机动预算] $5-8K留给Task 2的新方向素材脚本(等你竞品素材后产出)。',
]

for i, note in enumerate(notes):
    row = 56 + i
    ws.merge_cells(f'A{row}:P{row}')
    ws[f'A{row}'] = note
    ws[f'A{row}'].font = Font(name='Arial', size=10, color='333333')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[row].height = 22

# ======================================================
# Creative count for estimates
# ======================================================
row = 65
sec_hdr(ws, row, '▌素材数量与量级匹配（自检）')
row = 66
ws.merge_cells(f'A{row}:P{row}')
ws[f'A{row}'] = '共18个素材-出价组合(14个独立素材概念)。按$40-50K预算+CPI$5-8估算，可产出5,000-8,000个Install。每个素材平均DNU>300，统计上有意义。素材数量与预算量级匹配。'
ws[f'A{row}'].font = Font(name='Arial', size=10, color='333333')
ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

row = 67
ws.merge_cells(f'A{row}:P{row}')
ws[f'A{row}'] = '⚠ 如果预算被砍到$30K以下，优先砍Tier 3辅测+机动预算，保留Tier 1+2核心。'
ws[f'A{row}'].font = Font(name='Arial', size=10, color='CC0000', bold=True)
ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb.save('creative_final.xlsx')
print('Done! v2 sheet added with 18 creatives.')
