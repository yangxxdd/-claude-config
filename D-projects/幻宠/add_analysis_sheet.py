import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# ============================================================
# Load data
# ============================================================
cf = pd.read_excel('creative_final.xlsx', sheet_name='素材表现汇总')
cf['mat_name'] = cf['素材名称'].astype(str).str.strip()
cf['bid'] = cf['出价方式'].astype(str).str.strip()
cf['month'] = cf['月份'].astype(str).str.strip().str[:2]
cf['dnu'] = pd.to_numeric(cf['DNU'], errors='coerce')

data = cf[(cf['mat_name'].notna()) & (cf['mat_name'] != 'nan') &
          (~cf['月份'].astype(str).str.startswith('◆')) & (cf['R1'].notna())].copy()

# ============================================================
# Create analysis data
# ============================================================

# --- Part 1: Both months reliable ---
pivot = data.groupby(['mat_name', 'bid']).agg(
    months=('month', 'nunique'),
).reset_index()
both_names = pivot[pivot['months'] == 2]

both_data = []
for _, bm in both_names.iterrows():
    name, bid = bm['mat_name'], bm['bid']
    r4 = data[(data['mat_name'] == name) & (data['bid'] == bid) & (data['month'] == '4月')]
    r6 = data[(data['mat_name'] == name) & (data['bid'] == bid) & (data['month'] == '6月')]
    if len(r4) == 0 or len(r6) == 0: continue

    dnu4, dnu6 = r4.iloc[0]['dnu'], r6.iloc[0]['dnu']
    reliable = dnu4 >= 30 and dnu6 >= 30

    both_data.append({
        '素材名称': name, '出价方式': bid,
        '4月_CPI': r4.iloc[0]['CPI'], '4月_R1': r4.iloc[0]['R1'], '4月_R3': r4.iloc[0]['R3'], '4月_DNU': dnu4,
        '6月_CPI': r6.iloc[0]['CPI'], '6月_R1': r6.iloc[0]['R1'], '6月_R3': r6.iloc[0]['R3'], '6月_DNU': dnu6,
        'R1变化': r6.iloc[0]['R1'] - r4.iloc[0]['R1'],
        'R3变化': r6.iloc[0]['R3'] - r4.iloc[0]['R3'],
        'CPI变化': r6.iloc[0]['CPI'] - r4.iloc[0]['CPI'],
        '双月可靠': '是' if reliable else '否',
    })

df_both = pd.DataFrame(both_data)
df_both_reliable = df_both[df_both['双月可靠'] == '是'].sort_values('4月_R1', ascending=False)
df_both_all = df_both.sort_values('双月可靠', ascending=False)

# --- Part 2: Small sample outstanding ---
small = data[data['dnu'] < 30].copy()
small_filtered = small[small['dnu'] >= 5].copy()

top_r1 = small_filtered.nlargest(5, 'R1')[['mat_name', 'bid', 'month', 'R1', 'R2', 'R3', 'dnu', 'CPI', '花费(USD)']]
top_r3 = small_filtered.nlargest(5, 'R3')[['mat_name', 'bid', 'month', 'R1', 'R2', 'R3', 'dnu', 'CPI', '花费(USD)']]
small_filtered['composite'] = small_filtered['R1'] * 0.5 + small_filtered['R3'] * 0.5
top_comp = small_filtered.nlargest(5, 'composite')[['mat_name', 'bid', 'month', 'R1', 'R2', 'R3', 'dnu', 'CPI', '花费(USD)', 'composite']]
top_cpi = small_filtered.nsmallest(5, 'CPI')[['mat_name', 'bid', 'month', 'R1', 'R2', 'R3', 'dnu', 'CPI', '花费(USD)']]

# ============================================================
# Write to Excel
# ============================================================
wb = load_workbook('creative_final.xlsx')

# Remove existing analysis sheet if present
if '素材分析-双月对比与推荐' in wb.sheetnames:
    del wb['素材分析-双月对比与推荐']

ws = wb.create_sheet('素材分析-双月对比与推荐')

# Styles
title_font = Font(name='Arial', size=14, bold=True, color='1F4E79')
h2_font = Font(name='Arial', size=12, bold=True, color='2E75B6')
header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='2E75B6')
green_fill = PatternFill('solid', fgColor='C6EFCE')
yellow_fill = PatternFill('solid', fgColor='FFFFCC')
red_fill = PatternFill('solid', fgColor='FFC7CE')
light_blue_fill = PatternFill('solid', fgColor='DAEEF3')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', size=10, bold=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def write_header(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def write_data_row(ws, row, values, col_start=1, formats=None, fills=None):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start + i, value=v)
        cell.font = normal_font
        cell.alignment = center_align if i > 0 else left_align
        cell.border = thin_border
        if formats and i < len(formats) and formats[i]:
            cell.number_format = formats[i]
        if fills and i < len(fills) and fills[i]:
            cell.fill = fills[i]

r = 1  # current row

# ============================================================
# SECTION 1: Both tests reliable
# ============================================================
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
ws.cell(row=r, column=1, value='一、两次测试都数据好的素材（4月+6月 双月 DNU ≥ 30）').font = title_font
r += 1
ws.cell(row=r, column=1, value='标准：仅4个素材-出价组合满足双月样本充足。CPI和留存分开列，方便对比。').font = Font(name='Arial', size=9, color='666666')
r += 2

headers1 = ['素材名称', '出价方式', '4月 CPI', '4月 R1', '4月 R3', '4月 DNU',
            '6月 CPI', '6月 R1', '6月 R3', '6月 DNU', 'R1 变化', 'R3 变化', 'CPI 变化']
write_header(ws, r, headers1)
r += 1

for _, row in df_both_reliable.iterrows():
    name = row['素材名称']
    bid = row['出价方式']
    vals = [name, bid, row['4月_CPI'], row['4月_R1'], row['4月_R3'], row['4月_DNU'],
            row['6月_CPI'], row['6月_R1'], row['6月_R3'], row['6月_DNU'],
            row['R1变化'], row['R3变化'], row['CPI变化']]
    fmt = [None, None, '$#,##0.00', '0.0%', '0.0%', '#,##0',
           '$#,##0.00', '0.0%', '0.0%', '#,##0', '0.0%', '0.0%', '+$#,##0.00']

    # Color R1 change: green if up, red if down significantly
    fills = [None]*13
    r1_chg = row['R1变化']
    if r1_chg > 0.01:
        fills[10] = green_fill
    elif r1_chg < -0.05:
        fills[10] = red_fill

    write_data_row(ws, r, vals, formats=fmt, fills=fills)
    r += 1

r += 1

# --- All cross-month materials ---
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=13)
ws.cell(row=r, column=1, value='附：所有跨月素材总览（含小样本，仅供参考）').font = h2_font
r += 1

write_header(ws, r, ['素材名称', '出价方式', '4月 CPI', '4月 R1', '4月 R3', '4月 DNU',
                      '6月 CPI', '6月 R1', '6月 R3', '6月 DNU', 'R1 变化', 'R3 变化', '数据质量'])
r += 1

for _, row in df_both_all.iterrows():
    name = row['素材名称']
    bid = row['出价方式']
    reliable = row['双月可靠'] == '是'
    quality = '✅可靠' if reliable else ('⚠小样本' if (row['4月_DNU'] + row['6月_DNU']) >= 20 else '❌噪音')

    vals = [name, bid, row['4月_CPI'], row['4月_R1'], row['4月_R3'], row['4月_DNU'],
            row['6月_CPI'], row['6月_R1'], row['6月_R3'], row['6月_DNU'],
            row['R1变化'], row['R3变化'], quality]
    fmt = [None, None, '$#,##0.00', '0.0%', '0.0%', '#,##0',
           '$#,##0.00', '0.0%', '0.0%', '#,##0', '0.0%', '0.0%', None]

    fills = [None]*13
    if reliable:
        fills[0] = green_fill
    elif '噪音' in quality:
        fills[0] = red_fill
    else:
        fills[0] = yellow_fill

    write_data_row(ws, r, vals, formats=fmt, fills=fills)
    r += 1

r += 2

# ============================================================
# SECTION 2: Small sample outstanding
# ============================================================
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
ws.cell(row=r, column=1, value='二、样本不足但表现亮眼 — 建议重新测试').font = title_font
r += 1
ws.cell(row=r, column=1, value='筛选标准：DNU < 30 且 ≥ 5（过滤绝对噪音）。绿色=强烈推荐重测。').font = Font(name='Arial', size=9, color='666666')
r += 2

# 2a. Top 5 R1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
ws.cell(row=r, column=1, value='2a. Top 5 次日留存 (R1) 最高').font = h2_font
r += 1
write_header(ws, r, ['排名', '素材名称', '出价方式', '月份', 'R1', 'R2', 'R3', 'DNU', 'CPI', '推荐理由'])
r += 1

reasons_r1 = [
    '⭐必测！R1天花板52.9%',
    'R1=50%且R3=16.7%也不错',
    '⭐R1+R3双高+CPI极低',
    'R1亮眼但R3=0%需观察',
    'R1高但CPI偏贵',
]
for i, (_, row) in enumerate(top_r1.iterrows()):
    vals = [i+1, row['mat_name'], row['bid'], row['month'], row['R1'], row['R2'], row['R3'], row['dnu'], row['CPI'], reasons_r1[i]]
    fmt = [None, None, None, None, '0.0%', '0.0%', '0.0%', '#,##0', '$#,##0.00', None]
    fills = [green_fill if i < 3 else None]*10
    write_data_row(ws, r, vals, formats=fmt, fills=fills)
    r += 1

r += 1

# 2b. Top 5 R3
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
ws.cell(row=r, column=1, value='2b. Top 5 三日留存 (R3) 最高').font = h2_font
r += 1
write_header(ws, r, ['排名', '素材名称', '出价方式', '月份', 'R1', 'R2', 'R3', 'DNU', 'CPI', '推荐理由'])
r += 1

reasons_r3 = [
    '⭐R1=44%+R3=33%+CPI=$4.13三冠王',
    'R3异常高33%但R1仅17%需验证',
    'R1=R3=20%均衡但DNU仅5',
    'R1=36%+R3=18%但CPI贵',
    'R3=17%尚可，整体一般',
]
for i, (_, row) in enumerate(top_r3.iterrows()):
    vals = [i+1, row['mat_name'], row['bid'], row['month'], row['R1'], row['R2'], row['R3'], row['dnu'], row['CPI'], reasons_r3[i]]
    fmt = [None, None, None, None, '0.0%', '0.0%', '0.0%', '#,##0', '$#,##0.00', None]
    fills = [green_fill if i < 2 else None]*10
    write_data_row(ws, r, vals, formats=fmt, fills=fills)
    r += 1

r += 1

# 2c. Top 5 composite
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
ws.cell(row=r, column=1, value='2c. Top 5 综合最优 (R1×0.5 + R3×0.5)').font = h2_font
r += 1
write_header(ws, r, ['排名', '素材名称', '出价方式', '月份', 'R1', 'R2', 'R3', 'DNU', 'CPI', '综合分'])
r += 1

for i, (_, row) in enumerate(top_comp.iterrows()):
    vals = [i+1, row['mat_name'], row['bid'], row['month'], row['R1'], row['R2'], row['R3'], row['dnu'], row['CPI'], row['composite']]
    fmt = [None, None, None, None, '0.0%', '0.0%', '0.0%', '#,##0', '$#,##0.00', '0.0%']
    fills = [green_fill if i < 2 else None]*10
    write_data_row(ws, r, vals, formats=fmt, fills=fills)
    r += 1

r += 1

# 2d. Top 5 low CPI
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
ws.cell(row=r, column=1, value='2d. Top 5 最低CPI (获客成本最低)').font = h2_font
r += 1
write_header(ws, r, ['排名', '素材名称', '出价方式', '月份', 'CPI', 'R1', 'R2', 'R3', 'DNU', '推荐理由'])
r += 1

reasons_cpi = [
    '⭐CPI仅$3.68 + DNU=27差3个达标',
    '⭐CPI=$4.13+R1=44%+R3=33%',
    'CPI低但R1仅10%不推荐',
    'CPI低+R3=33%但R1仅17%',
    'CPI=$5.60但留存一般',
]
for i, (_, row) in enumerate(top_cpi.iterrows()):
    vals = [i+1, row['mat_name'], row['bid'], row['month'], row['CPI'], row['R1'], row['R2'], row['R3'], row['dnu'], reasons_cpi[i]]
    fmt = [None, None, None, None, '$#,##0.00', '0.0%', '0.0%', '0.0%', '#,##0', None]
    fills = [green_fill if i < 2 else None]*10
    write_data_row(ws, r, vals, formats=fmt, fills=fills)
    r += 1

r += 2

# ============================================================
# SECTION 3: Priority recommendations
# ============================================================
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
ws.cell(row=r, column=1, value='三、优先重测清单').font = title_font
r += 1
ws.cell(row=r, column=1, value='综合考虑留存、CPI、样本量，排序推荐下次测试优先加量的素材。').font = Font(name='Arial', size=9, color='666666')
r += 2

write_header(ws, r, ['优先级', '素材名称', '出价方式', '月份', 'R1', 'R3', 'DNU', 'CPI', '核心理由', '建议动作'])
r += 1

recommendations = [
    ['🥇 最优先', 'P-灾后重建', 'Install', '4月', 0.444, 0.333, 9, 4.13, 'R1=44%+R3=33%+CPI=$4.13三冠王，性价比无敌', '加预算冲50+ DNU验证'],
    ['🥈 第二', 'V-帕萌战斗-群殴打鸡', 'Install', '6月', 0.529, 0.059, 17, 8.21, 'R1=53%全素材最高，17 DNU差一点达标', '加预算冲50+ DNU，关注R3能否跟上'],
    ['🥉 第三', 'V-场景展示', 'AEO', '6月', 0.500, 0.167, 12, 8.89, 'R1=50%+R3=17%，且Install版(DNU=76)验证过R1=32%', 'AEO+Install双出价重测'],
    ['4', 'V-模拟经营-帕基玩法寒霜', 'Install', '6月', 0.364, 0.182, 11, 19.27, 'R1=36%+R3=18%留存好，但CPI太贵需看能否优化', '单独开campaign控CPI<$10重测'],
    ['5', 'P-宠物展示-二阶进化', 'Install', '6月', 0.333, 0.074, 27, 3.68, 'CPI=$3.68极低，DNU=27只差3个达标，R1=33%不错', '加$100预算即可验证'],
]

for rec in recommendations:
    fills = [green_fill] + [None]*9
    vals = rec
    fmt = [None, None, None, None, '0.0%', '0.0%', '#,##0', '$#,##0.00', None, None]
    write_data_row(ws, r, vals, formats=fmt, fills=fills)
    r += 1

r += 2

# ============================================================
# SECTION 4: Methodology note
# ============================================================
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
ws.cell(row=r, column=1, value='分析方法说明').font = h2_font
r += 1
notes = [
    '• 数据来源：fb媒体-源数据.xlsx → BI汇总 (素材ID匹配留存) → creative_final.xlsx',
    '• R1/R2/R3 = 次日/2日/3日留存 = r1_cnt/dnu, r2_cnt/dnu, r3_cnt/dnu (DNU加权平均)',
    '• "双月可靠" = 4月和6月DNU均≥30，留存率统计上有参考价值',
    '• "小样本" = DNU < 30，留存率仅供参考，需加大样本重新验证',
    '• "噪音" = DNU < 10，留存率不可靠，不作为决策依据',
    '• 综合分 = R1×0.5 + R3×0.5，兼顾短期和中期留存',
    '• 分析日期：2026-07-29',
]
for note in notes:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    ws.cell(row=r, column=1, value=note).font = Font(name='Arial', size=9, color='666666')
    r += 1

# ============================================================
# Column widths
# ============================================================
col_widths = {1: 18, 2: 26, 3: 12, 4: 8, 5: 10, 6: 10, 7: 10, 8: 10, 9: 12, 10: 28,
              11: 12, 12: 12, 13: 14}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# Freeze first row after the section headers? Actually freeze at row 1
ws.freeze_panes = 'A2'

# Add auto-filter on section 1
# ws.auto_filter.ref = f'A1:{get_column_letter(13)}{r}'

wb.save('creative_final.xlsx')
print('Sheet "素材分析-双月对比与推荐" added successfully!')
