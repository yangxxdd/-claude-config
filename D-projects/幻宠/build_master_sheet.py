import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

df = pd.read_excel('creative_final.xlsx', sheet_name='素材表现汇总')
df = df[df['月份'].isin(['4月','6月'])].copy()

# Tag mapping (from creative review)
tag_map = {
    'P-海啸': '天灾/生存',
    'P-幻想宠物3': '宠物展示, 收集',
    'P-灾后重建': '模拟经营, 建造',
    'V-场景展示': '模拟经营, 建造',
    'V-高效抓宠': '抓宠, 战斗, 宠物展示',
    'V-核心玩法': '战斗, 抓宠, 探索/冒险',
    'V-帕基世界冒险': '探索/冒险, 模拟经营',
    'V-收集升级': '模拟经营, 建造, 收集',
    'V-天灾重建-长版': '模拟经营, 建造, 收集',
    'V-抓宠战斗': '抓宠, 战斗',
    'P-宠物展示-二阶进化': '宠物进化, 宠物展示',
    'P-宠物展示-合成 3D': '宠物融合/合成, 宠物进化, 宠物展示',
    'P-宠物展示-巨物': '宠物展示',
    'P-宠物展示-帕基战斗': '战斗',
    'P-模拟经营-冬日写实': '模拟经营, 建造',
    'P-模拟经营-建造成长': '模拟经营, 建造, 收集',
    'P-抓宠经营-超梦': '抓宠',
    'P-抓宠经营-虐待': '抓宠',
    'P-抓宠经营-血腥': '经营',
    'P-抓宠经营-幽飘': '抓宠, 探索/冒险',
    'V-宠物展示-宠物合成': '宠物展示, 宠物融合/合成',
    'V-宠物展示-二阶合成': '宠物展示, 宠物进化',
    'V-模拟经营-砍树': '模拟经营, 建造, 收集, 战斗',
    'V-模拟经营-帕基玩法寒霜': '模拟经营, 探索/冒险, 收集',
    'V-模拟经营-七日建造': '模拟经营, 建造',
    'V-模拟经营-温馨': '模拟经营',
    'V-帕萌战斗-出狱打鸡': '战斗',
    'V-帕萌战斗-合成狙击': '战斗, 抓宠',
    'V-帕萌战斗-群殴打鸡': '战斗',
    'V-帕萌战斗-杀宠复刻': '战斗',
    'V-帕萌战斗-雪地竞品': '战斗',
    'V-抓宠经营-捕捞竞品': '抓宠, 模拟经营',
    'V-抓宠经营-狐狸': '战斗',
    'V-抓宠经营-虐待': '战斗, 抓宠',
    'V-抓宠经营-售卖帕基': '模拟经营, 建造',
    'V-抓宠经营-拯救可达鸭': '战斗, 建造',
}

def get_tags(name):
    for key in tag_map:
        if key in str(name):
            return tag_map[key]
    return '未分类'

# Core direction classification
def get_direction(tags_str):
    tags = tags_str.split(', ')
    dirs = []
    if '抓宠' in tags: dirs.append('捉宠')
    if '宠物融合/合成' in tags: dirs.append('融合')
    if '宠物进化' in tags: dirs.append('进化')
    if '模拟经营' in tags or '建造' in tags: dirs.append('模拟经营/建造')
    if '战斗' in tags: dirs.append('战斗')
    if '天灾/生存' in tags: dirs.append('天灾/生存')
    if '经营' in tags and '模拟经营' not in tags: dirs.append('经营/其他')
    if '宠物展示' in tags and not dirs: dirs.append('宠物展示')
    if not dirs: dirs.append('其他')
    return '/'.join(dirs)

# Build records
records = []
for _, row in df.iterrows():
    name = str(row['素材名称'])
    if name == 'nan' or '合计' in name or '◆' in name:
        continue
    try:
        dnu = int(float(row['DNU']))
        cpi = float(row['CPI'])
        r1 = float(row['R1'])
        r2 = float(row['R2'])
        r3 = float(row['R3'])
        cost22 = float(row['次留成本']) if not pd.isna(row['次留成本']) else None
        spend = float(row['花费(USD)'])
        cvr = float(row['CVR'])
        ctr = float(row['CTR'])
        installs = int(float(row['安装数']))
        impr = int(float(row['展示次数']))
        clicks = int(float(row['点击量']))
    except (ValueError, TypeError):
        continue

    tags = get_tags(name)
    direction = get_direction(tags)
    month = row['月份']
    bid = row['出价方式']
    note = str(row['备注']) if not pd.isna(row['备注']) else ''

    records.append({
        'name': name, 'direction': direction, 'tags': tags,
        'month': month, 'bid': bid, 'dnu': dnu, 'cpi': cpi,
        'r1': r1, 'r2': r2, 'r3': r3, 'cost22': cost22,
        'spend': spend, 'cvr': cvr, 'ctr': ctr, 'installs': installs,
        'impr': impr, 'clicks': clicks, 'note': note
    })

df_all = pd.DataFrame(records)

# Group by creative name, sort by direction priority then name
dir_order = {'捉宠': 1, '融合': 2, '进化': 3, '战斗': 4, '模拟经营/建造': 5, '天灾/生存': 6, '其他': 7}
unique_names = df_all.groupby('name').agg({
    'direction': 'first',
    'tags': 'first',
    'dnu': 'sum',
    'spend': 'sum'
}).reset_index()

# Assign primary direction for sorting
def sort_key(row):
    d = row['direction']
    for k in dir_order:
        if k in d:
            return dir_order[k]
    return 99

unique_names['sort_key'] = unique_names.apply(sort_key, axis=1)
unique_names = unique_names.sort_values(['sort_key', 'name'])

# Scoring function for color
def score_row(cpi, r1, r3, dnu, month):
    s = 0
    if cpi < 5: s += 3
    elif cpi < 7: s += 2
    elif cpi < 9: s += 1
    if r1 >= 0.35: s += 3
    elif r1 >= 0.28: s += 2
    elif r1 >= 0.20: s += 1
    if r3 >= 0.15: s += 3
    elif r3 >= 0.10: s += 2
    elif r3 >= 0.07: s += 1
    if dnu >= 30: s += 1
    if dnu < 10: s -= 1
    if month == '4月': s -= 1
    return s

# Excel
wb = load_workbook('creative_final.xlsx')
for s in ['素材全量数据-按素材', '素材全量数据-图例']:
    if s in wb.sheetnames:
        del wb[s]

ws = wb.create_sheet('素材全量数据-按素材', 0)

# Styles
hdr_f = Font(name='Arial', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='2F5496')
grp_fill = PatternFill('solid', fgColor='D6E4F0')
grp_font = Font(name='Arial', bold=True, size=11, color='1F4E79')
green_fill = PatternFill('solid', fgColor='C6EFCE')
yellow_fill = PatternFill('solid', fgColor='FFF2CC')
red_fill = PatternFill('solid', fgColor='F4B4C2')
gray_fill = PatternFill('solid', fgColor='E0E0E0')
white_fill = PatternFill('solid', fgColor='FFFFFF')
nf = Font(name='Arial', size=10)
bf = Font(name='Arial', size=10, bold=True)
rf = Font(name='Arial', size=10, color='CC0000')
thin_b = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
la = Alignment(horizontal='left', vertical='center', wrap_text=True)

headers = ['素材名称', '方向归类', '内容标签', '月份', '出价方式', 'DNU', 'CPI(USD)', 'R1', 'R2', 'R3',
           '次留成本', '花费(USD)', '安装数', '展示次数', '点击量', 'CTR', 'CVR', '备注']

col_w = [26, 16, 26, 8, 10, 8, 12, 10, 10, 10, 12, 12, 10, 12, 10, 10, 10, 30]
for i, w in enumerate(col_w, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title
row = 1
ws.merge_cells('A1:R1')
ws['A1'] = '幻宠素材全量数据 — 以素材为主维度，月份+出价方式为下钻（标色逻辑：绿=推荐 黄=可选 红=不推荐 灰=噪音/重复）'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 32

# Legend
row = 2
ws.merge_cells('A2:R2')
ws['A2'] = '🟢绿(≥7分) CPI<$7+R1>28%+R3>10% | 🟡黄(5-6分) 有亮点但有短板 | 🔴红(<5分) 不推荐 | ⬜灰 DNU<10噪音 | 排序：方向优先级(捉宠>融合>进化>战斗>模拟经营>天灾) → 素材名 → 6月优先 → CPI升序'
ws['A2'].font = Font(name='Arial', size=9, color='555555')
ws['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws.row_dimensions[2].height = 24

# Column headers
row = 3
for i, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.font = hdr_f; c.fill = hdr_fill; c.alignment = ca; c.border = thin_b
ws.row_dimensions[row].height = 32

# Track duplicate names across months
from collections import Counter
name_months = df_all.groupby('name')['month'].apply(set).to_dict()
dup_names = {n for n, ms in name_months.items() if len(ms) > 1}

row = 4
prev_name = None

for _, name_row in unique_names.iterrows():
    name = name_row['name']
    direction = name_row['direction']
    tags = name_row['tags']
    is_dup = name in dup_names

    # Get all records for this creative (June first, then April; CPI asc within month)
    name_data = df_all[df_all['name'] == name].copy()
    # Sort: June first, then April; within month by CPI asc
    name_data['month_order'] = name_data['month'].map({'6月': 0, '4月': 1})
    name_data = name_data.sort_values(['month_order', 'cpi'])

    is_first_row = True
    for _, rec in name_data.iterrows():
        score = score_row(rec['cpi'], rec['r1'], rec['r3'], rec['dnu'], rec['month'])
        is_noise = rec['dnu'] < 10

        if score >= 7 and not is_noise:
            fill = green_fill
        elif score >= 5 and not is_noise:
            fill = yellow_fill
        elif score < 5 and not is_noise:
            fill = red_fill
        else:
            fill = gray_fill

        ws.cell(row=row, column=1, value=name if is_first_row else '').font = bf
        ws.cell(row=row, column=1).alignment = la; ws.cell(row=row, column=1).border = thin_b
        ws.cell(row=row, column=1).fill = fill if is_first_row else white_fill

        ws.cell(row=row, column=2, value=direction if is_first_row else '').font = nf
        ws.cell(row=row, column=2).alignment = ca; ws.cell(row=row, column=2).border = thin_b
        ws.cell(row=row, column=2).fill = fill if is_first_row else white_fill

        ws.cell(row=row, column=3, value=tags if is_first_row else '').font = nf
        ws.cell(row=row, column=3).alignment = la; ws.cell(row=row, column=3).border = thin_b
        ws.cell(row=row, column=3).fill = fill if is_first_row else white_fill

        ws.cell(row=row, column=4, value=rec['month']).font = nf
        ws.cell(row=row, column=4).alignment = ca; ws.cell(row=row, column=4).border = thin_b
        ws.cell(row=row, column=4).fill = fill

        ws.cell(row=row, column=5, value=rec['bid']).font = nf
        ws.cell(row=row, column=5).alignment = ca; ws.cell(row=row, column=5).border = thin_b
        ws.cell(row=row, column=5).fill = fill

        ws.cell(row=row, column=6, value=rec['dnu']).font = nf
        ws.cell(row=row, column=6).alignment = ca; ws.cell(row=row, column=6).border = thin_b
        ws.cell(row=row, column=6).fill = fill

        ws.cell(row=row, column=7, value=round(rec['cpi'], 2)).font = nf
        ws.cell(row=row, column=7).alignment = ca; ws.cell(row=row, column=7).border = thin_b
        ws.cell(row=row, column=7).fill = fill

        ws.cell(row=row, column=8, value=f'{rec["r1"]:.1%}').font = nf
        ws.cell(row=row, column=8).alignment = ca; ws.cell(row=row, column=8).border = thin_b
        ws.cell(row=row, column=8).fill = fill

        ws.cell(row=row, column=9, value=f'{rec["r2"]:.1%}' if rec['r2'] > 0 else '—').font = nf
        ws.cell(row=row, column=9).alignment = ca; ws.cell(row=row, column=9).border = thin_b
        ws.cell(row=row, column=9).fill = fill

        ws.cell(row=row, column=10, value=f'{rec["r3"]:.1%}').font = nf
        ws.cell(row=row, column=10).alignment = ca; ws.cell(row=row, column=10).border = thin_b
        ws.cell(row=row, column=10).fill = fill

        cs_val = f'${rec["cost22"]:.2f}' if rec['cost22'] and rec['cost22'] > 0 else '—'
        ws.cell(row=row, column=11, value=cs_val).font = nf
        ws.cell(row=row, column=11).alignment = ca; ws.cell(row=row, column=11).border = thin_b
        ws.cell(row=row, column=11).fill = fill

        ws.cell(row=row, column=12, value=f'${rec["spend"]:.0f}').font = nf
        ws.cell(row=row, column=12).alignment = ca; ws.cell(row=row, column=12).border = thin_b
        ws.cell(row=row, column=12).fill = fill

        ws.cell(row=row, column=13, value=rec['installs']).font = nf
        ws.cell(row=row, column=13).alignment = ca; ws.cell(row=row, column=13).border = thin_b
        ws.cell(row=row, column=13).fill = fill

        ws.cell(row=row, column=14, value=rec['impr']).font = nf
        ws.cell(row=row, column=14).alignment = ca; ws.cell(row=row, column=14).border = thin_b
        ws.cell(row=row, column=14).fill = fill

        ws.cell(row=row, column=15, value=rec['clicks']).font = nf
        ws.cell(row=row, column=15).alignment = ca; ws.cell(row=row, column=15).border = thin_b
        ws.cell(row=row, column=15).fill = fill

        ws.cell(row=row, column=16, value=f'{rec["ctr"]:.2%}').font = nf
        ws.cell(row=row, column=16).alignment = ca; ws.cell(row=row, column=16).border = thin_b
        ws.cell(row=row, column=16).fill = fill

        ws.cell(row=row, column=17, value=f'{rec["cvr"]:.1%}').font = nf
        ws.cell(row=row, column=17).alignment = ca; ws.cell(row=row, column=17).border = thin_b
        ws.cell(row=row, column=17).fill = fill

        note_text = rec['note'] if rec['note'] else ''
        if rec['dnu'] < 10:
            note_text = ('⚠DNU<10噪音 ' + note_text).strip()
        if is_dup and any(kw in name for kw in ['场景展示', '高效抓宠', '核心玩法', '天灾重建', '抓宠战斗', '海啸']):
            note_text = ('📋4月6月内容相同 ' + note_text).strip()
        ws.cell(row=row, column=18, value=note_text if note_text else '—').font = nf
        ws.cell(row=row, column=18).alignment = la; ws.cell(row=row, column=18).border = thin_b
        ws.cell(row=row, column=18).fill = fill

        ws.row_dimensions[row].height = 22
        is_first_row = False
        row += 1

    # Blank separator row between creative groups
    ws.merge_cells(f'A{row}:R{row}')
    ws.row_dimensions[row].height = 6
    row += 1

wb.save('creative_final.xlsx')
print(f'Done! Sheet "素材全量数据-按素材" added, {row} rows total.')
print(f'Unique creatives: {len(unique_names)}')
