import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

wb = load_workbook('creative_final.xlsx')
ws = wb['素材全量数据-按素材']

# Column mapping (find by header row 3):
# A=素材名称, B=方向归类, C=内容标签, D=月份, E=出价方式
# F=DNU, G=CPI, H=R1, I=R2, J=R3, K=次留成本
# L=花费(USD), M=安装数, N=展示次数, O=点击量, P=CTR, Q=CVR, R=备注

fix_count = 0
r2_fix_count = 0
div0_fix_count = 0

for row in range(4, 92):
    k_cell = ws.cell(row=row, column=11)  # 次留成本
    f_cell = ws.cell(row=row, column=6)   # DNU
    l_cell = ws.cell(row=row, column=12)  # 花费
    h_cell = ws.cell(row=row, column=8)   # R1
    i_cell = ws.cell(row=row, column=9)   # R2

    # --- Fix 1: R2 ---
    # If R2 shows "—" but the source data has 0%, change to "0.0%"
    if i_cell.value is not None and str(i_cell.value).strip() == '—':
        # Check if this row should actually have 0% instead of —
        # R2 is "—" when r2_cnt=0 in source. Should show "0.0%" not "—"
        i_cell.value = '0.0%'
        r2_fix_count += 1

    # --- Fix 2: 次留成本 formula ---
    # User's formula: =L/(M*H) = 花费/(安装数×R1)
    # Correct formula: =L/(F*H) = 花费/(DNU×R1)
    # Because: 次留成本 = 花费/r1_cnt = 花费/(DNU×R1), not 花费/(安装数×R1)
    # DNU ≤ 安装数 due to BI matching loss
    if k_cell.value is not None and isinstance(k_cell.value, str) and k_cell.value.startswith('='):
        old_formula = k_cell.value
        # The formula references columns: L=12, M=13, H=8, F=6
        # User used M (安装数), should use F (DNU)
        # Pattern: =L{row}/(M{row}*H{row}) → =L{row}/(F{row}*H{row})
        new_formula = old_formula.replace(f'M{row}', f'F{row}')
        if new_formula != old_formula:
            k_cell.value = new_formula
            fix_count += 1

    # --- Fix 3: #DIV/0! when R1=0% ---
    # Already handled if we wrap in IFERROR... but the user's formula doesn't have IFERROR
    # Check if H cell has 0% - if so, wrap formula in IFERROR
    if k_cell.value is not None and isinstance(k_cell.value, str) and k_cell.value.startswith('='):
        try:
            h_val = float(str(h_cell.value).replace('%', '')) / 100 if h_cell.value else 0
        except:
            h_val = 0
        if h_val == 0:
            # Wrap in IFERROR to avoid #DIV/0!
            k_cell.value = f'=IFERROR({k_cell.value[1:]},"—")'
            div0_fix_count += 1

print(f'Fixed 次留成本 formula (安装数→DNU): {fix_count} rows')
print(f'Fixed R2 —→0.0%: {r2_fix_count} rows')
print(f'Fixed #DIV/0! with IFERROR: {div0_fix_count} rows')

# Verify fixes
print('\n=== Verification ===')
for row in [4, 10, 12, 53, 81, 86]:
    k = ws.cell(row=row, column=11)
    f = ws.cell(row=row, column=6)
    h = ws.cell(row=row, column=8)
    l = ws.cell(row=row, column=12)
    print(f'Row{row}: formula={k.value} | DNU={f.value} | R1={h.value} | 花费={l.value}')

wb.save('creative_final.xlsx')
print('\nSaved.')
