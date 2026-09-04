import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy
from collections import defaultdict

wb = load_workbook('creative_final.xlsx')
ws1 = wb['素材全量数据-按素材']

BLUE_FONT = Font(name='Arial', size=10, color='0066CC')
BLUE_BOLD = Font(name='Arial', size=10, color='0066CC', bold=True)
BLACK_FONT = Font(name='Arial', size=10, color='000000')

# =====================================================
# PHASE 1: Define absolute tier for every creative+bidding
# =====================================================
# Read all data rows from Sheet 1 (handle merged cells properly)
rows_data = []
current_name = None
for row in range(4, 92):
    raw_name = ws1.cell(row=row, column=1).value
    month = ws1.cell(row=row, column=4).value
    bid = ws1.cell(row=row, column=5).value

    # Handle merged cells: propagate name from parent row
    if raw_name and str(raw_name).strip():
        current_name = str(raw_name).strip()

    if not month:
        continue

    name = current_name  # use propagated name

    dnu_v = ws1.cell(row=row, column=6).value
    cpi_v = ws1.cell(row=row, column=7).value
    r1_v = ws1.cell(row=row, column=8).value
    r3_v = ws1.cell(row=row, column=10).value
    note_v = ws1.cell(row=row, column=19).value

    try:
        dnu = int(float(dnu_v)) if dnu_v else 0
        cpi = float(cpi_v) if cpi_v else 0
        r1 = float(str(r1_v).replace('%',''))/100 if r1_v else 0
        r3 = float(str(r3_v).replace('%',''))/100 if r3_v else 0
    except:
        continue

    rows_data.append({
        'row': row, 'name': str(name), 'month': str(month), 'bid': str(bid),
        'dnu': dnu, 'cpi': cpi, 'r1': r1, 'r3': r3,
        'note': str(note_v) if note_v else ''
    })

# ---- Tier logic (ABSOLUTE, 6月-primary, 4月-secondary) ----
# 🥇 强烈推荐: DNU>=10, CPI competitive + retention has clear strength
#   - 6月: CPI<$9 AND (R1>=28% OR R3>=10%) AND not both weak
#   - 4月: Same thresholds BUT labeled "4月优异,需6月重验"
# 🥈 可选: Has value but with significant caveat
#   - CPI<$12, not terrible retention, OR 4月 great data needing verification
# 🥉 待观察: DNU<10 with interesting signal worth tiny test
# ❌ 不推荐: Clear failure — high CPI + weak retention, or DNU<10 noise

def get_tier_and_reason(d):
    dnu, cpi, r1, r3, month = d['dnu'], d['cpi'], d['r1'], d['r3'], d['month']
    noise = dnu < 10
    small = dnu < 30
    is_june = month == '6月'

    reasons = []
    tier = None

    # Build reasons
    if cpi < 5:
        reasons.append(f'CPI=${cpi:.2f}极低')
    elif cpi < 7:
        reasons.append(f'CPI=${cpi:.2f}较低')
    elif cpi < 9:
        reasons.append(f'CPI=${cpi:.2f}适中')
    elif cpi < 13:
        reasons.append(f'CPI=${cpi:.2f}偏高')
    else:
        reasons.append(f'CPI=${cpi:.2f}过高')

    if r1 >= 0.35:
        reasons.append(f'R1={r1:.0%}优秀(≥35%)')
    elif r1 >= 0.28:
        reasons.append(f'R1={r1:.0%}尚可')
    elif r1 >= 0.20:
        reasons.append(f'R1={r1:.0%}偏低')
    else:
        reasons.append(f'R1={r1:.0%}差')

    if r3 >= 0.15:
        reasons.append(f'R3={r3:.0%}优秀(≥15%)')
    elif r3 >= 0.10:
        reasons.append(f'R3={r3:.0%}达标(≥10%)')
    elif r3 >= 0.07:
        reasons.append(f'R3={r3:.0%}一般')
    elif r3 > 0:
        reasons.append(f'R3={r3:.0%}偏低')
    else:
        reasons.append(f'R3=0%无三日留存')

    if noise:
        reasons.append(f'⚠DNU={dnu}<10样本不足')
    elif small:
        reasons.append(f'⚠DNU={dnu}偏小需扩量验证')
    else:
        reasons.append(f'DNU={dnu}样本可靠')

    if not is_june:
        reasons.append('⚠仅4月数据需6月重验')

    reason_str = '; '.join(reasons)

    # Determine tier
    if noise:
        # Noise: only recommend if metrics are exceptionally good
        if cpi < 7 and r1 >= 0.35 and r3 >= 0.10:
            tier = '🥈 可选'
        else:
            tier = '🥉 待观察' if (cpi < 10 and (r1 >= 0.30 or r3 >= 0.15)) else '❌ 不推荐'
    elif not is_june:
        # 4月 only: downgrade by one notch
        if cpi < 5 and r1 >= 0.35 and r3 >= 0.10:
            tier = '🥇 强烈推荐'  # exceptionally good even for 4月
        elif cpi < 8 and r1 >= 0.28:
            tier = '🥈 可选'
        elif cpi < 10 and (r1 >= 0.25 or r3 >= 0.10):
            tier = '🥉 待观察'
        else:
            tier = '❌ 不推荐'
    else:
        # 6月 data, DNU>=10 — main evaluation
        cpi_good = cpi < 7
        cpi_ok = cpi < 9
        r1_good = r1 >= 0.28
        r1_ok = r1 >= 0.20
        r3_good = r3 >= 0.10
        r3_ok = r3 >= 0.07

        if cpi_good and r1_good and r3_good:
            tier = '🥇 强烈推荐'  # all 3 metrics good
        elif cpi_good and (r1_good or r3_good):
            tier = '🥇 强烈推荐'  # CPI very good + at least one retention good
        elif cpi_ok and r1_good and r3_good:
            tier = '🥇 强烈推荐'  # CPI OK but both retention good
        elif cpi_good and r1_ok:
            tier = '🥈 可选'  # CPI great but retention mediocre
        elif cpi_ok and (r1_good or r3_good):
            tier = '🥈 可选'  # CPI OK + one retention good
        elif cpi_ok and r1_ok and r3_ok:
            tier = '🥉 待观察'  # everything mid
        elif cpi < 12 and (r1 >= 0.30 or r3 >= 0.15):
            tier = '🥉 待观察'  # high CPI but exceptional retention
        else:
            tier = '❌ 不推荐'

    return tier, reason_str

# Apply to all rows
for d in rows_data:
    tier, reason = get_tier_and_reason(d)
    d['tier'] = tier
    d['reason'] = reason

# Print summary for review
from collections import Counter
tier_counts = Counter(d['tier'] for d in rows_data)
print('=== Tier Distribution ===')
for t, c in tier_counts.most_common():
    print(f'  {t}: {c}')

# =====================================================
# PHASE 2: Update Sheet 1 — add tier column + update 备注
# =====================================================

# Add tier column at Col21 (after 备注 Col19 + 确认 Col20)
tier_col = 21
print(f'Adding tier column at Col{tier_col}')

# Check if tier header already exists
existing = ws1.cell(row=3, column=tier_col).value
if existing:
    print(f'  Warning: Col{tier_col} already has value: {existing}')
    # Clear old tier values
    for row in range(4, 92):
        ws1.cell(row=row, column=tier_col).value = None

# Write tier header
ws1.cell(row=3, column=tier_col, value='推荐等级').font = Font(name='Arial', bold=True, size=10, color='0066CC')
ws1.cell(row=3, column=tier_col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws1.cell(row=3, column=tier_col).border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

# Unmerge any ranges that include tier_col to avoid MergedCell errors
ranges_to_unmerge = []
for mc in list(ws1.merged_cells.ranges):
    if mc.min_col <= tier_col <= mc.max_col and mc.min_row >= 4:
        ranges_to_unmerge.append(str(mc))
        ws1.unmerge_cells(str(mc))
if ranges_to_unmerge:
    print(f'Unmerged {len(ranges_to_unmerge)} ranges affecting Col{tier_col}')

# Update 备注 and write tier for each row
for d in rows_data:
    row = d['row']
    # Update 备注 (unmerge first to avoid MergedCell errors)
    # Check if note cell is part of a merged range
    note_cell = ws1.cell(row=row, column=19)
    is_merged_note = False
    for mc in ws1.merged_cells.ranges:
        if mc.min_row <= row <= mc.max_row and mc.min_col <= 19 <= mc.max_col:
            is_merged_note = True
            break
    if is_merged_note:
        continue  # skip merged 备注 cells (part of separator rows)
    current_note = str(note_cell.value or '')
    old_patterns = ['📋4月6月内容相同', 'CPI适中', 'CPI极低', 'CPI较低', 'CPI偏高', 'CPI过高',
                    '⚠DNU<10噪音', '⚠仅4月数据', 'DNU=', 'R1=', 'R3=']
    is_old_style = any(p in current_note for p in old_patterns) if current_note and current_note != 'None' else True
    is_empty = not current_note or current_note == 'None' or current_note.strip() == '—'

    if is_old_style or is_empty:
        note_cell.value = d['reason']
        note_cell.font = BLUE_FONT

    # Write tier (skip if row is part of a merged separator row)
    tier_cell = ws1.cell(row=row, column=tier_col)
    try:
        tier_cell.value = d['tier']
    except AttributeError:
        continue  # MergedCell - separator row, skip
    tier_cell.font = BLUE_BOLD
    tier_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tier_cell.border = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    # Color tier cell
    if '🥇' in d['tier']:
        tier_cell.fill = PatternFill('solid', fgColor='C6EFCE')
    elif '🥈' in d['tier']:
        tier_cell.fill = PatternFill('solid', fgColor='BDD7EE')
    elif '🥉' in d['tier']:
        tier_cell.fill = PatternFill('solid', fgColor='FCE4D6')
    else:
        tier_cell.fill = PatternFill('solid', fgColor='F4B4C2')

# =====================================================
# PHASE 3: Rebuild Sheet 2 — direction summary (derived from Sheet 1 tiers)
# =====================================================
if '9月推荐-按方向' in wb.sheetnames:
    del wb['9月推荐-按方向']

ws2 = wb.create_sheet('9月推荐-按方向')

# Group by direction (handle merged cells for Col2)
direction_groups = defaultdict(list)
current_dir = None
for row in range(4, 92):
    # Read direction from Col2, propagate through merged cells
    raw_dir = ws1.cell(row=row, column=2).value
    month = ws1.cell(row=row, column=4).value
    if not month:
        continue
    if raw_dir and str(raw_dir).strip():
        current_dir = str(raw_dir).strip()
    # Find matching row_data
    for d in rows_data:
        if d['row'] == row:
            if current_dir:
                direction_groups[current_dir].append(d)
            break

# Sort directions by priority
dir_priority = {'捉宠':1,'捉宠/战斗':2,'捉宠/模拟经营/建造':3,'融合':4,'融合/进化':5,'进化':6,
                '战斗':7,'模拟经营/建造':8,'模拟经营/建造/战斗':9,'宠物展示':10,
                '天灾/生存':11,'经营/其他':12,'其他':13}
sorted_dirs = sorted(direction_groups.keys(), key=lambda x: dir_priority.get(x.split('/')[0], 99))

# Write to Sheet 2
hdr_f = Font(name='Arial', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='2F5496')
sec_fill = PatternFill('solid', fgColor='D6E4F0')
nf = Font(name='Arial', size=10)
thin_b = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
la = Alignment(horizontal='left', vertical='center', wrap_text=True)

row = 1
ws2.merge_cells('A1:H1')
ws2['A1'] = '幻宠 9月测试素材推荐 — 按方向归类（统一推荐等级，由Sheet1衍生）'
ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
ws2.row_dimensions[1].height = 30

for i, w in enumerate([6, 12, 26, 10, 10, 10, 10, 50], 1):
    ws2.column_dimensions[chr(64+i)].width = w

row = 3

for dir_name in sorted_dirs:
    dir_data = direction_groups[dir_name]
    # Sort by tier then CPI
    tier_order = {'🥇 强烈推荐':0, '🥈 可选':1, '🥉 待观察':2, '❌ 不推荐':3}
    dir_data.sort(key=lambda x: (tier_order.get(x['tier'],99), x['cpi']))

    # Section header
    ws2.merge_cells(f'A{row}:H{row}')
    ws2[f'A{row}'] = f'▌{dir_name}（{len(dir_data)}个素材-出价组合）'
    ws2[f'A{row}'].font = Font(name='Arial', bold=True, size=12, color='1F4E79')
    for c in range(1, 9):
        ws2.cell(row=row, column=c).fill = sec_fill
    row += 1

    # Headers
    for i, h in enumerate(['序号','推荐等级','素材名称','月份','出价方式','DNU','CPI', '推荐理由'], 1):
        c = ws2.cell(row=row, column=i, value=h)
        c.font = hdr_f; c.fill = hdr_fill; c.alignment = ca; c.border = thin_b
    row += 1

    for idx, d in enumerate(dir_data, 1):
        tier_fill = PatternFill()
        if '🥇' in d['tier']: tier_fill = PatternFill('solid', fgColor='C6EFCE')
        elif '🥈' in d['tier']: tier_fill = PatternFill('solid', fgColor='BDD7EE')
        elif '🥉' in d['tier']: tier_fill = PatternFill('solid', fgColor='FCE4D6')
        else: tier_fill = PatternFill('solid', fgColor='F4B4C2')

        vals = [idx, d['tier'], d['name'], d['month'], d['bid'], d['dnu'], d['cpi'], d['reason']]
        for i, v in enumerate(vals, 1):
            c = ws2.cell(row=row, column=i, value=v)
            c.font = BLUE_FONT if i in [2,8] else nf
            c.alignment = la if i == 8 else ca
            c.border = thin_b
            c.fill = tier_fill
        ws2.row_dimensions[row].height = 24
        row += 1

    row += 1  # gap

# =====================================================
# PHASE 4: Rebuild Sheet 3 — direction summary (derived from Sheet 1)
# =====================================================
if '9月推荐-方向汇总' in wb.sheetnames:
    del wb['9月推荐-方向汇总']

ws3 = wb.create_sheet('9月推荐-方向汇总')

row = 1
ws3.merge_cells('A1:G1')
ws3['A1'] = '9月测试素材推荐 — 方向级汇总（由Sheet1统一推荐等级衍生）'
ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')

for i, w in enumerate([28, 10, 10, 10, 10, 55, 55], 1):
    ws3.column_dimensions[chr(64+i)].width = w

row = 3
for i, h in enumerate(['方向','强烈推荐','可选','待观察','不推荐','核心结论','素材清单(推荐等级+理由)'], 1):
    c = ws3.cell(row=row, column=i, value=h)
    c.font = hdr_f; c.fill = hdr_fill; c.alignment = ca; c.border = thin_b
ws3.row_dimensions[row].height = 30
row += 1

for dir_name in sorted_dirs:
    dir_data = direction_groups[dir_name]
    count_1 = sum(1 for d in dir_data if '🥇' in d['tier'])
    count_2 = sum(1 for d in dir_data if '🥈' in d['tier'])
    count_3 = sum(1 for d in dir_data if '🥉' in d['tier'])
    count_4 = sum(1 for d in dir_data if '❌' in d['tier'])

    # Pick top recommendations
    top = [d for d in dir_data if '🥇' in d['tier'] or '🥈' in d['tier']][:4]
    pick_list = '\n'.join([f"{d['tier'][:2]} {d['name']}({d['bid']},CPI${d['cpi']:.2f},R1={d['r1']:.0%},R3={d['r3']:.0%})" for d in top])

    # Generate conclusion
    if count_1 >= 2:
        conclusion = f'方向储备充足。{count_1}个强烈推荐素材可做主力量。'
    elif count_1 == 1:
        conclusion = f'有1个强烈推荐素材可作为方向核心。'
    elif count_2 >= 1:
        conclusion = f'无强烈推荐,{count_2}个可选素材需进一步验证。'
    else:
        conclusion = f'方向素材储备不足或数据不支撑推荐。'

    if count_4 > len(dir_data)*0.5:
        conclusion += f' {count_4}/{len(dir_data)}个素材不推荐,方向整体偏弱。'

    vals = [dir_name, count_1, count_2, count_3, count_4, conclusion, pick_list]
    for i, v in enumerate(vals, 1):
        c = ws3.cell(row=row, column=i, value=v)
        c.font = BLUE_FONT if i in [6,7] else nf
        c.alignment = la if i in [6,7] else ca
        c.border = thin_b
    ws3.row_dimensions[row].height = 80
    row += 1

# =====================================================
# Save
# =====================================================
wb.save('creative_final.xlsx')
print('\nDone! All changes marked in BLUE font.')
print(f'Sheet 1: Updated 备注 + added tier column (Col{tier_col})')
print('Sheet 2: Rebuilt from unified tiers')
print('Sheet 3: Rebuilt from unified tiers')
