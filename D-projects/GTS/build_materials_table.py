# -*- coding: utf-8 -*-
"""GTS 素材全量数据-按素材.xlsx 构建脚本
数据源：
  2月：飞书 Sheet HljIsZhGOhiccAt2vnpcJU3Kn4c / erKbRv（美国，Meta 激活口径，无留存）
  4月：复盘报告 KB6pd2KlVotXhZxnUMtc7O7GnBe §3.6（BI 重拆口径，权威）
       CTR/CVR 来自《GTS两次测试对比分析》docx（仅5支素材有）
       ctr*cvr 来自 Sheet JhDasabrthuYLht2ECwcpVXxnRb / CfgROC / wb9oM0
  7月：复盘报告 §3.1/§3.2/§5.1（BI 留存为准，UTC-4，菲律宾剔除）
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GREEN = PatternFill('solid', fgColor='C6EFCE')
YELLOW = PatternFill('solid', fgColor='FFEB9C')
RED = PatternFill('solid', fgColor='FFC7CE')
GREY = PatternFill('solid', fgColor='D9D9D9')
WHITE = PatternFill('solid', fgColor='FFFFFF')
HDR = PatternFill('solid', fgColor='4472C4')
THIN = Border(*[Side(style='thin', color='BFBFBF')]*4)

# 评分：Install / AEO 分别按 GTS KPI 基线
def score_row(bid, cpi, d1, d3):
    s = 0
    if cpi is not None:
        if bid == 'Install':
            s += 2 if cpi <= 1.50 else (1 if cpi <= 2.80 else 0)
        else:
            s += 2 if cpi <= 2.00 else (1 if cpi <= 2.60 else 0)
    if d1 is not None:
        if bid == 'Install':
            s += 2 if d1 >= 30 else (1 if d1 >= 25 else 0)
        else:
            s += 2 if d1 >= 35 else (1 if d1 >= 30 else 0)
    if d3 is not None:
        if bid == 'Install':
            s += 2 if d3 >= 13 else (1 if d3 >= 10 else 0)
        else:
            s += 2 if d3 >= 15 else (1 if d3 >= 12 else 0)
    return s

def grade(score, dnu, feb=False, installs=None):
    if feb:
        if installs is not None and installs < 20:
            return '⬜ 噪音', GREY
        return '⚪ 仅吸量参考', WHITE
    if dnu is not None and dnu < 20:
        return '⬜ 噪音', GREY
    if score >= 5:
        return '🥇 强烈推荐', GREEN
    if score >= 3:
        return '🥈 可选', YELLOW
    return '🔴 不推荐', RED

def pct(x):
    return None if x is None else f'{x:.2f}%'.rstrip('0').rstrip('.') + ('%' if not f'{x:.2f}%'.endswith('%') else '')

def p2(x):
    """xx.xx -> 'xx.xx%'，保留1-2位小数"""
    if x is None:
        return None
    return f'{x:.2f}'.rstrip('0').rstrip('.') + '%'

# 每条数据行: dict(name, direction, tags, month, bid, dnu, cpi, r1, r2, r3,
#                  d1_cost, spend, installs, imps, clicks, ctr, cvr, cc, note)
ROWS = []
def add(name, direction, tags, month, bid, dnu=None, cpi=None, r1=None, r2=None,
        r3=None, d1_cost=None, spend=None, installs=None, imps=None, clicks=None,
        ctr=None, cvr=None, cc=None, note=''):
    ROWS.append(dict(name=name, direction=direction, tags=tags, month=month, bid=bid,
                     dnu=dnu, cpi=cpi, r1=r1, r2=r2, r3=r3, d1_cost=d1_cost,
                     spend=spend, installs=installs, imps=imps, clicks=clicks,
                     ctr=ctr, cvr=cvr, cc=cc, note=note))

# ============ 7月 Install（复盘报告 §3.1，展示/点击由 CPM/CTR 反推） ============
JUL_INS = [
    # name, 花费, 安装, CPI, CTR%, CVR%, CPM, cc%, DNU, D1%, D3%, D1cost
    ('V-浴血黑帮',        703, 196, 3.58, 3.19, 17.85, 20.39, 0.5694, 169, 28.99, 9.5,  14.34),
    ('V-晋级失败被捕',     600, 202, 2.97, 3.34, 18.33, 18.17, 0.6122, 175, 24.6,  11.4, 13.95),
    ('P-披萨店',          260, 71,  3.65, 0.56, 14.55, 2.97,  0.0815, 64,  23.44, 4.7,  17.30),
    ('V-立绘展示',        218, 60,  3.63, 4.43, 18.75, 30.15, 0.8306, 57,  24.56, 3.5,  15.56),
    ('V-招募表演',        190, 48,  3.95, 3.73, 12.53, 18.48, 0.4674, 48,  39.58, 16.7, 9.99),
    ('V-鸡公大侠',        187, 56,  3.34, 4.70, 13.56, 21.29, 0.6373, 51,  27.50, 5.9,  13.36),
    ('V-模拟经营（原版）', 187, 51,  3.66, 6.03, 10.74, 23.71, 0.6476, 48,  12.50, 8.3,  31.13),
    ('P-炸鸡店',          173, 64,  2.70, 0.47, 18.99, 2.43,  0.0893, 59,  22.00, 10.2, 13.28),
    ('V-模拟经营（新版）', 127, 27,  4.69, 4.77, 10.15, 22.73, 0.4842, 26,  15.38, 7.7,  31.69),
    ('P-美漫分镜',        126, 48,  2.62, 2.07, 18.97, 10.26, 0.3927, 43,  25.58, 2.33, 11.42),
    ('V-爽感战斗',        122, 45,  2.70, 4.62, 20.09, 25.10, 0.9282, 35,  8.6,   8.6,  40.57),
    ('V-跟谁混',          104, 39,  2.66, 3.26, 21.55, 18.71, 0.7025, 37,  16.22, 10.8, 17.28),
    ('V-特殊设备视角',     102, 26,  3.93, 4.69, 14.86, 27.42, 0.6969, 23,  17.39, 8.7,  25.57),
]
for n, sp, ins, cpi, ctr, cvr, cpm, cc, dnu, d1, d3, cost in JUL_INS:
    imps = round(sp / cpm * 1000)
    clks = round(imps * ctr / 100)
    add(n, '', '', '7月', 'Install', dnu=dnu, cpi=cpi, r1=d1, r3=d3, d1_cost=cost,
        spend=sp, installs=ins, imps=imps, clicks=clks, ctr=ctr, cvr=cvr, cc=cc/100)

# 赌场博弈 Install（§5.1，精确展示/点击）
add('V-赌场博弈', '', '', '7月', 'Install', dnu=10, cpi=6.10, r1=0.0, r3=0.0,
    d1_cost=None, spend=55, installs=9, imps=2024, clicks=59, ctr=2.92, cvr=15.25,
    cc=None, note='')

# ============ 7月 AEO（§3.2） ============
JUL_AEO = [
    ('V-浴血黑帮',        517, 148, 3.49, 7.4,  17.39, 31.97, 0.9147, 142, 29.58, 9.2,  12.31),
    ('V-模拟经营（原版）', 465, 105, 4.43, 5.60, 12.98, 32.21, 0.7269, 95,  24.21, 7.4,  20.22),
    ('V-模拟经营（新版）', 465, 48,  9.68, 5.87, 10.2,  39.16, 0.4050, 41,  29.27, 9.8,  38.71),
    ('V-爽感战斗',        302, 70,  4.31, 6.47, 13.28, 37.09, 0.8592, 68,  27.94, 4.4,  15.90),
    ('V-招募表演',        265, 73,  3.63, 4.58, 14.72, 24.48, 0.6742, 72,  29.17, 13.9, 12.62),
    ('V-特殊设备视角',     200, 48,  4.16, 7.56, 11.43, 35.93, 0.8641, 48,  22.92, 18.75, 18.14),
    ('V-鸡公大侠',        187, 60,  3.12, 6.74, 16.62, 35.01, 1.1202, 58,  25.86, 6.9,  12.50),
    ('V-立绘展示',        139, 48,  2.90, 5.52, 24.37, 39.03, 1.3452, 42,  19.05, 2.4,  17.41),
]
for n, sp, ins, cpi, ctr, cvr, cpm, cc, dnu, d1, d3, cost in JUL_AEO:
    imps = round(sp / cpm * 1000)
    clks = round(imps * ctr / 100)
    add(n, '', '', '7月', 'AEO', dnu=dnu, cpi=cpi, r1=d1, r3=d3, d1_cost=cost,
        spend=sp, installs=ins, imps=imps, clicks=clks, ctr=ctr, cvr=cvr, cc=cc/100)

add('V-赌场博弈', '', '', '7月', 'AEO', dnu=13, cpi=5.58, r1=0.0, r3=0.0,
    d1_cost=None, spend=78, installs=14, imps=1909, clicks=87, ctr=4.56, cvr=16.09,
    cc=None, note='')

# ============ 4月（复盘报告 §3.6，BI 重拆口径） ============
# name, 花费, 安装(媒体), CPI, BI DNU, BI D1%, BI D3%, D1cost
APR_INS = [
    ('V-模拟经营（原版）', 118, 63,  1.87, 51, 23.5, 3.9,  9.84),
    ('V-场景展示',        93,  76,  1.22, 56, 10.7, 8.9,  15.46),
    ('V-晋级失败被捕',     165, 106, 1.56, 89, 23.6, 7.9,  7.86),
    ('V-爽感战斗',        145, 91,  1.59, 77, 18.2, 6.5,  10.33),
    ('V-鸡公大侠',        109, 71,  1.54, 59, 22.0, 3.4,  8.38),
    ('V-浴血黑帮',        55,  48,  1.14, 37, 32.4, 10.8, 4.55),
    ('V-超速被抓',        121, 81,  1.49, 63, 11.1, 3.2,  17.25),
    ('V-升级变装',        86,  40,  2.15, 33, 18.2, 15.2, 14.36),
    ('V-战斗形式',        35,  12,  2.89, 9,  22.2, 0.0,  17.32),
    ('V-越狱混剪',        16,  6,   2.68, 8,  12.5, 0.0,  16.05),
]
APR_INS_CTR_CVR = {  # 来自《两次测试对比分析》docx，仅这5支有
    'P-炸鸡店': (1.3, 25.5), 'V-模拟经营（原版）': (3.16, 31.34),
    'V-晋级失败被捕': (3.92, 27.75), 'V-战斗形式': (3.43, 24.49), 'V-升级变装': (4.72, 21.62),
}
APR_INS_CC = {  # CfgROC 列N（%）
    'V-晋级失败被捕': 1.09, 'V-爽感战斗': 1.38, 'V-超速被抓': 1.05, 'V-场景展示': 1.61,
    'V-鸡公大侠': 1.46, 'V-模拟经营（原版）': 0.99, 'V-浴血黑帮': 1.44, 'P-炸鸡店': 0.33,
    'V-升级变装': 1.02, 'V-战斗形式': 0.84, 'V-越狱混剪': 0.73,
}
for n, sp, ins, cpi, dnu, d1, d3, cost in APR_INS:
    ctr, cvr = APR_INS_CTR_CVR.get(n, (None, None))
    cc = APR_INS_CC.get(n)
    add(n, '', '', '4月', 'Install', dnu=dnu, cpi=cpi, r1=d1, r3=d3, d1_cost=cost,
        spend=sp, installs=ins, ctr=ctr, cvr=cvr, cc=cc/100 if cc else None)

APR_AEO = [
    ('V-模拟经营（原版）', 297, 137, 2.17, 117, 35.9, 12.0, 7.07),
    ('V-场景展示',        91,  45,  2.01, 40,  25.0, 12.5, 9.05),
    ('V-爽感战斗',        139, 63,  2.21, 54,  29.6, 7.4,  8.71),
    ('V-浴血黑帮',        34,  24,  1.40, 20,  45.0, 10.0, 3.74),
    ('V-超速被抓',        110, 48,  2.29, 47,  10.6, 10.6, 21.98),
    ('V-鸡公大侠',        39,  17,  2.28, 15,  26.7, 0.0,  9.68),
    ('V-升级变装',        28,  7,   3.97, 7,   28.6, 14.3, 13.91),
    ('P-炸鸡店',          61,  21,  2.88, 19,  21.1, 5.3,  15.13),
    ('V-晋级失败被捕',     31,  9,   3.44, 7,   14.3, 28.6, 21.69),
    ('P-角色展示',        9,   1,   9.26, 1,   None, None, None),
]
APR_AEO_CC = {  # wb9oM0 列N（%）
    'V-模拟经营（原版）': 1.22, 'V-爽感战斗': 1.46, 'V-超速被抓': 0.80, 'V-场景展示': 1.32,
    'P-炸鸡店': 0.35, 'V-浴血黑帮': 1.71, 'V-鸡公大侠': 1.53, 'V-晋级失败被捕': 0.63,
    'V-升级变装': 0.64, 'P-角色展示': 0.11,
}
for n, sp, ins, cpi, dnu, d1, d3, cost in APR_AEO:
    cc = APR_AEO_CC.get(n)
    add(n, '', '', '4月', 'AEO', dnu=dnu, cpi=cpi, r1=d1, r3=d3, d1_cost=cost,
        spend=sp, installs=ins, cc=cc/100 if cc else None)

# ============ 2月（Sheet erKbRv，Meta 激活口径，无留存） ============
# name, 花费, 激活, 激活单价, cc(%) CTR%, CVR%, CPM, 点击, 展示
FEB = [
    ('P-角色展示',       206.76, 120, 1.72, 0.62, 2.54, 24.39, 10.68, 492, 19366),
    ('V-晋级失败被捕',    183.34, 116, 1.58, 1.06, 4.63, 22.79, 16.66, 509, 11005),
    ('V-跟谁混',          183.98, 101, 1.82, 1.25, 5.37, 23.27, 22.76, 434, 8083),
    ('V-战斗形式',        251.18, 95,  2.64, 1.09, 7.02, 15.52, 28.80, 612, 8723),
    ('P-主角展示',        115.84, 65,  1.78, 0.65, 2.87, 22.81, 11.66, 285, 9935),
    ('V-追逐战',          108.36, 54,  2.01, 0.92, 4.39, 21.01, 18.52, 257, 5852),
    ('V-diy人物&抢劫',    38.07,  19,  2.00, 0.82, 5.13, 15.97, 16.42, 119, 2318),
    ('V-升级变装',        49.04,  18,  2.72, 0.90, 5.39, 16.67, 24.47, 108, 2004),
    ('V-越狱混剪',        12.40,  4,   3.10, 0.60, 6.34, 9.52,  18.73, 42,  662),
    ('V-警匪追击',        4.26,   0,   None, 0.00, 5.45, 0.00,  21.09, 11,  202),
    ('V-模拟经营（原版）', 52.81,  36,  1.47, 1.37, 4.79, 28.57, 20.10, 126, 2628),
    ('P-炸鸡店',          54.81,  40,  1.37, 0.88, 3.55, 24.84, 12.09, 161, 4533),
]
for n, sp, ins, cpi, cc, ctr, cvr, cpm, clks, imps in FEB:
    add(n, '', '', '2月', 'Install', cpi=cpi, spend=sp, installs=ins, imps=imps,
        clicks=clks, ctr=ctr, cvr=cvr, cc=cc/100)

# ============ 素材元信息（方向归类/内容标签） ============
META = {
    'V-浴血黑帮':       ('氛围混剪', '复古黑帮, 世界观混剪, 电影致敬'),
    'V-模拟经营（原版）': ('模拟经营', '产业经营, 荒诞犯罪循环'),
    'V-模拟经营（新版）': ('模拟经营', '产业经营, 玩法闭环对齐'),
    'V-场景展示':       ('模拟经营', '地盘建设, 场景展示'),
    'P-炸鸡店':         ('模拟经营/幽默反差', '产业经营, 认知反差'),
    'P-披萨店':         ('模拟经营/幽默反差', '产业经营, 认知反差'),
    'V-晋级失败被捕':    ('复仇逆袭', '命运反转, 权力欲, 互动选择'),
    'V-越狱混剪':       ('复仇逆袭', '越狱突围, 反抗逆袭'),
    'V-爽感战斗':       ('战斗', 'AOE爽感, 特效包装'),
    'V-战斗形式':       ('战斗', '裸录屏战斗'),
    'V-特殊设备视角':    ('战斗', '无人机视角, 战斗演出'),
    'V-鸡公大侠':       ('幽默反差', '恶搞战斗, 打鸡升级'),
    'V-diy人物&抢劫':   ('幽默反差', '搞笑道具, 荒诞元素'),
    'V-跟谁混':         ('小弟招募', '帮派选择, 权力博弈'),
    'V-招募表演':       ('小弟招募', '招募仪式感, AI制作'),
    'P-美漫分镜':       ('美术风格', '美式漫画, 分镜'),
    'P-主角展示':       ('美术风格', '赛博朋克, 霓虹犯罪'),
    'V-立绘展示':       ('角色展示', '立绘, AI制作'),
    'P-角色展示':       ('角色展示', '动漫风, 等级展示'),
    'V-升级变装':       ('角色养成', '升级变装'),
    'V-赌场博弈':       ('休闲副玩法', '赌场小游戏'),
    'V-追逐战':         ('休闲副玩法', '赛车追车'),
    'V-超速被抓':       ('街头任务', '警匪追捕, 货不对板'),
    'V-警匪追击':       ('街头任务', '超现实追车'),
}
# 方向排序优先级
DIR_ORDER = ['氛围混剪', '模拟经营', '模拟经营/幽默反差', '幽默反差', '复仇逆袭',
             '小弟招募', '战斗', '美术风格', '角色展示', '角色养成', '休闲副玩法', '街头任务']
MONTH_ORDER = {'7月': 0, '4月': 1, '2月': 2}
BID_ORDER = {'Install': 0, 'AEO': 1}

# ============ 备注生成 ============
def cpi_word(bid, cpi):
    if bid == 'Install':
        return '优秀' if cpi <= 1.50 else ('合格' if cpi <= 2.80 else '超标')
    return '优秀' if cpi <= 2.00 else ('合格' if cpi <= 2.60 else '超标')

def d1_word(bid, d1):
    if bid == 'Install':
        return '优秀' if d1 >= 30 else ('合格' if d1 >= 25 else '待改进')
    return '优秀' if d1 >= 35 else ('合格' if d1 >= 30 else '待改进')

def d3_word(bid, d3):
    if bid == 'Install':
        return '优秀' if d3 >= 13 else ('合格' if d3 >= 10 else '待改进')
    return '优秀' if d3 >= 15 else ('合格' if d3 >= 12 else '待改进')

def dnu_word(dnu):
    if dnu is None: return ''
    if dnu < 20: return f'DNU={dnu}样本不足'
    if dnu < 45: return f'DNU={dnu}偏小'
    return f'DNU={dnu}可靠'

NOTES = {  # 特殊备注（key: (name, month, bid)）
    ('V-浴血黑帮', '7月', 'Install'): 'CPI超标系7月大盘普涨；D1 28.99%双线领先，CBO下获量能力验证，Install核心主力',
    ('V-浴血黑帮', '7月', 'AEO'): 'D1 29.58%为AEO最高之一，次留成本$12.31全线最优；较4月45%回落15pp但仍居首',
    ('V-浴血黑帮', '4月', 'Install'): 'CPI $1.14全场最低+D1 32.4%最高；DNU=37样本偏小→7月扩量验证',
    ('V-浴血黑帮', '4月', 'AEO'): 'CPI $1.40最低+D1 45%最高；DNU=20样本小，次留成本$3.74全场最优',
    ('V-招募表演', '7月', 'Install'): 'D1 39.58%全场最高，次留成本$9.99最优；新方向最强信号',
    ('V-招募表演', '7月', 'AEO'): 'D1 29.17%稳定，双线均达标方向',
    ('V-模拟经营（原版）', '7月', 'Install'): 'D1仅12.50%严重翻车（4月23.5%）；AEO重开补测后24.21%，方向需重新评估',
    ('V-模拟经营（原版）', '7月', 'AEO'): 'D1 24.21%较4月35.9%大幅下滑；4月标杆方向本轮不复现',
    ('V-模拟经营（原版）', '4月', 'AEO'): '4月唯一摸到行业线的方向：D1 35.9%+量最大（DNU=117）',
    ('V-模拟经营（新版）', '7月', 'AEO'): 'CPI $9.68为AEO灾难级，CVR仅10.2%；同公式新做未复现原版',
    ('V-爽感战斗', '7月', 'Install'): 'D1 8.6%严重异常（AEO线27.94%正常），疑似投放/回传问题，见复盘第5章',
    ('V-赌场博弈', '7月', 'Install'): '双线D1=0%（23个DNU无一留存），CPI超$5.5；方向与游戏内容完全不匹配，已剔除汇总',
    ('V-赌场博弈', '7月', 'AEO'): '同上，双线关停，方向淘汰',
    ('P-美漫分镜', '7月', 'Install'): '新方向图片素材，CPI $2.62图片中合格，D1 25.58%擦线；D3仅2.33%偏弱',
    ('V-立绘展示', '7月', 'AEO'): 'CPI $2.90 AEO最低+CVR 24.37%最高，但D1仅19.05%——高转化不等于高留存',
    ('V-特殊设备视角', '7月', 'AEO'): 'D3 18.75%为全场最高，建议下轮确认',
    ('P-炸鸡店', '4月', 'Install'): '图片素材天花板：D1 30.77%级别表现，CPI稳定低位',
    ('P-炸鸡店', '4月', 'AEO'): '图片AEO次留21.1% vs Install 30.77%——图片不适合AEO出价',
    ('V-晋级失败被捕', '4月', 'AEO'): 'AEO D1 14.3%远低于Install 23.6%，AEO无增益，仅投Install',
    ('V-超速被抓', '4月', 'Install'): '货不对板，D1 11.1%量够可定论，淘汰',
    ('V-战斗形式', '4月', 'Install'): '裸录屏接不住点击，淘汰',
    ('V-警匪追击', '2月', 'Install'): 'CVR 0%、激活0，超现实方向直接淘汰',
}

def auto_note(r):
    key = (r['name'], r['month'], r['bid'])
    if key in NOTES:
        return NOTES[key]
    parts = []
    if r['month'] == '2月':
        parts.append('2月仅测吸量，无留存数据')
        if r['cpi'] is not None:
            parts.append(f"CPI ${r['cpi']:.2f}{cpi_word('Install', r['cpi'])}")
        return '；'.join(parts)
    if r['cpi'] is not None:
        parts.append(f"CPI ${r['cpi']:.2f}{cpi_word(r['bid'], r['cpi'])}")
    if r['r1'] is not None:
        parts.append(f"D1 {r['r1']:.1f}%{d1_word(r['bid'], r['r1'])}")
    if r['r3'] is not None:
        parts.append(f"D3 {r['r3']:.1f}%{d3_word(r['bid'], r['r3'])}")
    dw = dnu_word(r['dnu'])
    if dw: parts.append(dw)
    parts.append(f"{r['month']}数据")
    return '；'.join(parts)

# ============ 组装：按素材分块 ============
materials = {}
for r in ROWS:
    materials.setdefault(r['name'], []).append(r)

def sort_key(item):
    name, rows = item
    d = META.get(name, ('zz', ''))[0]
    di = DIR_ORDER.index(d) if d in DIR_ORDER else 99
    return (di, name)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '素材全量数据-按素材'

ws['A1'] = 'GTS素材全量数据 — 以素材为主维度，月份+出价方式为下钻（标色逻辑：绿=推荐 黄=可选 红=不推荐 灰=噪音/样本不足）'
ws['A2'] = ('🟢绿(≥5分) | 🟡黄(3-4分) | 🔴红(<3分) | ⬜灰 DNU<20 | ⚪白 2月仅吸量无留存 | '
            '评分= CPI/D1/D3各0-2分（Install: ≤$1.5/≥30%/≥13%满分；AEO: ≤$2.0/≥35%/≥15%满分） | '
            '排序：方向优先级 → 素材名 → 7月→4月→2月 → Install→AEO | '
            '数据源：7月&4月=复盘报告(BI留存,UTC-4,剔除菲律宾)；2月=Meta激活口径；4月CTR/CVR仅部分素材有')
ws['A1'].font = Font(bold=True, size=12)
ws['A2'].font = Font(size=9, color='555555')

headers = ['素材名称', '方向归类', '内容标签', '月份', '出价方式', 'DNU', 'CPI', 'R1', 'R2', 'R3',
           '次留成本', '花费(USD)', '安装数', '展示次数', '点击量', 'CTR', 'CVR', 'CTR*CVR', '备注', None, '推荐等级']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = HDR
    cell.border = THIN
    cell.alignment = Alignment(horizontal='center', vertical='center')

row_i = 4
for name, rows in sorted(materials.items(), key=sort_key):
    direction, tags = META.get(name, ('', ''))
    rows.sort(key=lambda r: (MONTH_ORDER[r['month']], BID_ORDER[r['bid']]))
    start = row_i
    for r in rows:
        feb = r['month'] == '2月'
        sc = score_row(r['bid'], r['cpi'], r['r1'], r['r3']) if not feb else 0
        g, fill = grade(sc, r['dnu'], feb=feb, installs=r['installs'])
        if r['name'] == 'V-赌场博弈':  # 用户确认：双线D1=0%异常，强制标红
            g, fill = '🔴 不推荐', RED
        vals = [name, direction, tags, r['month'], r['bid'], r['dnu'],
                r['cpi'], p2(r['r1']), '—' if r['r2'] is None else p2(r['r2']), p2(r['r3']),
                r['d1_cost'], f"${r['spend']:,.0f}" if r['spend'] else None, r['installs'],
                r['imps'], r['clicks'], p2(r['ctr']), p2(r['cvr']), r['cc'],
                auto_note(r), None, g]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=c, value=v)
            cell.border = THIN
            cell.fill = fill
            cell.alignment = Alignment(vertical='center', wrap_text=(c in (19,)))
        ws.cell(row=row_i, column=7).number_format = '$#,##0.00'
        ws.cell(row=row_i, column=11).number_format = '$#,##0.00'
        ws.cell(row=row_i, column=13).number_format = '#,##0'
        ws.cell(row=row_i, column=14).number_format = '#,##0'
        ws.cell(row=row_i, column=15).number_format = '#,##0'
        ws.cell(row=row_i, column=18).number_format = '0.0000%'
        row_i += 1
    # 合并素材名称/方向/标签
    if row_i - 1 > start:
        for c in (1, 2, 3):
            ws.merge_cells(start_row=start, start_column=c, end_row=row_i - 1, end_column=c)
    row_i += 1  # 空行分隔

widths = [18, 14, 22, 6, 9, 7, 8, 8, 6, 8, 9, 10, 8, 9, 8, 8, 8, 10, 52, 4, 12]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w
ws.freeze_panes = 'A4'

out = r'D:\claude-projects\projects\GTS\GTS素材全量数据-按素材.xlsx'
wb.save(out)
print('saved:', out, 'rows:', row_i - 1)

# 自检：打印各月行数
from collections import Counter
cnt = Counter((r['month'], r['bid']) for r in ROWS)
print(cnt)
